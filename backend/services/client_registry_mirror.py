from __future__ import annotations

import datetime as dt
import io
import os

from googleapiclient.http import MediaIoBaseUpload
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import joinedload

from database import Client, SessionLocal
from services.client_promotorias import valid_promotoria


DEFAULT_CLIENT_REGISTRY_FOLDER_ID = "1z8xgaGU9ZRnLy1bZkGBbXZ_C7fjS3eJZ"
CLIENT_REGISTRY_FILENAME = "Clientes - Registro Maestro.xlsx"


def _folder_id() -> str:
    return os.getenv("GOOGLE_DRIVE_CLIENT_REGISTRY_FOLDER_ID", DEFAULT_CLIENT_REGISTRY_FOLDER_ID).strip()


def _build_writable_drive_service():
    from google.auth import default
    from googleapiclient.discovery import build

    credentials, _ = default(scopes=["https://www.googleapis.com/auth/drive"])
    return build("drive", "v3", credentials=credentials, cache_discovery=False)


def _workbook_bytes(clients: list[Client]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Clientes"
    headers = [
        "Nombre", "RFC", "Correo", "Telefono", "Promotorias", "Estado identidad",
        "Estatus", "Expediente", "Nombre expediente", "Ultima actualizacion",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17365D")
    for client in clients:
        sheet.append([
            client.full_name,
            client.rfc or "",
            client.email or "",
            client.phone or "",
            ", ".join(sorted({
                row.promotoria for row in client.promotorias if valid_promotoria(row.promotoria)
            })),
            client.identity_status,
            client.status,
            client.drive_folder_url or "",
            client.drive_folder_name or "",
            client.updated_at.isoformat(timespec="seconds") if client.updated_at else "",
        ])
    widths = [42, 18, 36, 18, 36, 20, 14, 65, 48, 22]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    metadata = workbook.create_sheet("Control")
    metadata.append(["Campo", "Valor"])
    metadata.append(["Generado UTC", dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds")])
    metadata.append(["Registros activos", len(clients)])
    metadata.append(["Fuente maestra", "TAIICO CRM / tabla clients"])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def sync_client_registry_mirror() -> dict:
    db = SessionLocal()
    try:
        clients = (
            db.query(Client)
            .options(joinedload(Client.promotorias))
            .filter(Client.status != "inactive")
            .order_by(Client.full_name)
            .all()
        )
        content = _workbook_bytes(clients)
    finally:
        db.close()

    service = _build_writable_drive_service()
    escaped_name = CLIENT_REGISTRY_FILENAME.replace("'", "\\'")
    matches = service.files().list(
        q=f"'{_folder_id()}' in parents and name='{escaped_name}' and trashed=false",
        fields="files(id,name,webViewLink)",
        pageSize=10,
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute().get("files", [])
    media = MediaIoBaseUpload(
        io.BytesIO(content),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False,
    )
    if matches:
        saved = service.files().update(
            fileId=matches[0]["id"],
            media_body=media,
            fields="id,name,webViewLink,modifiedTime",
            supportsAllDrives=True,
        ).execute()
        action = "updated"
    else:
        saved = service.files().create(
            body={"name": CLIENT_REGISTRY_FILENAME, "parents": [_folder_id()]},
            media_body=media,
            fields="id,name,webViewLink,modifiedTime",
            supportsAllDrives=True,
        ).execute()
        action = "created"
    return {"action": action, "clients": len(clients), **saved}


def sync_client_registry_mirror_best_effort() -> None:
    try:
        sync_client_registry_mirror()
    except Exception as exc:
        print(f"Client registry mirror unavailable: {type(exc).__name__}: {exc}")
