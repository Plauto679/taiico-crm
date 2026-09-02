from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import re
import shutil
import tempfile
import threading
import unicodedata
import uuid
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from xml.etree import ElementTree

from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
from sqlalchemy import asc, desc, func, or_

from config import FINANCE_INVOICES_PATH, FINANCE_SOURCE_FILE_IDS, FINANCE_SOURCE_PATHS
from database import (
    FinanceClassificationRule,
    FinanceBudgetItem,
    FinanceIngestion,
    FinanceInvoice,
    FinanceInvoiceMatch,
    FinanceMovement,
    FinanceProjection,
    FinanceRecurringDecision,
    FinanceRuleApplication,
    FinanceSourceState,
    SessionLocal,
)
from services.auth import AccessProfile
from services.authorization import require_module_access
from drive.client import download_drive_file_bytes


router = APIRouter(prefix="/finanzas", tags=["finanzas"])
_sync_lock = threading.RLock()
CANONICAL_COLUMNS = (
    "id_movimiento", "empresa", "banco", "tipo_cuenta", "naturaleza_cuenta",
    "cuenta", "clabe", "moneda", "fecha_operacion", "fecha_liquidacion",
    "descripcion_original", "referencia", "contraparte", "cargo", "abono",
    "importe_neto", "saldo", "titular", "categoria", "subcategoria",
    "recurrente", "impuesto", "nomina", "requiere_factura", "factura_uuid",
    "estatus_conciliacion_factura", "estatus_revision", "periodo_estado",
    "archivo_fuente", "pagina_fuente", "hash_fuente",
)
SOURCE_META = {
    "tla_amex": ("TLA", "AMEX"),
    "tla_bbva": ("TLA", "BBVA"),
    "tla_banorte": ("TLA", "BANORTE"),
    "ts_bbva": ("TS", "BBVA"),
}
BANK_ALIASES = {
    "AMEX": {"AMEX", "AMERICAN EXPRESS"},
}
BOOL_TRUE = {"1", "true", "si", "sí", "yes", "x"}
RUNTIME_DIR = Path(__file__).resolve().parents[2] / ".runtime" / "finanzas"


class MovementPatch(BaseModel):
    categoria: str | None = Field(default=None, max_length=255)
    subcategoria: str | None = Field(default=None, max_length=255)
    recurrente: bool | None = None
    impuesto: bool | None = None
    nomina: bool | None = None
    requiere_factura: bool | None = None
    estatus_revision: str | None = Field(default=None, max_length=50)


class BulkClassification(BaseModel):
    movement_ids: list[str] = Field(min_length=1, max_length=1000)
    categoria: str = Field(min_length=1, max_length=255)
    subcategoria: str | None = Field(default=None, max_length=255)


class RecurringInput(BaseModel):
    status: str = Field(pattern="^(confirmado|descartado|pendiente)$")
    note: str | None = Field(default=None, max_length=2000)


class ProjectionInput(BaseModel):
    company: str = Field(pattern="^(TLA|TS)$")
    due_date: date
    concept: str = Field(min_length=1, max_length=500)
    amount: Decimal
    scenario: str = Field(default="base", pattern="^(base|optimista|estres)$")


class RuleInput(BaseModel):
    name: str = Field(min_length=1, max_length=255)
    priority: int = Field(default=100, ge=1, le=9999)
    field: str = Field(pattern="^(descripcion_original|contraparte|referencia|banco)$")
    operator: str = Field(pattern="^(contiene|igual|empieza_con|regex)$")
    value: str = Field(min_length=1, max_length=500)
    company: str | None = Field(default=None, pattern="^(TLA|TS)$")
    category: str = Field(min_length=1, max_length=255)
    subcategory: str | None = Field(default=None, max_length=255)
    enabled: bool = True
    exclusion: bool = False


class InvoiceMatchInput(BaseModel):
    movement_id: str
    status: str = Field(default="confirmado", pattern="^(confirmado|rechazado)$")


class BudgetInput(BaseModel):
    company: str = Field(pattern="^(TLA|TS)$")
    month: date
    category: str = Field(min_length=1, max_length=255)
    amount: Decimal = Field(ge=0)


def _text(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def _canonical_bank(value: object) -> str:
    bank = _text(value).upper()
    for canonical, aliases in BANK_ALIASES.items():
        if bank in aliases:
            return canonical
    return bank


def _decimal(value: object, *, default: Decimal | None = Decimal("0")) -> Decimal | None:
    raw = str(value or "").strip().replace(",", "").replace("$", "")
    if not raw:
        return default
    if raw.startswith("(") and raw.endswith(")"):
        raw = f"-{raw[1:-1]}"
    try:
        return Decimal(raw)
    except InvalidOperation as exc:
        raise ValueError(f"Importe inválido: {value}") from exc


def _date(value: object, *, required: bool = False) -> date | None:
    raw = _text(value)
    if not raw:
        if required:
            raise ValueError("Fecha de operación vacía")
        return None
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y", "%Y%m%d"):
        try:
            return datetime.strptime(raw[:10], pattern).date()
        except ValueError:
            continue
    raise ValueError(f"Fecha inválida: {raw}")


def _bool(value: object) -> bool:
    normalized = unicodedata.normalize("NFKD", _text(value)).encode("ascii", "ignore").decode().casefold()
    return normalized in BOOL_TRUE


def _sha256(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _source_state(db, key: str) -> FinanceSourceState:
    company, bank = SOURCE_META[key]
    path = FINANCE_SOURCE_PATHS[key]
    source_path = str(path) if path.is_file() else f"google-drive:{key}"
    state = db.get(FinanceSourceState, key)
    if not state:
        state = FinanceSourceState(key=key, company=company, bank=bank, source_path=source_path)
        db.add(state)
        db.flush()
    else:
        state.source_path = source_path
    return state


def _read_source(key: str) -> tuple[bytes, datetime | None]:
    path = FINANCE_SOURCE_PATHS[key]
    if path.is_file():
        return path.read_bytes(), datetime.fromtimestamp(path.stat().st_mtime)
    file_id = _text(FINANCE_SOURCE_FILE_IDS.get(key))
    if file_id:
        return download_drive_file_bytes(file_id), None
    raise FileNotFoundError("La fuente no está montada ni configurada en Google Drive")


def _validate_headers(headers: list[str] | None) -> None:
    actual = tuple(headers or ())
    missing = [column for column in CANONICAL_COLUMNS if column not in actual]
    if missing:
        raise ValueError("Faltan columnas canónicas: " + ", ".join(missing))


def _decode_csv(content: bytes) -> str:
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("No se pudo identificar la codificación del CSV")


def _csv_reader(content: bytes) -> csv.DictReader:
    decoded = _decode_csv(content)
    try:
        dialect = csv.Sniffer().sniff(decoded[:4096], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    return csv.DictReader(io.StringIO(decoded), dialect=dialect)


def _header_key(value: object) -> str:
    normalized = unicodedata.normalize("NFKD", _text(value)).encode("ascii", "ignore").decode().casefold()
    return re.sub(r"[^a-z0-9]+", "_", normalized).strip("_")


def _amex_date(value: object, *, required: bool = False) -> date | None:
    raw = _text(value)
    if not raw:
        if required:
            raise ValueError("Fecha de operación vacía")
        return None
    month_aliases = {"ene": "jan", "abr": "apr", "ago": "aug", "dic": "dec", "sept": "sep"}
    normalized = unicodedata.normalize("NFKD", raw).encode("ascii", "ignore").decode().casefold().replace(".", "")
    for spanish, english in month_aliases.items():
        normalized = re.sub(rf"\b{spanish}\b", english, normalized)
    for pattern in ("%d %b %Y", "%d-%b-%y", "%d-%b-%Y", "%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(normalized, pattern).date()
        except ValueError:
            continue
    raise ValueError(f"Fecha AMEX inválida: {raw}")


def _amex_account(value: object) -> str:
    account = _text(value)
    if re.fullmatch(r"-\d{4}", account):
        return f"-0{account[1:]}"
    return account


def _amex_source_filename(filename: str) -> str:
    return f"TLA/Estados Mensuales Amex/{Path(filename or 'estado.csv').name}"


def parse_amex_monthly_csv(content: bytes, *, filename: str = "estado.csv") -> list[dict[str, object]]:
    reader = _csv_reader(content)
    headers = list(reader.fieldnames or ())
    keyed_headers = {_header_key(header): header for header in headers}
    description_header = next((header for header in headers if _header_key(header).startswith("descripci")), None)
    aliases = {
        "operation_date": keyed_headers.get("fecha"),
        "settlement_date": keyed_headers.get("fecha_de_compra") or keyed_headers.get("fecha_compra"),
        "description": keyed_headers.get("descripcion") or description_header,
        "holder": keyed_headers.get("titular_de_la_tarjeta") or keyed_headers.get("titular_tarjeta"),
        "account": keyed_headers.get("cuenta"),
        "amount": keyed_headers.get("importe"),
    }
    if any(not header for header in aliases.values()):
        raise ValueError(
            "El CSV mensual de AMEX debe incluir las columnas: Fecha, Fecha de Compra, "
            "Descripción, Titular de la Tarjeta, Cuenta e Importe"
        )

    rows: list[dict[str, object]] = []
    occurrences: Counter[str] = Counter()
    source_filename = _amex_source_filename(filename)
    source_hash = _sha256(content)
    for line, source in enumerate(reader, start=2):
        if not any(_text(value) for value in source.values()):
            continue
        try:
            operation_date = _amex_date(source.get(aliases["operation_date"]), required=True)
            settlement_date = _amex_date(source.get(aliases["settlement_date"]))
            amount = _decimal(source.get(aliases["amount"]), default=None)
            if amount is None:
                raise ValueError("Importe vacío")
        except ValueError as exc:
            raise ValueError(f"Fila {line}: {exc}") from exc
        description = _text(source.get(aliases["description"]))
        holder = _text(source.get(aliases["holder"]))
        account = _amex_account(source.get(aliases["account"]))
        fingerprint = json.dumps({
            "fecha": operation_date.isoformat(),
            "fecha_compra": settlement_date.isoformat() if settlement_date else "",
            "descripcion": description,
            "titular": holder,
            "cuenta": account,
            "importe": str(amount),
        }, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        occurrences[fingerprint] += 1
        row_hash = _sha256(f"{fingerprint}|{occurrences[fingerprint]}".encode())
        debit = amount if amount > 0 else Decimal("0")
        credit = -amount if amount < 0 else Decimal("0")
        rows.append({
            "external_id": row_hash[:24], "company": "TLA", "bank": "AMEX",
            "account_type": "Tarjeta", "account_nature": "Crédito", "account": account,
            "clabe": "", "currency": "MXN", "operation_date": operation_date,
            "settlement_date": settlement_date, "original_description": description,
            "reference": "", "counterparty": "", "debit": debit, "credit": credit,
            "net_amount": -amount, "balance": None, "holder": holder,
            "source_category": "", "source_subcategory": "", "recurring": False,
            "tax": False, "payroll": False, "requires_invoice": False,
            "invoice_uuid": "", "invoice_reconciliation_status": "", "review_status": "",
            "statement_period": operation_date.strftime("%Y-%m"),
            "source_filename": source_filename, "source_page": None, "source_hash": source_hash,
        })
    return rows


def parse_canonical_csv(content: bytes) -> list[dict[str, object]]:
    reader = _csv_reader(content)
    _validate_headers(reader.fieldnames)
    rows: list[dict[str, object]] = []
    seen: set[str] = set()
    for line, source in enumerate(reader, start=2):
        external_id = _text(source.get("id_movimiento"))
        if not external_id:
            raise ValueError(f"Fila {line}: id_movimiento vacío")
        if external_id in seen:
            raise ValueError(f"Fila {line}: id_movimiento duplicado ({external_id})")
        seen.add(external_id)
        debit = _decimal(source.get("cargo")) or Decimal("0")
        credit = _decimal(source.get("abono")) or Decimal("0")
        net = _decimal(source.get("importe_neto"), default=None)
        if net is None:
            net = credit - debit
        rows.append({
            "external_id": external_id,
            "company": _text(source.get("empresa")).upper(),
            "bank": _canonical_bank(source.get("banco")),
            "account_type": _text(source.get("tipo_cuenta")),
            "account_nature": _text(source.get("naturaleza_cuenta")),
            "account": _text(source.get("cuenta")),
            "clabe": _text(source.get("clabe")),
            "currency": _text(source.get("moneda")) or "MXN",
            "operation_date": _date(source.get("fecha_operacion"), required=True),
            "settlement_date": _date(source.get("fecha_liquidacion")),
            "original_description": _text(source.get("descripcion_original")),
            "reference": _text(source.get("referencia")),
            "counterparty": _text(source.get("contraparte")),
            "debit": debit,
            "credit": credit,
            "net_amount": net,
            "balance": _decimal(source.get("saldo"), default=None),
            "holder": _text(source.get("titular")),
            "source_category": _text(source.get("categoria")),
            "source_subcategory": _text(source.get("subcategoria")),
            "recurring": _bool(source.get("recurrente")),
            "tax": _bool(source.get("impuesto")),
            "payroll": _bool(source.get("nomina")),
            "requires_invoice": _bool(source.get("requiere_factura")),
            "invoice_uuid": _text(source.get("factura_uuid")),
            "invoice_reconciliation_status": _text(source.get("estatus_conciliacion_factura")),
            "review_status": _text(source.get("estatus_revision")),
            "statement_period": _text(source.get("periodo_estado")),
            "source_filename": _text(source.get("archivo_fuente")),
            "source_page": int(_decimal(source.get("pagina_fuente")) or 0) or None,
            "source_hash": _text(source.get("hash_fuente")),
        })
    return rows


def parse_ingestion_csv(source_key: str, content: bytes, *, filename: str = "estado.csv") -> list[dict[str, object]]:
    headers = list(_csv_reader(content).fieldnames or ())
    if all(column in headers for column in CANONICAL_COLUMNS):
        return parse_canonical_csv(content)
    if source_key == "tla_amex":
        return parse_amex_monthly_csv(content, filename=filename)
    _validate_headers(headers)
    return []


def serialize_canonical_csv(rows: list[dict[str, object]]) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=CANONICAL_COLUMNS)
    writer.writeheader()
    for row in rows:
        writer.writerow({
            "id_movimiento": row["external_id"], "empresa": row["company"], "banco": row["bank"],
            "tipo_cuenta": row["account_type"], "naturaleza_cuenta": row["account_nature"],
            "cuenta": row["account"], "clabe": row["clabe"], "moneda": row["currency"],
            "fecha_operacion": row["operation_date"].isoformat(),
            "fecha_liquidacion": row["settlement_date"].isoformat() if row["settlement_date"] else "",
            "descripcion_original": row["original_description"], "referencia": row["reference"],
            "contraparte": row["counterparty"], "cargo": row["debit"], "abono": row["credit"],
            "importe_neto": row["net_amount"], "saldo": row["balance"] if row["balance"] is not None else "",
            "titular": row["holder"], "categoria": row["source_category"],
            "subcategoria": row["source_subcategory"], "recurrente": str(row["recurring"]).lower(),
            "impuesto": str(row["tax"]).lower(), "nomina": str(row["payroll"]).lower(),
            "requiere_factura": str(row["requires_invoice"]).lower(), "factura_uuid": row["invoice_uuid"],
            "estatus_conciliacion_factura": row["invoice_reconciliation_status"],
            "estatus_revision": row["review_status"], "periodo_estado": row["statement_period"],
            "archivo_fuente": row["source_filename"], "pagina_fuente": row["source_page"] or "",
            "hash_fuente": row["source_hash"],
        })
    return output.getvalue().encode("utf-8-sig")


def _update_drive_csv(file_id: str, content: bytes) -> None:
    try:
        from google.auth import default
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaIoBaseUpload
    except ImportError as exc:
        raise RuntimeError("Faltan las dependencias de escritura de Google Drive") from exc
    credentials, _ = default(scopes=["https://www.googleapis.com/auth/drive"])
    service = build("drive", "v3", credentials=credentials, cache_discovery=False)
    media = MediaIoBaseUpload(io.BytesIO(content), mimetype="text/csv", resumable=False)
    service.files().update(
        fileId=file_id,
        media_body=media,
        fields="id,modifiedTime,size",
        supportsAllDrives=True,
    ).execute()


def _combined_canonical_csv(current_content: bytes, incoming_content: bytes) -> bytes:
    current_rows = parse_canonical_csv(current_content)
    incoming_rows = parse_canonical_csv(incoming_content)
    existing_ids = {row["external_id"] for row in current_rows}
    if any(row["external_id"] in existing_ids for row in incoming_rows):
        raise HTTPException(409, "La carga contiene movimientos ya existentes; vuelve a previsualizar")
    return serialize_canonical_csv(current_rows + incoming_rows)


def sync_source(key: str, *, force: bool = False) -> dict[str, object]:
    if key not in FINANCE_SOURCE_PATHS:
        raise ValueError("Fuente financiera no válida")
    path = FINANCE_SOURCE_PATHS[key]
    db = SessionLocal()
    try:
        state = _source_state(db, key)
        try:
            content, modified = _read_source(key)
        except FileNotFoundError as exc:
            state.available = False
            state.last_error = str(exc)
            db.commit()
            return _serialize_source(state)
        content_hash = _sha256(content)
        if not force and state.content_hash == content_hash:
            state.available = True
            if modified is not None:
                state.last_modified_at = modified
            state.last_error = None
            db.commit()
            return _serialize_source(state)
        rows = parse_canonical_csv(content)
        with _sync_lock:
            existing = {
                row.external_id: row
                for row in db.query(FinanceMovement).filter(FinanceMovement.source_key == key).all()
            }
            incoming_ids = {str(values["external_id"]) for values in rows}
            # The external history is canonical. If a backed-up/reverted file no
            # longer contains an indexed row, remove that stale index entry too.
            for external_id, movement in existing.items():
                if external_id not in incoming_ids:
                    db.delete(movement)
            for values in rows:
                movement = existing.get(str(values["external_id"]))
                if movement is None:
                    movement = FinanceMovement(source_key=key, **values)
                    db.add(movement)
                else:
                    # Controlled enrichment columns are intentionally omitted.
                    for field, value in values.items():
                        setattr(movement, field, value)
                    movement.indexed_at = datetime.utcnow()
            state.available = True
            state.content_hash = content_hash
            state.row_count = len(rows)
            if modified is not None:
                state.last_modified_at = modified
            state.last_synced_at = datetime.utcnow()
            state.last_error = None
            db.commit()
        return _serialize_source(state)
    except Exception as exc:
        db.rollback()
        state = _source_state(db, key)
        state.available = path.is_file() or bool(_text(FINANCE_SOURCE_FILE_IDS.get(key)))
        state.last_error = str(exc)[:2000]
        db.commit()
        raise
    finally:
        db.close()


def sync_all_sources(*, force: bool = False) -> list[dict[str, object]]:
    results = []
    for key in FINANCE_SOURCE_PATHS:
        try:
            results.append(sync_source(key, force=force))
        except Exception as exc:
            results.append({"key": key, "available": FINANCE_SOURCE_PATHS[key].is_file() or bool(_text(FINANCE_SOURCE_FILE_IDS.get(key))), "error": str(exc)})
    return results


def _serialize_source(row: FinanceSourceState) -> dict[str, object]:
    return {
        "key": row.key, "company": row.company, "bank": row.bank,
        "available": row.available, "row_count": row.row_count,
        "last_modified_at": row.last_modified_at.isoformat() if row.last_modified_at else None,
        "last_synced_at": row.last_synced_at.isoformat() if row.last_synced_at else None,
        "error": row.last_error,
    }


def _movement_dict(row: FinanceMovement) -> dict[str, object]:
    category = row.category_override or row.source_category or ""
    subcategory = row.subcategory_override or row.source_subcategory or ""
    return {
        "id": row.id, "id_movimiento": row.external_id, "empresa": row.company,
        "banco": row.bank, "tipo_cuenta": row.account_type, "naturaleza_cuenta": row.account_nature,
        "moneda": row.currency, "fecha_operacion": row.operation_date.isoformat(),
        "fecha_liquidacion": row.settlement_date.isoformat() if row.settlement_date else None,
        "descripcion_original": row.original_description, "referencia": row.reference,
        "contraparte": row.counterparty, "cargo": float(row.debit), "abono": float(row.credit),
        "importe_neto": float(row.net_amount), "saldo": float(row.balance) if row.balance is not None else None,
        "categoria": category, "subcategoria": subcategory, "recurrente": row.recurring,
        "impuesto": row.tax, "nomina": row.payroll, "requiere_factura": row.requires_invoice,
        "factura_uuid": row.invoice_uuid, "estatus_conciliacion_factura": row.invoice_reconciliation_status,
        "estatus_revision": row.review_status, "periodo_estado": row.statement_period,
        "archivo_fuente": row.source_filename, "pagina_fuente": row.source_page,
    }


def _company_filter(query, company: str):
    return query if company == "CONSOLIDADO" else query.filter(FinanceMovement.company == company)


def _movement_scope(query, company: str = "CONSOLIDADO", bank: str = "", start_date: date | None = None, end_date: date | None = None):
    query = _company_filter(query, company)
    if bank:
        canonical_bank = _canonical_bank(bank)
        query = query.filter(FinanceMovement.bank.in_(BANK_ALIASES.get(canonical_bank, {canonical_bank})))
    if start_date:
        query = query.filter(FinanceMovement.operation_date >= start_date)
    if end_date:
        query = query.filter(FinanceMovement.operation_date <= end_date)
    return query


def overview(company: str = "CONSOLIDADO", bank: str = "", start_date: date | None = None, end_date: date | None = None) -> dict[str, object]:
    db = SessionLocal()
    try:
        query = _movement_scope(db.query(FinanceMovement), company, bank, start_date, end_date)
        movements = query.all()
        today = date.today()
        month_start = today.replace(day=1)
        period_movements = movements if start_date or end_date else [row for row in movements if row.operation_date >= month_start]
        entries = sum((row.net_amount for row in period_movements if row.net_amount > 0), Decimal("0"))
        exits = sum((-row.net_amount for row in period_movements if row.net_amount < 0), Decimal("0"))
        latest_by_account: dict[tuple[str, str, str], FinanceMovement] = {}
        for row in movements:
            key = (row.company, row.bank, row.account or row.source_key)
            if key not in latest_by_account or row.operation_date > latest_by_account[key].operation_date:
                latest_by_account[key] = row
        cash = sum((row.balance or Decimal("0") for row in latest_by_account.values() if "credito" not in _normalized(row.account_nature)), Decimal("0"))
        credit_liability = sum((abs(row.balance or Decimal("0")) for row in latest_by_account.values() if "credito" in _normalized(row.account_nature)), Decimal("0"))
        unclassified = sum(not (row.category_override or row.source_category) for row in movements)
        invoice_gaps = sum(row.requires_invoice and not row.invoice_uuid for row in movements)
        tax_total = sum((-row.net_amount for row in period_movements if row.tax and row.net_amount < 0), Decimal("0"))
        recurring_pending = sum(1 for item in recurring_groups(db, company, bank, start_date, end_date) if item["status"] == "pendiente")
        projections = db.query(FinanceProjection).filter(FinanceProjection.status == "activa", FinanceProjection.due_date >= today)
        if company != "CONSOLIDADO": projections = projections.filter(FinanceProjection.company == company)
        projected = sum((row.amount for row in projections.all()), Decimal("0"))
        source_query = db.query(FinanceSourceState)
        if company != "CONSOLIDADO":
            source_query = source_query.filter(FinanceSourceState.company == company)
        if bank:
            source_query = source_query.filter(FinanceSourceState.bank == bank)
        sources = source_query.order_by(FinanceSourceState.company, FinanceSourceState.bank).all()
        return {
            "company": company,
            "as_of": datetime.utcnow().isoformat(),
            "kpis": {
                "active_cash": float(cash), "credit_liability": float(credit_liability),
                "net_flow_month": float(entries - exits), "entries_month": float(entries),
                "exits_month": float(exits), "unclassified": unclassified,
                "recurring_pending": recurring_pending, "invoice_gaps": invoice_gaps,
                "tax_month": float(tax_total), "future_commitments": float(projected),
            },
            "monthly": _monthly_series(movements),
            "sources": [_serialize_source(item) for item in sources],
        }
    finally:
        db.close()


def _monthly_series(rows: list[FinanceMovement]) -> list[dict[str, object]]:
    values: dict[str, dict[str, Decimal]] = defaultdict(lambda: {"entries": Decimal("0"), "exits": Decimal("0")})
    for row in rows:
        month = row.operation_date.strftime("%Y-%m")
        if row.net_amount >= 0: values[month]["entries"] += row.net_amount
        else: values[month]["exits"] += -row.net_amount
    return [{"month": month, "entries": float(data["entries"]), "exits": float(data["exits"]), "net": float(data["entries"] - data["exits"])} for month, data in sorted(values.items())[-12:]]


def _normalized(value: object) -> str:
    return unicodedata.normalize("NFKD", _text(value)).encode("ascii", "ignore").decode().casefold()


def _recurring_fingerprint(row: FinanceMovement) -> str:
    seed = f"{row.company}|{_normalized(row.counterparty or row.original_description)}"
    return hashlib.sha256(seed.encode()).hexdigest()


def recurring_groups(db, company: str = "CONSOLIDADO", bank: str = "", start_date: date | None = None, end_date: date | None = None) -> list[dict[str, object]]:
    query = _movement_scope(db.query(FinanceMovement), company, bank, start_date, end_date)
    groups: dict[str, list[FinanceMovement]] = defaultdict(list)
    for row in query.order_by(FinanceMovement.operation_date).all():
        groups[_recurring_fingerprint(row)].append(row)
    decisions = {row.fingerprint: row for row in db.query(FinanceRecurringDecision).all()}
    result = []
    for fingerprint, rows in groups.items():
        months = sorted({row.operation_date.strftime("%Y-%m") for row in rows})
        decision = decisions.get(fingerprint)
        suggested = len(months) >= 2
        if not suggested and not decision:
            continue
        amounts = [abs(row.net_amount) for row in rows]
        result.append({
            "fingerprint": fingerprint, "company": rows[0].company,
            "label": rows[0].counterparty or rows[0].original_description,
            "occurrences": len(rows), "months": len(months),
            "average_amount": float(sum(amounts, Decimal("0")) / len(amounts)),
            "last_date": rows[-1].operation_date.isoformat(),
            "status": decision.status if decision else "pendiente",
            "note": decision.note if decision else None,
            "basis": "Mismo beneficiario/descripción en al menos dos meses" if suggested else "Decisión humana conservada",
        })
    return sorted(result, key=lambda item: (item["status"] != "pendiente", item["last_date"]), reverse=True)


def _rule_matches(rule: FinanceClassificationRule, movement: FinanceMovement) -> bool:
    if rule.company and rule.company != movement.company:
        return False
    source = _text(getattr(movement, {"descripcion_original": "original_description", "contraparte": "counterparty", "referencia": "reference", "banco": "bank"}[rule.field]))
    source_normalized, value_normalized = _normalized(source), _normalized(rule.value)
    if rule.operator == "igual": return source_normalized == value_normalized
    if rule.operator == "empieza_con": return source_normalized.startswith(value_normalized)
    if rule.operator == "regex":
        try: return bool(re.search(rule.value, source, flags=re.IGNORECASE))
        except re.error: return False
    return value_normalized in source_normalized


def _projection_dict(row: FinanceProjection) -> dict[str, object]:
    return {"id": row.id, "company": row.company, "due_date": row.due_date.isoformat(), "concept": row.concept, "amount": float(row.amount), "scenario": row.scenario, "status": row.status, "source": row.source}


def _rule_dict(row: FinanceClassificationRule) -> dict[str, object]:
    return {"id": row.id, "name": row.name, "priority": row.priority, "field": row.field, "operator": row.operator, "value": row.value, "company": row.company, "category": row.category, "subcategory": row.subcategory, "enabled": row.enabled, "exclusion": row.exclusion, "updated_at": row.updated_at.isoformat()}


def _budget_dict(row: FinanceBudgetItem, actual: Decimal = Decimal("0")) -> dict[str, object]:
    return {"id": row.id, "company": row.company, "month": row.month.isoformat(), "category": row.category, "budget": float(row.amount), "actual": float(actual), "variance": float(row.amount - actual)}


def _invoice_dict(row: FinanceInvoice) -> dict[str, object]:
    return {"id": row.id, "filename": Path(row.file_path).name, "file_type": row.file_type, "uuid": row.uuid, "issuer_rfc": row.issuer_rfc, "receiver_rfc": row.receiver_rfc, "issued_at": row.issued_at.isoformat() if row.issued_at else None, "total": float(row.total) if row.total is not None else None, "currency": row.currency, "status": row.status, "parse_error": row.parse_error}


def _xml_invoice(path: Path) -> dict[str, object]:
    root = ElementTree.fromstring(path.read_bytes())
    attrs = {key.split("}")[-1].casefold(): value for key, value in root.attrib.items()}
    uuid_value = None
    issuer_rfc = receiver_rfc = None
    for element in root.iter():
        tag = element.tag.split("}")[-1].casefold()
        lower = {key.casefold(): value for key, value in element.attrib.items()}
        if tag == "timbrefiscaldigital": uuid_value = lower.get("uuid")
        elif tag == "emisor": issuer_rfc = lower.get("rfc")
        elif tag == "receptor": receiver_rfc = lower.get("rfc")
    issued = attrs.get("fecha")
    return {"uuid": uuid_value, "issuer_rfc": issuer_rfc, "receiver_rfc": receiver_rfc, "issued_at": datetime.fromisoformat(issued.replace("Z", "+00:00")).replace(tzinfo=None) if issued else None, "total": _decimal(attrs.get("total"), default=None), "currency": attrs.get("moneda"), "payment_method": attrs.get("metodopago"), "parse_error": None}


def scan_invoices() -> dict[str, object]:
    db = SessionLocal()
    indexed = errors = 0
    try:
        if not FINANCE_INVOICES_PATH.is_dir():
            return {"available": False, "indexed": 0, "errors": 0, "message": "La carpeta de facturas no está montada en este servidor"}
        for path in FINANCE_INVOICES_PATH.rglob("*"):
            if not path.is_file() or path.suffix.casefold() not in {".xml", ".pdf"}:
                continue
            content = path.read_bytes(); digest = _sha256(content)
            invoice = db.query(FinanceInvoice).filter(FinanceInvoice.file_path == str(path)).one_or_none()
            if invoice and invoice.file_hash == digest:
                continue
            values: dict[str, object] = {"uuid": None, "issuer_rfc": None, "receiver_rfc": None, "issued_at": None, "total": None, "currency": None, "payment_method": None, "parse_error": None}
            if path.suffix.casefold() == ".xml":
                try: values.update(_xml_invoice(path))
                except Exception as exc: values["parse_error"] = f"XML no interpretable: {exc}"; errors += 1
            else:
                values["parse_error"] = "PDF indexado como evidencia; se requiere XML para extraer datos CFDI"
            if not invoice:
                invoice = FinanceInvoice(file_path=str(path), file_hash=digest, file_type=path.suffix.casefold()[1:])
                db.add(invoice)
            invoice.file_hash = digest
            for field, value in values.items(): setattr(invoice, field, value)
            invoice.indexed_at = datetime.utcnow(); indexed += 1
        db.commit()
        return {"available": True, "indexed": indexed, "errors": errors}
    finally:
        db.close()


@router.get("/sources")
def list_sources():
    db = SessionLocal()
    try:
        for key in FINANCE_SOURCE_PATHS: _source_state(db, key)
        db.commit()
        return {"sources": [_serialize_source(row) for row in db.query(FinanceSourceState).order_by(FinanceSourceState.company, FinanceSourceState.bank).all()]}
    finally: db.close()


@router.post("/sources/sync")
def synchronize_sources(force: bool = False, _profile=Depends(require_module_access("finanzas", operation=True))):
    return {"sources": sync_all_sources(force=force)}


@router.get("/overview")
def get_overview(company: str = Query(default="CONSOLIDADO", pattern="^(CONSOLIDADO|TLA|TS)$"), bank: str = "", start_date: date | None = None, end_date: date | None = None):
    return overview(company, bank, start_date, end_date)


@router.get("/movements")
def list_movements(company: str = "CONSOLIDADO", search: str = "", category: str = "", bank: str = "", start_date: date | None = None, end_date: date | None = None, page: int = Query(1, ge=1), page_size: int = Query(50, ge=1, le=5000), sort: str = "operation_date", direction: str = "desc"):
    db = SessionLocal()
    try:
        query = _movement_scope(db.query(FinanceMovement), company, bank, start_date, end_date)
        if search:
            pattern = f"%{search.strip()}%"
            query = query.filter(or_(FinanceMovement.original_description.ilike(pattern), FinanceMovement.counterparty.ilike(pattern), FinanceMovement.reference.ilike(pattern), FinanceMovement.external_id.ilike(pattern)))
        if category: query = query.filter(or_(FinanceMovement.category_override == category, (FinanceMovement.category_override.is_(None)) & (FinanceMovement.source_category == category)))
        total = query.count()
        allowed_sort = {"operation_date": FinanceMovement.operation_date, "net_amount": FinanceMovement.net_amount, "bank": FinanceMovement.bank, "company": FinanceMovement.company}
        order = allowed_sort.get(sort, FinanceMovement.operation_date)
        rows = query.order_by(desc(order) if direction == "desc" else asc(order), desc(FinanceMovement.id)).offset((page - 1) * page_size).limit(page_size).all()
        categories = sorted({item for pair in db.query(FinanceMovement.category_override, FinanceMovement.source_category).all() for item in pair if item})
        return {"items": [_movement_dict(row) for row in rows], "total": total, "page": page, "page_size": page_size, "categories": categories}
    finally: db.close()


@router.get("/movements/export")
def export_movements(company: str = "CONSOLIDADO", search: str = "", bank: str = "", start_date: date | None = None, end_date: date | None = None):
    db = SessionLocal()
    try:
        query = _movement_scope(db.query(FinanceMovement), company, bank, start_date, end_date)
        if search:
            pattern = f"%{search.strip()}%"
            query = query.filter(or_(FinanceMovement.original_description.ilike(pattern), FinanceMovement.counterparty.ilike(pattern), FinanceMovement.reference.ilike(pattern), FinanceMovement.external_id.ilike(pattern)))

        headers = ["ID", "Empresa", "Banco", "Fecha", "Descripción", "Contraparte", "Moneda", "Cargo", "Abono", "Importe neto", "Saldo", "Categoría", "Subcategoría", "UUID factura", "Archivo fuente", "Página"]
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Movimientos"
        sheet.append(headers)

        for row in query.order_by(FinanceMovement.operation_date, FinanceMovement.id).all():
            item = _movement_dict(row)
            sheet.append([
                item["id_movimiento"], item["empresa"], item["banco"], row.operation_date,
                item["descripcion_original"], item["contraparte"], item["moneda"],
                item["cargo"], item["abono"], item["importe_neto"], item["saldo"],
                item["categoria"], item["subcategoria"], item["factura_uuid"],
                item["archivo_fuente"], item["pagina_fuente"],
            ])

        header_fill = PatternFill("solid", fgColor="17365D")
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_number in range(2, sheet.max_row + 1):
            sheet.cell(row_number, 4).number_format = "dd/mm/yyyy"
            for column_number in (8, 9, 10, 11):
                sheet.cell(row_number, column_number).number_format = '$#,##0.00;[Red]-$#,##0.00'

        widths = (24, 12, 16, 13, 48, 32, 12, 16, 16, 18, 16, 24, 24, 38, 36, 10)
        for column_number, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(column_number)].width = width
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False

        output = io.BytesIO()
        workbook.save(output)
        filename = f"movimientos-{company.casefold()}-{date.today():%Y%m%d}.xlsx"
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    finally: db.close()


@router.patch("/movements/{movement_id}")
def update_movement(movement_id: str, payload: MovementPatch, profile: AccessProfile = Depends(require_module_access("finanzas", operation=True))):
    db = SessionLocal()
    try:
        row = db.get(FinanceMovement, movement_id)
        if not row: raise HTTPException(404, "Movimiento no encontrado")
        values = payload.model_dump(exclude_unset=True)
        mapping = {"categoria": "category_override", "subcategoria": "subcategory_override", "recurrente": "recurring", "impuesto": "tax", "nomina": "payroll", "requiere_factura": "requires_invoice", "estatus_revision": "review_status"}
        for key, value in values.items(): setattr(row, mapping[key], value)
        row.enrichment_updated_by = profile.username; row.enrichment_updated_at = datetime.utcnow()
        db.commit(); db.refresh(row)
        return {"movement": _movement_dict(row)}
    finally: db.close()


@router.post("/movements/bulk-classify")
def bulk_classify(payload: BulkClassification, profile: AccessProfile = Depends(require_module_access("finanzas", operation=True))):
    db = SessionLocal()
    try:
        rows = db.query(FinanceMovement).filter(FinanceMovement.id.in_(payload.movement_ids)).all()
        for row in rows:
            row.category_override = payload.categoria; row.subcategory_override = payload.subcategoria
            row.enrichment_updated_by = profile.username; row.enrichment_updated_at = datetime.utcnow()
        db.commit(); return {"updated": len(rows)}
    finally: db.close()


@router.get("/recurring")
def list_recurring(company: str = "CONSOLIDADO", bank: str = "", start_date: date | None = None, end_date: date | None = None):
    db = SessionLocal()
    try: return {"items": recurring_groups(db, company, bank, start_date, end_date)}
    finally: db.close()


@router.put("/recurring/{fingerprint}")
def decide_recurring(fingerprint: str, payload: RecurringInput, profile: AccessProfile = Depends(require_module_access("finanzas", operation=True))):
    db = SessionLocal()
    try:
        matches = [row for row in db.query(FinanceMovement).all() if _recurring_fingerprint(row) == fingerprint]
        if not matches: raise HTTPException(404, "Grupo recurrente no encontrado")
        row = db.get(FinanceRecurringDecision, fingerprint)
        if not row:
            row = FinanceRecurringDecision(fingerprint=fingerprint, company=matches[0].company, label=matches[0].counterparty or matches[0].original_description, status=payload.status, decided_by=profile.username)
            db.add(row)
        row.status = payload.status; row.note = payload.note; row.decided_by = profile.username; row.updated_at = datetime.utcnow()
        for movement in matches: movement.recurring = payload.status == "confirmado"
        db.commit(); return {"success": True}
    finally: db.close()


@router.get("/invoices")
def list_invoices(status: str = ""):
    db = SessionLocal()
    try:
        query = db.query(FinanceInvoice)
        if status: query = query.filter(FinanceInvoice.status == status)
        return {"items": [_invoice_dict(row) for row in query.order_by(desc(FinanceInvoice.issued_at), desc(FinanceInvoice.indexed_at)).limit(1000).all()], "folder_available": FINANCE_INVOICES_PATH.is_dir()}
    finally: db.close()


@router.get("/invoices/{invoice_id}/suggestions")
def invoice_suggestions(invoice_id: str, limit: int = Query(10, ge=1, le=50)):
    db = SessionLocal()
    try:
        invoice = db.get(FinanceInvoice, invoice_id)
        if not invoice: raise HTTPException(404, "Factura no encontrada")
        if invoice.total is None or invoice.issued_at is None: return {"items": [], "reason": "La factura no tiene total y fecha extraíbles"}
        start, end = invoice.issued_at.date() - timedelta(days=15), invoice.issued_at.date() + timedelta(days=15)
        candidates = db.query(FinanceMovement).filter(FinanceMovement.operation_date.between(start, end), FinanceMovement.net_amount < 0).all()
        ranked = []
        for movement in candidates:
            difference = abs(abs(movement.net_amount) - invoice.total)
            amount_score = max(Decimal("0"), Decimal("1") - (difference / max(invoice.total, Decimal("1"))))
            day_gap = abs((movement.operation_date - invoice.issued_at.date()).days)
            date_score = Decimal(str(max(0, 1 - day_gap / 15)))
            confidence = (amount_score * Decimal("0.8") + date_score * Decimal("0.2")) * 100
            if confidence < 50: continue
            ranked.append({"movement": _movement_dict(movement), "confidence": round(float(confidence), 1), "rationale": f"Diferencia {money_text(difference)}; {day_gap} días respecto a emisión"})
        ranked.sort(key=lambda item: item["confidence"], reverse=True)
        return {"items": ranked[:limit], "reason": "Sugerencias por importe y cercanía de fecha; requieren confirmación humana"}
    finally: db.close()


def money_text(value: Decimal) -> str:
    return f"${value:,.2f}"


@router.post("/invoices/scan")
def index_invoices(_profile=Depends(require_module_access("finanzas", operation=True))):
    return scan_invoices()


@router.post("/invoices/{invoice_id}/match")
def match_invoice(invoice_id: str, payload: InvoiceMatchInput, profile: AccessProfile = Depends(require_module_access("finanzas", operation=True))):
    db = SessionLocal()
    try:
        invoice, movement = db.get(FinanceInvoice, invoice_id), db.get(FinanceMovement, payload.movement_id)
        if not invoice or not movement: raise HTTPException(404, "Factura o movimiento no encontrado")
        match = db.query(FinanceInvoiceMatch).filter_by(invoice_id=invoice_id, movement_id=payload.movement_id).one_or_none()
        if not match: match = FinanceInvoiceMatch(invoice_id=invoice_id, movement_id=payload.movement_id, status=payload.status); db.add(match)
        match.status = payload.status; match.confirmed_by = profile.username
        if payload.status == "confirmado":
            invoice.status = "conciliada"; movement.invoice_uuid = invoice.uuid; movement.invoice_reconciliation_status = "conciliada"
        db.commit(); return {"success": True}
    finally: db.close()


@router.get("/cash-flow")
def cash_flow(company: str = "CONSOLIDADO", scenario: str = "base", days: int = Query(90, ge=7, le=730)):
    db = SessionLocal()
    try:
        start, end = date.today(), date.today() + timedelta(days=days)
        projections = db.query(FinanceProjection).filter(FinanceProjection.status == "activa", FinanceProjection.due_date.between(start, end), FinanceProjection.scenario == scenario)
        if company != "CONSOLIDADO": projections = projections.filter(FinanceProjection.company == company)
        items = projections.order_by(FinanceProjection.due_date).all()
        return {"items": [_projection_dict(row) for row in items], "total": float(sum((row.amount for row in items), Decimal("0")))}
    finally: db.close()


@router.get("/budgets")
def list_budgets(company: str = "CONSOLIDADO", month: date | None = None):
    target = (month or date.today()).replace(day=1)
    db = SessionLocal()
    try:
        query = db.query(FinanceBudgetItem).filter(FinanceBudgetItem.month == target)
        if company != "CONSOLIDADO": query = query.filter(FinanceBudgetItem.company == company)
        rows = query.order_by(FinanceBudgetItem.company, FinanceBudgetItem.category).all()
        result = []
        month_end = (target.replace(day=28) + timedelta(days=4)).replace(day=1)
        for row in rows:
            movements = db.query(FinanceMovement).filter(FinanceMovement.company == row.company, FinanceMovement.operation_date >= target, FinanceMovement.operation_date < month_end, FinanceMovement.net_amount < 0, or_(FinanceMovement.category_override == row.category, (FinanceMovement.category_override.is_(None)) & (FinanceMovement.source_category == row.category))).all()
            actual = sum((-movement.net_amount for movement in movements), Decimal("0"))
            result.append(_budget_dict(row, actual))
        return {"items": result, "month": target.isoformat(), "budget": sum(item["budget"] for item in result), "actual": sum(item["actual"] for item in result)}
    finally: db.close()


@router.put("/budgets")
def upsert_budget(payload: BudgetInput, profile: AccessProfile = Depends(require_module_access("finanzas", operation=True))):
    db = SessionLocal()
    try:
        month = payload.month.replace(day=1)
        row = db.query(FinanceBudgetItem).filter_by(company=payload.company, month=month, category=payload.category).one_or_none()
        if not row:
            row = FinanceBudgetItem(company=payload.company, month=month, category=payload.category, amount=payload.amount, created_by=profile.username); db.add(row)
        else: row.amount = payload.amount
        db.commit(); db.refresh(row); return {"budget": _budget_dict(row)}
    finally: db.close()


@router.post("/projections", status_code=201)
def create_projection(payload: ProjectionInput, profile: AccessProfile = Depends(require_module_access("finanzas", operation=True))):
    db = SessionLocal()
    try:
        row = FinanceProjection(**payload.model_dump(), created_by=profile.username); db.add(row); db.commit(); db.refresh(row)
        return {"projection": _projection_dict(row)}
    finally: db.close()


@router.put("/projections/{projection_id}")
def update_projection(projection_id: str, payload: ProjectionInput, _profile=Depends(require_module_access("finanzas", operation=True))):
    db = SessionLocal()
    try:
        row = db.get(FinanceProjection, projection_id)
        if not row: raise HTTPException(404, "Proyección no encontrada")
        for key, value in payload.model_dump().items(): setattr(row, key, value)
        db.commit(); db.refresh(row); return {"projection": _projection_dict(row)}
    finally: db.close()


@router.delete("/projections/{projection_id}")
def cancel_projection(projection_id: str, _profile=Depends(require_module_access("finanzas", operation=True))):
    db = SessionLocal()
    try:
        row = db.get(FinanceProjection, projection_id)
        if not row: raise HTTPException(404, "Proyección no encontrada")
        row.status = "cancelada"; db.commit(); return {"success": True}
    finally: db.close()


@router.get("/rules")
def list_rules():
    db = SessionLocal()
    try: return {"items": [_rule_dict(row) for row in db.query(FinanceClassificationRule).order_by(FinanceClassificationRule.priority, FinanceClassificationRule.name).all()]}
    finally: db.close()


@router.post("/rules", status_code=201)
def create_rule(payload: RuleInput, profile: AccessProfile = Depends(require_module_access("finanzas", operation=True))):
    if payload.operator == "regex":
        try: re.compile(payload.value)
        except re.error as exc: raise HTTPException(422, f"Expresión regular inválida: {exc}") from exc
    db = SessionLocal()
    try:
        row = FinanceClassificationRule(**payload.model_dump(), created_by=profile.username); db.add(row); db.commit(); db.refresh(row)
        return {"rule": _rule_dict(row)}
    finally: db.close()


@router.delete("/rules/{rule_id}")
def delete_rule(rule_id: str, _profile=Depends(require_module_access("finanzas", operation=True))):
    db = SessionLocal()
    try:
        row = db.get(FinanceClassificationRule, rule_id)
        if not row: raise HTTPException(404, "Regla no encontrada")
        db.delete(row); db.commit(); return {"success": True}
    finally: db.close()


@router.post("/rules/{rule_id}/preview")
def preview_rule(rule_id: str, limit: int = Query(25, ge=1, le=200), _profile=Depends(require_module_access("finanzas", operation=True))):
    db = SessionLocal()
    try:
        rule = db.get(FinanceClassificationRule, rule_id)
        if not rule: raise HTTPException(404, "Regla no encontrada")
        matches = [row for row in db.query(FinanceMovement).all() if _rule_matches(rule, row)]
        conflicts = sum(bool(row.category_override or row.source_category) and (row.category_override or row.source_category) != rule.category for row in matches)
        return {"total": len(matches), "conflicts": conflicts, "sample": [_movement_dict(row) for row in matches[:limit]]}
    finally: db.close()


@router.post("/rules/{rule_id}/apply")
def apply_rule(rule_id: str, profile: AccessProfile = Depends(require_module_access("finanzas", operation=True))):
    db = SessionLocal()
    try:
        rule = db.get(FinanceClassificationRule, rule_id)
        if not rule: raise HTTPException(404, "Regla no encontrada")
        updated = 0; run_id = str(uuid.uuid4())
        for row in db.query(FinanceMovement).all():
            if not _rule_matches(rule, row): continue
            db.add(FinanceRuleApplication(run_id=run_id, rule_id=rule.id, movement_id=row.id, before_category=row.category_override, before_subcategory=row.subcategory_override, before_review_status=row.review_status, applied_by=profile.username))
            if rule.exclusion: row.review_status = "excluido_regla"
            else: row.category_override = rule.category; row.subcategory_override = rule.subcategory
            row.enrichment_updated_by = profile.username; row.enrichment_updated_at = datetime.utcnow(); updated += 1
        db.commit(); return {"updated": updated, "run_id": run_id}
    finally: db.close()


@router.post("/rules/{rule_id}/revert")
def revert_rule(rule_id: str, _profile=Depends(require_module_access("finanzas", operation=True))):
    db = SessionLocal()
    try:
        latest = db.query(FinanceRuleApplication).filter(FinanceRuleApplication.rule_id == rule_id, FinanceRuleApplication.reverted_at.is_(None)).order_by(desc(FinanceRuleApplication.created_at)).first()
        if not latest: raise HTTPException(409, "La regla no tiene una ejecución reversible")
        applications = db.query(FinanceRuleApplication).filter(FinanceRuleApplication.run_id == latest.run_id, FinanceRuleApplication.reverted_at.is_(None)).all()
        reverted_at = datetime.utcnow(); restored = 0
        for application in applications:
            movement = db.get(FinanceMovement, application.movement_id)
            if movement:
                movement.category_override = application.before_category; movement.subcategory_override = application.before_subcategory; movement.review_status = application.before_review_status; restored += 1
            application.reverted_at = reverted_at
        db.commit(); return {"restored": restored, "run_id": latest.run_id}
    finally: db.close()


@router.post("/ingestions/preview")
async def preview_ingestion(source_key: str, file: UploadFile = File(...), profile: AccessProfile = Depends(require_module_access("finanzas", operation=True))):
    if source_key not in FINANCE_SOURCE_PATHS: raise HTTPException(422, "Fuente financiera no válida")
    content = await file.read()
    if len(content) > 25 * 1024 * 1024: raise HTTPException(413, "El archivo supera 25 MB")
    digest = _sha256(content)
    db = SessionLocal()
    try:
        duplicate = db.query(FinanceIngestion).filter(FinanceIngestion.file_hash == digest, FinanceIngestion.status == "publicada").first()
        if duplicate: raise HTTPException(409, "Este archivo ya fue publicado")
        if (file.filename or "").casefold().endswith(".pdf"):
            raise HTTPException(422, "El PDF se conserva como estado original, pero este formato bancario aún no tiene un parser validado en este servidor. Usa el CSV canónico para no interpretar movimientos incorrectamente.")
        try: rows = parse_ingestion_csv(source_key, content, filename=file.filename or "estado.csv")
        except ValueError as exc: raise HTTPException(422, str(exc)) from exc
        existing = {value for (value,) in db.query(FinanceMovement.external_id).filter(FinanceMovement.source_key == source_key).all()}
        RUNTIME_DIR.mkdir(parents=True, exist_ok=True)
        ingestion_id = str(uuid.uuid4()); staging = RUNTIME_DIR / f"{ingestion_id}.csv"
        canonical_content = serialize_canonical_csv(rows)
        temporary = staging.with_suffix(".tmp"); temporary.write_bytes(canonical_content); os.chmod(temporary, 0o600); os.replace(temporary, staging)
        record = FinanceIngestion(id=ingestion_id, source_key=source_key, filename=Path(file.filename or "estado.csv").name, file_hash=digest, status="previsualizada", row_count=len(rows), new_rows=sum(row["external_id"] not in existing for row in rows), duplicate_rows=sum(row["external_id"] in existing for row in rows), staging_path=str(staging), created_by=profile.username)
        db.add(record); db.commit()
        return {"ingestion_id": record.id, "rows": record.row_count, "new_rows": record.new_rows, "duplicates": record.duplicate_rows, "sample": [{"id_movimiento": row["external_id"], "fecha_operacion": row["operation_date"].isoformat(), "descripcion_original": row["original_description"], "importe_neto": float(row["net_amount"])} for row in rows[:20]]}
    finally: db.close()


@router.post("/ingestions/{ingestion_id}/publish")
def publish_ingestion(ingestion_id: str, _profile=Depends(require_module_access("finanzas", operation=True))):
    db = SessionLocal()
    try:
        record = db.get(FinanceIngestion, ingestion_id)
        if not record or record.status != "previsualizada": raise HTTPException(409, "La carga no está disponible para publicar")
        staging, destination = Path(record.staging_path or ""), FINANCE_SOURCE_PATHS[record.source_key]
        if not staging.is_file(): raise HTTPException(410, "La previsualización expiró")
        file_id = _text(FINANCE_SOURCE_FILE_IDS.get(record.source_key))
        use_drive = not destination.is_file() and bool(file_id)
        if not use_drive and not destination.parent.is_dir():
            raise HTTPException(503, "La fuente canónica no está disponible localmente ni en Google Drive")
        try:
            current_content, _modified = _read_source(record.source_key)
            combined_content = _combined_canonical_csv(current_content, staging.read_bytes())
        except ValueError as exc:
            raise HTTPException(422, str(exc)) from exc
        if use_drive:
            RUNTIME_DIR.mkdir(parents=True, exist_ok=True)
            backup = RUNTIME_DIR / f"{record.id}.backup.csv"
            temporary_backup = backup.with_suffix(".tmp")
            temporary_backup.write_bytes(current_content); os.chmod(temporary_backup, 0o600); os.replace(temporary_backup, backup)
            record.backup_path = str(backup)
            _update_drive_csv(file_id, combined_content)
        else:
            backup = destination.with_name(f"{destination.stem}.backup-{datetime.utcnow():%Y%m%dT%H%M%SZ}{destination.suffix}")
            if destination.exists(): shutil.copy2(destination, backup); record.backup_path = str(backup)
            with tempfile.NamedTemporaryFile("wb", dir=destination.parent, delete=False) as handle:
                handle.write(combined_content); temp_name = handle.name
            os.replace(temp_name, destination)
        record.status = "publicada"; record.published_at = datetime.utcnow(); db.commit()
        sync_source(record.source_key, force=True)
        return {"success": True, "backup_created": bool(record.backup_path), "published_rows": record.new_rows}
    finally: db.close()


@router.post("/ingestions/{ingestion_id}/revert")
def revert_ingestion(ingestion_id: str, _profile=Depends(require_module_access("finanzas", operation=True))):
    db = SessionLocal()
    try:
        record = db.get(FinanceIngestion, ingestion_id)
        if not record or record.status != "publicada" or not record.backup_path: raise HTTPException(409, "La carga no tiene un respaldo reversible")
        backup, destination = Path(record.backup_path), FINANCE_SOURCE_PATHS[record.source_key]
        if not backup.is_file(): raise HTTPException(410, "El respaldo ya no está disponible")
        file_id = _text(FINANCE_SOURCE_FILE_IDS.get(record.source_key))
        if not destination.is_file() and file_id:
            _update_drive_csv(file_id, backup.read_bytes())
        else:
            if not destination.parent.is_dir(): raise HTTPException(503, "La fuente canónica no está disponible")
            with tempfile.NamedTemporaryFile("wb", dir=destination.parent, delete=False) as handle: handle.write(backup.read_bytes()); temp_name = handle.name
            os.replace(temp_name, destination)
        record.status = "revertida"; record.reverted_at = datetime.utcnow(); db.commit(); sync_source(record.source_key, force=True)
        return {"success": True}
    finally: db.close()
