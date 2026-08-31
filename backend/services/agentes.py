from __future__ import annotations

import copy
import datetime as dt
import hashlib
import io
import json
import os
import threading
from dataclasses import dataclass

from fastapi import APIRouter, Depends, HTTPException
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from pydantic import BaseModel, Field

from services.auth import AccessProfile
from services.authorization import require_module_access
from services.data_cache import data_cache
from services.metlife_agent_directory import (
    AGENTS_FILE_ID_ENV,
    DEFAULT_AGENTS_FILE_ID,
    clear_agent_directory_cache,
    normalize_agent_key,
)
from services.pendientes import _download_workbook


router = APIRouter(prefix="/agentes", tags=["agentes"])
_write_lock = threading.RLock()

SHEET_NAME = "Datos"
SOURCE_COLUMNS = {
    "nombres": "Nombres",
    "apellido_paterno": "Apellido_Paterno",
    "apellido_materno": "Apellido_Materno",
    "clave_arranque": "CLAVE_ARRANQUE",
    "clave_definitiva": "CLAVE_DEFINITIVA",
    "promotoria": "Promotoria",
    "rfc": "RFC",
    "telefono_particular": "Telefono_Particular",
    "correo_personal": "Correo_Personal",
    "inicio_vigencia_cedula": "Inicio_Vigencia_Cedula",
    "fin_vigencia_cedula": "Fin_Vigencia_Cedula",
    "clasificacion_comercial": "Clasificación Comercial",
    "estatus_met": "Estatus_Met",
}
REQUIRED_COLUMNS = set(SOURCE_COLUMNS.values()) | {"Nombre"}


class AgentFields(BaseModel):
    nombres: str = Field(min_length=1, max_length=160)
    apellido_paterno: str = Field(default="", max_length=100)
    apellido_materno: str = Field(default="", max_length=100)
    clave_arranque: str = Field(default="", max_length=50)
    clave_definitiva: str = Field(default="", max_length=50)
    promotoria: str = Field(min_length=1, max_length=100)
    rfc: str = Field(default="", max_length=20)
    telefono_particular: str = Field(default="", max_length=40)
    correo_personal: str = Field(default="", max_length=320)
    inicio_vigencia_cedula: dt.date | None = None
    fin_vigencia_cedula: dt.date | None = None
    clasificacion_comercial: str = Field(default="", max_length=120)
    estatus_met: str = Field(default="", max_length=100)


class CreateAgentPayload(AgentFields):
    version: str = Field(min_length=64, max_length=64)


class UpdateAgentPayload(AgentFields):
    version: str = Field(min_length=64, max_length=64)
    fingerprint: str = Field(min_length=64, max_length=64)


@dataclass(frozen=True)
class WorkbookContext:
    workbook: object
    sheet: object
    columns: dict[str, int]


def _file_id() -> str:
    return os.getenv(AGENTS_FILE_ID_ENV, "").strip() or DEFAULT_AGENTS_FILE_ID


def _source_url() -> str:
    return f"https://drive.google.com/file/d/{_file_id()}/view"


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _date_to_iso(value: object, *, epoch: dt.datetime | None = None) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, (int, float)) and epoch is not None:
        try:
            parsed = from_excel(value, epoch=epoch)
            return (parsed.date() if isinstance(parsed, dt.datetime) else parsed).isoformat()
        except (TypeError, ValueError, OverflowError):
            return ""
    text = _clean_text(value).split(" ", 1)[0]
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y", "%m/%d/%Y"):
        try:
            return dt.datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    return ""


def _workbook_context(workbook_bytes: bytes) -> WorkbookContext:
    workbook = load_workbook(io.BytesIO(workbook_bytes))
    sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
    columns = {
        _clean_text(cell.value): cell.column
        for cell in sheet[1]
        if _clean_text(cell.value)
    }
    missing = sorted(REQUIRED_COLUMNS.difference(columns))
    if missing:
        raise ValueError("La base de agentes no contiene: " + ", ".join(missing))
    return WorkbookContext(workbook=workbook, sheet=sheet, columns=columns)


def _serialized_row(sheet, row_number: int) -> list[str]:
    serialized: list[str] = []
    for cell in sheet[row_number]:
        value = cell.value
        if isinstance(value, (dt.datetime, dt.date)):
            serialized.append(value.isoformat())
        else:
            serialized.append(_clean_text(value))
    return serialized


def _row_fingerprint(sheet, row_number: int) -> str:
    payload = json.dumps(
        _serialized_row(sheet, row_number),
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def _agent_from_row(context: WorkbookContext, row_number: int) -> dict:
    sheet = context.sheet
    columns = context.columns
    get = lambda source: sheet.cell(row_number, columns[source]).value
    names = [_clean_text(get(SOURCE_COLUMNS[field])) for field in (
        "nombres", "apellido_paterno", "apellido_materno"
    )]
    full_name = " ".join(value for value in names if value) or _clean_text(get("Nombre"))
    return {
        "row_number": row_number,
        "fingerprint": _row_fingerprint(sheet, row_number),
        "nombre": full_name,
        "nombres": names[0],
        "apellido_paterno": names[1],
        "apellido_materno": names[2],
        "clave_arranque": normalize_agent_key(get(SOURCE_COLUMNS["clave_arranque"])),
        "clave_definitiva": normalize_agent_key(get(SOURCE_COLUMNS["clave_definitiva"])),
        "promotoria": _clean_text(get(SOURCE_COLUMNS["promotoria"])),
        "rfc": _clean_text(get(SOURCE_COLUMNS["rfc"])).upper().replace(" ", ""),
        "telefono_particular": _clean_text(get(SOURCE_COLUMNS["telefono_particular"])),
        "correo_personal": _clean_text(get(SOURCE_COLUMNS["correo_personal"])).lower(),
        "inicio_vigencia_cedula": _date_to_iso(
            get(SOURCE_COLUMNS["inicio_vigencia_cedula"]), epoch=context.workbook.epoch
        ),
        "fin_vigencia_cedula": _date_to_iso(
            get(SOURCE_COLUMNS["fin_vigencia_cedula"]), epoch=context.workbook.epoch
        ),
        "clasificacion_comercial": _clean_text(
            get(SOURCE_COLUMNS["clasificacion_comercial"])
        ),
        "estatus_met": _clean_text(get(SOURCE_COLUMNS["estatus_met"])),
    }


def build_agent_directory(workbook_bytes: bytes, *, can_operate: bool = False) -> dict:
    context = _workbook_context(workbook_bytes)
    agents = []
    for row_number in range(2, context.sheet.max_row + 1):
        if not any(_clean_text(cell.value) for cell in context.sheet[row_number]):
            continue
        agents.append(_agent_from_row(context, row_number))
    agents.sort(key=lambda row: (row["nombre"].casefold(), row["row_number"]))
    return {
        "version": hashlib.sha256(workbook_bytes).hexdigest(),
        "can_operate": can_operate,
        "source_url": _source_url(),
        "agents": agents,
        "catalogs": {
            "promotorias": sorted({row["promotoria"] for row in agents if row["promotoria"]}),
            "clasificaciones": sorted({row["clasificacion_comercial"] for row in agents if row["clasificacion_comercial"]}),
            "estatus_met": sorted({row["estatus_met"] for row in agents if row["estatus_met"]}),
        },
    }


def _normalized_payload(payload: AgentFields) -> dict[str, object]:
    values = payload.model_dump()
    for field in ("nombres", "apellido_paterno", "apellido_materno", "promotoria", "telefono_particular", "correo_personal", "clasificacion_comercial", "estatus_met"):
        values[field] = _clean_text(values[field])
    values["clave_arranque"] = normalize_agent_key(values["clave_arranque"])
    values["clave_definitiva"] = normalize_agent_key(values["clave_definitiva"])
    values["rfc"] = _clean_text(values["rfc"]).upper().replace(" ", "")
    values["correo_personal"] = str(values["correo_personal"]).lower()
    return values


def _validate_unique(context: WorkbookContext, values: dict[str, object], *, skip_row: int | None = None) -> None:
    checks = (
        ("clave_arranque", "clave de arranque"),
        ("clave_definitiva", "clave definitiva"),
        ("rfc", "RFC"),
    )
    for field, label in checks:
        wanted = _clean_text(values[field]).casefold()
        if not wanted:
            continue
        column = context.columns[SOURCE_COLUMNS[field]]
        for row_number in range(2, context.sheet.max_row + 1):
            if row_number == skip_row:
                continue
            existing = _clean_text(context.sheet.cell(row_number, column).value).casefold()
            if existing == wanted:
                raise ValueError(f"Ya existe otro agente con {label} {values[field]}")


def _write_agent(context: WorkbookContext, row_number: int, values: dict[str, object]) -> None:
    for field, source_column in SOURCE_COLUMNS.items():
        cell = context.sheet.cell(row_number, context.columns[source_column])
        cell.value = values[field] or None
        if field in {"inicio_vigencia_cedula", "fin_vigencia_cedula"} and values[field]:
            cell.number_format = "dd/mm/yyyy"
    full_name = " ".join(
        str(values[field]) for field in ("nombres", "apellido_paterno", "apellido_materno")
        if values[field]
    )
    context.sheet.cell(row_number, context.columns["Nombre"]).value = full_name


def mutate_agent_workbook(
    workbook_bytes: bytes,
    payload: AgentFields,
    *,
    row_number: int | None = None,
    fingerprint: str | None = None,
) -> bytes:
    context = _workbook_context(workbook_bytes)
    values = _normalized_payload(payload)
    if row_number is None:
        row_number = context.sheet.max_row + 1
        template_row = max(2, row_number - 1)
        for column in range(1, context.sheet.max_column + 1):
            source = context.sheet.cell(template_row, column)
            target = context.sheet.cell(row_number, column)
            if source.has_style:
                target._style = copy.copy(source._style)
            target.number_format = source.number_format
        if context.sheet.row_dimensions[template_row].height:
            context.sheet.row_dimensions[row_number].height = context.sheet.row_dimensions[template_row].height
    elif row_number < 2 or row_number > context.sheet.max_row:
        raise ValueError("El agente ya no existe en la base")
    elif fingerprint and _row_fingerprint(context.sheet, row_number) != fingerprint:
        raise RuntimeError("El agente cambió desde que abriste el registro; actualiza la página")

    _validate_unique(context, values, skip_row=row_number)
    _write_agent(context, row_number, values)
    output = io.BytesIO()
    context.workbook.save(output)
    return output.getvalue()


def _upload_workbook(file_id: str, workbook_bytes: bytes) -> None:
    try:
        from google.auth import default
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaIoBaseUpload
    except ImportError as exc:
        raise RuntimeError("No están instaladas las dependencias de Google Drive") from exc
    credentials, _ = default(scopes=["https://www.googleapis.com/auth/drive"])
    drive = build("drive", "v3", credentials=credentials, cache_discovery=False)
    media = MediaIoBaseUpload(
        io.BytesIO(workbook_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False,
    )
    drive.files().update(fileId=file_id, media_body=media).execute()


def _save_mutation(payload: AgentFields, version: str, *, row_number: int | None = None, fingerprint: str | None = None) -> dict:
    with _write_lock:
        current = _download_workbook(_file_id())
        if hashlib.sha256(current).hexdigest() != version:
            raise HTTPException(
                status_code=409,
                detail="La base de agentes cambió desde que abriste la pantalla; actualiza antes de guardar",
            )
        try:
            updated = mutate_agent_workbook(
                current,
                payload,
                row_number=row_number,
                fingerprint=fingerprint,
            )
        except RuntimeError as exc:
            raise HTTPException(status_code=409, detail=str(exc)) from exc
        except ValueError as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc
        try:
            _upload_workbook(_file_id(), updated)
        except Exception as exc:
            raise HTTPException(status_code=502, detail=f"No se pudo actualizar la base en Drive: {exc}") from exc
        data_cache.invalidate("cumpleanos:agentes")
        clear_agent_directory_cache()
        return updated


@router.get("")
def get_agents(
    profile: AccessProfile = Depends(require_module_access("agentes")),
):
    try:
        workbook = _download_workbook(_file_id())
        return build_agent_directory(workbook, can_operate=profile.can_operate("agentes"))
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"No se pudo leer la base de agentes: {exc}") from exc


@router.post("")
def create_agent(
    payload: CreateAgentPayload,
    profile: AccessProfile = Depends(require_module_access("agentes", operation=True)),
):
    updated = _save_mutation(payload, payload.version)
    return build_agent_directory(updated, can_operate=profile.can_operate("agentes"))


@router.patch("/{row_number}")
def update_agent(
    row_number: int,
    payload: UpdateAgentPayload,
    profile: AccessProfile = Depends(require_module_access("agentes", operation=True)),
):
    updated = _save_mutation(
        payload,
        payload.version,
        row_number=row_number,
        fingerprint=payload.fingerprint,
    )
    return build_agent_directory(updated, can_operate=profile.can_operate("agentes"))
