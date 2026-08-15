from __future__ import annotations

import io
import json
import mimetypes
import os
import re
import secrets
import threading
import time
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pandas as pd
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from pydantic import BaseModel, Field, model_validator
from sqlalchemy import or_

from adapters.metlife_quotation_portal import (
    MetLifeQuotationAnswer,
    MetLifeQuotationPortalAdapter,
    MetLifeQuotationTask,
    quotation_credentials_configured,
    quotation_result_to_dict,
)
from database import Client, SessionLocal
from drive.client import download_drive_file_bytes
from services.auth import AccessProfile
from services.authorization import current_access_profile
from services.client_email_directory import lookup_client_email
from services.client_folders import (
    FOLDER_MIME_TYPE,
    build_client_folder_drive_service,
    client_folders_parent_id,
    normalize_client_name,
    normalize_rfc,
    valid_client_rfc,
)
from services.mail_configuration import smtp_settings_for
from services.renovaciones import send_email_smtp


router = APIRouter(prefix="/cotizaciones", tags=["cotizaciones"])

QUOTES_FILE_ID_ENV = "GOOGLE_DRIVE_QUOTES_FILE_ID"
DEFAULT_QUOTES_FILE_ID = "1uP-G9GAz75SyO4nUhrJlaHDhX5zJ6vk4"
AGENTS_FILE_ID_ENV = "GOOGLE_DRIVE_AGENTS_METLIFE_FILE_ID"
DEFAULT_AGENTS_FILE_ID = "1IoeLDCQe4T3DofStiBSaI09xjX2-RSby"
PRODUCTS = {
    "GMM": ("Medicalife Familiar", "Medicalife PG", "Primordial"),
    "Vida": ("Metalife", "Totalife", "Flexilife", "Horizonte", "Temporal"),
}
INITIAL_STATUS = "Pendiente de cotización"
READY_STATUS = "Lista para cotizar"
MAX_QUOTE_DOCUMENT_BYTES = 25 * 1024 * 1024
QUOTE_DOCUMENT_COLUMNS = {
    "cotizaciones": "Cotizaciones",
    "documentos_adicionales": "Documentos Adicionales",
}
QUOTE_DOCUMENT_FOLDER_PREFIXES = {
    "cotizaciones": "Cotización",
    "documentos_adicionales": "Documentos Adicionales",
    "solicitud_datos": "Solicitud de datos",
}
HEADERS = (
    "Agente",
    "Promotoría",
    "Aseguradora",
    "Clave de agente",
    "Cliente / Prospecto",
    "RFC",
    "Ramo",
    "Producto",
    "Estatus",
    "Cotizaciones",
    "Documentos Adicionales",
    "ID",
)
_workbook_lock = threading.RLock()
_agent_cache_lock = threading.RLock()
_agent_cache: tuple[float, list[dict[str, str]]] | None = None
_quotation_browser_lock = threading.RLock()
_active_quotation_browser_quote_id: str | None = None
_data_request_lock = threading.RLock()


class QuoteCreate(BaseModel):
    client_id: str | None = None
    prospect_name: str | None = Field(default=None, max_length=255)
    ramo: str
    producto: str
    agent_rfc: str | None = None
    agent_promotoria: str | None = None

    @model_validator(mode="after")
    def validate_client_and_product(self):
        if bool(self.client_id) == bool(str(self.prospect_name or "").strip()):
            raise ValueError("Selecciona un cliente existente o captura un prospecto")
        products = PRODUCTS.get(self.ramo)
        if not products:
            raise ValueError("El ramo debe ser GMM o Vida")
        if self.producto not in products:
            raise ValueError(f"El producto no corresponde al ramo {self.ramo}")
        return self


class QuoteUpdate(BaseModel):
    cliente: str = Field(min_length=2, max_length=255)
    rfc: str | None = Field(default=None, max_length=50)
    ramo: str
    producto: str
    agent_rfc: str | None = None
    agent_promotoria: str | None = None

    @model_validator(mode="after")
    def validate_product(self):
        products = PRODUCTS.get(self.ramo)
        if not products:
            raise ValueError("El ramo debe ser GMM o Vida")
        if self.producto not in products:
            raise ValueError(f"El producto no corresponde al ramo {self.ramo}")
        return self


class QuoteStartRequest(BaseModel):
    rfc: str | None = Field(default=None, max_length=50)


class QuoteBrowserAnswerRequest(BaseModel):
    question_id: str = Field(min_length=1, max_length=80)
    option_id: str = Field(min_length=1, max_length=120)
    value: str | None = Field(default=None, max_length=255)


class QuoteEmailRequest(BaseModel):
    file_ids: list[str] = Field(min_length=1, max_length=10)
    recipients: list[str] = Field(min_length=1, max_length=20)
    subject: str = Field(min_length=3, max_length=255)
    body: str = Field(min_length=3, max_length=10000)


def _file_id() -> str:
    return os.getenv(QUOTES_FILE_ID_ENV, "").strip() or DEFAULT_QUOTES_FILE_ID


def _build_writable_drive_service():
    try:
        from google.auth import default
        from googleapiclient.discovery import build
    except ImportError as exc:
        raise RuntimeError("Faltan las dependencias de Google Drive") from exc
    credentials, _ = default(scopes=["https://www.googleapis.com/auth/drive"])
    return build("drive", "v3", credentials=credentials, cache_discovery=False)


def _upload_workbook(workbook_bytes: bytes) -> None:
    try:
        from googleapiclient.http import MediaIoBaseUpload
    except ImportError as exc:
        raise RuntimeError("Faltan las dependencias de Google Drive") from exc
    media = MediaIoBaseUpload(
        io.BytesIO(workbook_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False,
    )
    _build_writable_drive_service().files().update(
        fileId=_file_id(), media_body=media, supportsAllDrives=True
    ).execute()


def _load_workbook():
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl es necesario para administrar cotizaciones") from exc
    return load_workbook(io.BytesIO(download_drive_file_bytes(_file_id())))


def _sheet_and_headers(workbook):
    sheet = workbook.active
    headers: dict[str, int] = {}
    for index, cell in enumerate(sheet[1], start=1):
        value = str(cell.value or "").strip()
        if value:
            headers[value] = index
    for header in HEADERS:
        if header not in headers:
            column = sheet.max_column + 1
            sheet.cell(row=1, column=column, value=header)
            headers[header] = column
    return sheet, headers


def _serialize_row(sheet, headers: dict[str, int], row_number: int) -> dict[str, str]:
    stored_id = str(sheet.cell(row=row_number, column=headers["ID"]).value or "").strip()
    return {
        "id": stored_id or f"ROW-{row_number}",
        "cliente": str(sheet.cell(row=row_number, column=headers["Cliente / Prospecto"]).value or ""),
        "rfc": str(sheet.cell(row=row_number, column=headers["RFC"]).value or ""),
        "ramo": str(sheet.cell(row=row_number, column=headers["Ramo"]).value or ""),
        "producto": str(sheet.cell(row=row_number, column=headers["Producto"]).value or ""),
        "estatus": str(sheet.cell(row=row_number, column=headers["Estatus"]).value or ""),
        "cotizaciones": str(sheet.cell(row=row_number, column=headers["Cotizaciones"]).value or ""),
        "documentos_adicionales": str(sheet.cell(row=row_number, column=headers["Documentos Adicionales"]).value or ""),
        "agente": str(sheet.cell(row=row_number, column=headers["Agente"]).value or ""),
        "promotoria": str(sheet.cell(row=row_number, column=headers["Promotoría"]).value or ""),
        "aseguradora": str(sheet.cell(row=row_number, column=headers["Aseguradora"]).value or ""),
        "clave_agente": str(sheet.cell(row=row_number, column=headers["Clave de agente"]).value or ""),
    }


def _quote_row_by_id(sheet, headers: dict[str, int], quote_id: str) -> int:
    normalized_id = str(quote_id or "").strip()
    if normalized_id.startswith("ROW-"):
        try:
            candidate = int(normalized_id.removeprefix("ROW-"))
        except ValueError:
            candidate = 0
        if 2 <= candidate <= sheet.max_row:
            return candidate
    for candidate in range(2, sheet.max_row + 1):
        value = str(sheet.cell(row=candidate, column=headers["ID"]).value or "").strip()
        if value == normalized_id:
            return candidate
    raise KeyError("Cotización no encontrada")


def _drive_query_literal(value: str) -> str:
    return value.replace("\\", "\\\\").replace("'", "\\'")


def _folder_link(folder: dict[str, str]) -> str:
    link = str(folder.get("webViewLink") or "").strip()
    if link:
        return link
    folder_id = str(folder.get("id") or "").strip()
    return f"https://drive.google.com/drive/folders/{folder_id}" if folder_id else ""


def _folder_id_from_link(value: str) -> str:
    text = str(value or "").strip()
    patterns = (
        r"/folders/([A-Za-z0-9_-]+)",
        r"[?&]id=([A-Za-z0-9_-]+)",
    )
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1)
    return ""


def _safe_upload_filename(filename: str | None, fallback: str = "documento") -> str:
    raw = str(filename or "").strip() or fallback
    raw = raw.replace("/", "-").replace("\\", "-")
    cleaned = re.sub(r"\s+", " ", raw).strip()
    return cleaned[:180] or fallback


def _quote_record_date(quote_id: str) -> str:
    match = re.search(r"COT-(\d{4})(\d{2})(\d{2})-", str(quote_id or ""))
    if match:
        year, month, day = match.groups()
        return f"{day}/{month}/{year[-2:]}"
    return f"{datetime.now(timezone.utc):%d/%m/%y}"


def _quote_document_folder_name(document_kind: str, product: str, quote_id: str) -> str:
    prefix = QUOTE_DOCUMENT_FOLDER_PREFIXES[document_kind]
    product_name = re.sub(r"\s+", " ", str(product or "").strip()) or "Producto"
    product_name = product_name.replace("/", "-").replace("\\", "-")
    return f"{prefix}-{product_name}-{_quote_record_date(quote_id)}"


def quote_client_folder_name(rfc: str, client_name: str) -> str:
    normalized_rfc = normalize_rfc(rfc)
    normalized_name = normalize_client_name(client_name, normalized_rfc)
    if not normalized_name:
        normalized_name = "Cliente sin nombre"
    return f"{normalized_rfc} - {normalized_name}"


def find_or_create_quote_client_folder(
    service,
    *,
    rfc: str,
    client_name: str,
    parent_id: str | None = None,
) -> dict[str, str]:
    normalized_rfc = normalize_rfc(rfc)
    if not valid_client_rfc(normalized_rfc):
        raise ValueError("Captura un RFC válido antes de iniciar la cotización")
    parent = (parent_id or client_folders_parent_id()).strip()
    expected_name = quote_client_folder_name(normalized_rfc, client_name)
    response = service.files().list(
        q=(
            f"'{_drive_query_literal(parent)}' in parents and "
            f"mimeType = '{FOLDER_MIME_TYPE}' and trashed = false"
        ),
        spaces="drive",
        fields="files(id,name,mimeType,webViewLink)",
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
        pageSize=1000,
    ).execute()
    folders = [
        item for item in response.get("files", [])
        if normalize_rfc(str(item.get("name", "")).split(" - ", 1)[0]) == normalized_rfc
    ]
    if folders:
        return folders[0]
    return service.files().create(
        body={
            "name": expected_name,
            "mimeType": FOLDER_MIME_TYPE,
            "parents": [parent],
        },
        fields="id,name,mimeType,webViewLink",
        supportsAllDrives=True,
    ).execute()


def _drive_folder_metadata(service, folder_id: str) -> dict[str, object]:
    return service.files().get(
        fileId=folder_id,
        fields="id,name,webViewLink,parents",
        supportsAllDrives=True,
    ).execute()


def _is_quote_document_subfolder_name(name: str) -> bool:
    return any(
        str(name or "").startswith(f"{prefix}-")
        for prefix in QUOTE_DOCUMENT_FOLDER_PREFIXES.values()
    )


def _resolve_quote_client_folder(service, row: dict[str, str], rfc: str) -> dict[str, str]:
    for link in (row["cotizaciones"], row["documentos_adicionales"]):
        folder_id = _folder_id_from_link(link)
        if not folder_id:
            continue
        try:
            folder = _drive_folder_metadata(service, folder_id)
        except Exception:
            continue
        name = str(folder.get("name") or "")
        if _is_quote_document_subfolder_name(name):
            parents = folder.get("parents") or []
            if parents:
                try:
                    parent = _drive_folder_metadata(service, str(parents[0]))
                    return {
                        "id": str(parent.get("id") or parents[0]),
                        "name": str(parent.get("name") or ""),
                        "webViewLink": str(parent.get("webViewLink") or f"https://drive.google.com/drive/folders/{parents[0]}"),
                    }
                except Exception:
                    return {"id": str(parents[0]), "name": "", "webViewLink": f"https://drive.google.com/drive/folders/{parents[0]}"}
        return {
            "id": str(folder.get("id") or folder_id),
            "name": name,
            "webViewLink": str(folder.get("webViewLink") or f"https://drive.google.com/drive/folders/{folder_id}"),
        }
    return find_or_create_quote_client_folder(service, rfc=rfc, client_name=row["cliente"])


def find_or_create_quote_document_subfolder(
    service,
    *,
    parent_id: str,
    folder_name: str,
) -> dict[str, str]:
    response = service.files().list(
        q=(
            f"'{_drive_query_literal(parent_id)}' in parents and "
            f"mimeType = '{FOLDER_MIME_TYPE}' and trashed = false and "
            f"name = '{_drive_query_literal(folder_name)}'"
        ),
        spaces="drive",
        fields="files(id,name,mimeType,webViewLink)",
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
        pageSize=10,
    ).execute()
    folders = response.get("files", [])
    if folders:
        return folders[0]
    return service.files().create(
        body={
            "name": folder_name,
            "mimeType": FOLDER_MIME_TYPE,
            "parents": [parent_id],
        },
        fields="id,name,mimeType,webViewLink",
        supportsAllDrives=True,
    ).execute()


def _list_drive_files_in_folder(service, folder_id: str) -> list[dict[str, str]]:
    if not folder_id:
        return []
    response = service.files().list(
        q=(
            f"'{_drive_query_literal(folder_id)}' in parents and "
            "trashed = false and "
            f"mimeType != '{FOLDER_MIME_TYPE}'"
        ),
        spaces="drive",
        fields="files(id,name,mimeType,webViewLink,modifiedTime,size)",
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
        pageSize=1000,
        orderBy="modifiedTime desc",
    ).execute()
    return [
        {
            "id": str(item.get("id") or ""),
            "name": str(item.get("name") or ""),
            "mimeType": str(item.get("mimeType") or "application/octet-stream"),
            "webViewLink": str(item.get("webViewLink") or ""),
            "modifiedTime": str(item.get("modifiedTime") or ""),
            "size": str(item.get("size") or ""),
        }
        for item in response.get("files", [])
        if item.get("id")
    ]


def _data_request_store_path() -> Path:
    path = Path(__file__).resolve().parents[1] / "local-data" / "quote_data_requests.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def _load_data_requests() -> dict[str, dict[str, object]]:
    path = _data_request_store_path()
    if not path.exists():
        return {}
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    if not isinstance(raw, dict):
        return {}
    return {str(key): value for key, value in raw.items() if isinstance(value, dict)}


def _save_data_requests(records: dict[str, dict[str, object]]) -> None:
    path = _data_request_store_path()
    path.write_text(json.dumps(records, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")


def _parse_iso_datetime(value: object) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    except ValueError:
        return None


def _public_quote_payload(row: dict[str, str]) -> dict[str, str]:
    return {
        "id": row["id"],
        "cliente": row["cliente"],
        "rfc": row["rfc"],
        "ramo": row["ramo"],
        "producto": row["producto"],
        "estatus": row["estatus"],
        "agente": row["agente"],
        "promotoria": row["promotoria"],
        "aseguradora": row["aseguradora"],
    }


def create_quote_data_request_link(quote_id: str, profile: AccessProfile) -> dict[str, object]:
    row = _quote_by_id(quote_id)
    rfc = normalize_rfc(row.get("rfc"))
    if not valid_client_rfc(rfc):
        raise ValueError("Captura un RFC válido antes de solicitar datos")
    now = datetime.now(timezone.utc)
    token = secrets.token_urlsafe(32)
    expires_at = now + timedelta(hours=24)
    record = {
        "token": token,
        "quote_id": row["id"],
        "created_at": now.isoformat(),
        "expires_at": expires_at.isoformat(),
        "created_by": profile.username,
        "used_at": None,
        "folder_id": None,
        "folder_link": None,
    }
    with _data_request_lock:
        records = _load_data_requests()
        records[token] = record
        _save_data_requests(records)
    return {
        "token": token,
        "path": f"/solicitud-datos/{token}",
        "expires_at": expires_at.isoformat(),
        "quote": _public_quote_payload(row),
    }


def _active_data_request(token: str) -> tuple[dict[str, object], dict[str, str]]:
    with _data_request_lock:
        records = _load_data_requests()
        record = records.get(token)
    if not record:
        raise KeyError("Liga no encontrada")
    expires_at = _parse_iso_datetime(record.get("expires_at"))
    if not expires_at or expires_at <= datetime.now(timezone.utc):
        raise ValueError("La liga expiró")
    if record.get("used_at"):
        raise ValueError("La solicitud ya fue enviada")
    return record, _quote_by_id(str(record.get("quote_id") or ""))


def get_quote_data_request(token: str) -> dict[str, object]:
    with _data_request_lock:
        records = _load_data_requests()
        record = records.get(token)
    if not record:
        raise KeyError("Liga no encontrada")
    row = _quote_by_id(str(record.get("quote_id") or ""))
    expires_at = _parse_iso_datetime(record.get("expires_at"))
    expired = not expires_at or expires_at <= datetime.now(timezone.utc)
    submitted = bool(record.get("used_at"))
    return {
        "token": token,
        "expires_at": record.get("expires_at"),
        "expired": expired,
        "submitted": submitted,
        "quote": _public_quote_payload(row),
    }


def _upload_bytes_to_drive(service, *, folder_id: str, filename: str, content: bytes, mime_type: str) -> dict[str, str]:
    try:
        from googleapiclient.http import MediaIoBaseUpload
    except ImportError as exc:
        raise RuntimeError("Faltan las dependencias de Google Drive") from exc
    media = MediaIoBaseUpload(io.BytesIO(content), mimetype=mime_type, resumable=False)
    return service.files().create(
        body={"name": _safe_upload_filename(filename), "parents": [folder_id]},
        media_body=media,
        fields="id,name,mimeType,webViewLink",
        supportsAllDrives=True,
    ).execute()


def _notify_data_request_submission(row: dict[str, str], record: dict[str, object], folder_link: str) -> tuple[bool, str | None]:
    recipients = []
    seen: set[str] = set()
    for candidate in (str(record.get("created_by") or ""), "alberto.alfaro@taiico.com"):
        email = candidate.strip().casefold()
        if "@" in email and email not in seen:
            recipients.append(email)
            seen.add(email)
    if not recipients:
        return False, "No hay destinatarios válidos para notificar"

    sender = str(record.get("created_by") or "").strip()
    settings = smtp_settings_for(sender) if sender else None
    if not settings:
        settings = smtp_settings_for("alberto.alfaro@taiico.com")
    if not settings:
        return False, "No hay configuración SMTP disponible para enviar la notificación"

    subject = f"Datos recibidos para emisión - {row.get('cliente') or row.get('rfc')}"
    body = (
        "El prospecto/cliente ya proporcionó sus datos para continuar con la emisión.\n\n"
        f"Cliente: {row.get('cliente') or ''}\n"
        f"RFC: {row.get('rfc') or ''}\n"
        f"Producto: {row.get('producto') or ''}\n"
        f"Carpeta de solicitud: {folder_link}\n"
    )
    send_email_smtp(subject, body, recipients, [], cc_recipients=[], settings=settings)
    return True, None


async def submit_quote_data_request(token: str, payload: str, documents: list[UploadFile] | None) -> dict[str, object]:
    record, row = _active_data_request(token)
    try:
        data = json.loads(payload)
    except json.JSONDecodeError as exc:
        raise ValueError("Los datos enviados no tienen formato JSON válido") from exc
    if not isinstance(data, dict):
        raise ValueError("Los datos enviados deben ser un objeto JSON")

    rfc = normalize_rfc(row.get("rfc"))
    if not valid_client_rfc(rfc):
        raise ValueError("La cotización no tiene un RFC válido")

    service = build_client_folder_drive_service()
    client_folder = _resolve_quote_client_folder(service, row, rfc)
    request_folder = find_or_create_quote_document_subfolder(
        service,
        parent_id=str(client_folder["id"]),
        folder_name=_quote_document_folder_name("solicitud_datos", row.get("producto") or "", row.get("id") or ""),
    )
    folder_id = str(request_folder["id"])
    folder_link = _folder_link(request_folder)

    submitted_at = datetime.now(timezone.utc).isoformat()
    json_payload = {
        "quote": _public_quote_payload(row),
        "submitted_at": submitted_at,
        "data": data,
    }
    json_name = f"datos-solicitud-{rfc}-{datetime.now(timezone.utc):%Y%m%d-%H%M%S}.json"
    uploaded_files = [
        _upload_bytes_to_drive(
            service,
            folder_id=folder_id,
            filename=json_name,
            content=json.dumps(json_payload, ensure_ascii=False, indent=2).encode("utf-8"),
            mime_type="application/json",
        )
    ]

    for index, document in enumerate(documents or [], start=1):
        content = await document.read(MAX_QUOTE_DOCUMENT_BYTES + 1)
        if not content:
            continue
        if len(content) > MAX_QUOTE_DOCUMENT_BYTES:
            raise ValueError(f"El archivo {document.filename or index} supera el límite de 25 MB")
        uploaded_files.append(
            _upload_bytes_to_drive(
                service,
                folder_id=folder_id,
                filename=f"{index:02d}-{_safe_upload_filename(document.filename, 'documento')}",
                content=content,
                mime_type=document.content_type or mimetypes.guess_type(document.filename or "")[0] or "application/octet-stream",
            )
        )

    with _data_request_lock:
        records = _load_data_requests()
        stored = records.get(token, {})
        stored.update(
            {
                "used_at": submitted_at,
                "folder_id": folder_id,
                "folder_link": folder_link,
                "uploaded_files": [str(item.get("name") or "") for item in uploaded_files],
            }
        )
        records[token] = stored
        _save_data_requests(records)

    notification_sent, notification_warning = _notify_data_request_submission(row, record, folder_link)
    return {
        "submitted": True,
        "folder_link": folder_link,
        "uploaded_count": len(uploaded_files),
        "notification_sent": notification_sent,
        "notification_warning": notification_warning,
    }


def _quote_email_body(row: dict[str, str]) -> str:
    agent_name = row.get("agente") or "TAIICO"
    return (
        f"Hola {row.get('cliente') or ''},\n\n"
        f"Te comparto adjunta la cotización de {row.get('producto') or 'tu seguro'} para tu revisión.\n\n"
        "Quedo atento(a) a tus comentarios o a cualquier ajuste que quieras que realicemos.\n\n"
        f"Saludos,\n{agent_name}"
    )


def _quote_email_subject(row: dict[str, str]) -> str:
    product = row.get("producto") or "seguro"
    client = row.get("cliente") or "cliente"
    return f"Cotización {product} - {client}"


def _lookup_quote_client_email(row: dict[str, str]) -> str | None:
    normalized_rfc = normalize_rfc(row.get("rfc"))
    db = SessionLocal()
    try:
        if normalized_rfc:
            client = (
                db.query(Client)
                .filter(Client.rfc.ilike(normalized_rfc))
                .first()
            )
            if client and str(client.email or "").strip():
                return str(client.email).strip().casefold()
        normalized_name = " ".join(str(row.get("cliente") or "").split())
        if normalized_name:
            client = (
                db.query(Client)
                .filter(Client.full_name.ilike(normalized_name))
                .first()
            )
            if client and str(client.email or "").strip():
                return str(client.email).strip().casefold()
    finally:
        db.close()
    try:
        return lookup_client_email(row.get("cliente"))
    except Exception:
        return None


def _quote_by_id(quote_id: str) -> dict[str, str]:
    workbook = _load_workbook()
    sheet, headers = _sheet_and_headers(workbook)
    row_number = _quote_row_by_id(sheet, headers, quote_id)
    return _serialize_row(sheet, headers, row_number)


def _quote_document_folder_for_listing(
    service,
    row: dict[str, str],
    document_kind: str = "cotizaciones",
) -> dict[str, str] | None:
    folder_id = _folder_id_from_link(row.get(document_kind, ""))
    if folder_id:
        try:
            return _drive_folder_metadata(service, folder_id)
        except Exception:
            pass
    rfc = normalize_rfc(row.get("rfc"))
    if not valid_client_rfc(rfc):
        return None
    client_folder = _resolve_quote_client_folder(service, row, rfc)
    folder_name = _quote_document_folder_name(document_kind, row["producto"], row["id"])
    return find_or_create_quote_document_subfolder(
        service,
        parent_id=str(client_folder.get("id") or ""),
        folder_name=folder_name,
    )


def quote_email_draft(quote_id: str) -> dict[str, object]:
    row = _quote_by_id(quote_id)
    service = build_client_folder_drive_service()
    folder = _quote_document_folder_for_listing(service, row, "cotizaciones")
    folder_id = str((folder or {}).get("id") or "")
    email = _lookup_quote_client_email(row)
    return {
        "quote": row,
        "files": _list_drive_files_in_folder(service, folder_id),
        "folder_link": _folder_link(folder or {}) if folder else "",
        "default_recipients": [email] if email else [],
        "default_subject": _quote_email_subject(row),
        "default_body": _quote_email_body(row),
    }


def _download_drive_file_attachment(service, file_id: str, allowed: dict[str, dict[str, str]]) -> dict[str, object]:
    file_info = allowed.get(file_id)
    if not file_info:
        raise ValueError("Uno de los archivos seleccionados no pertenece a la carpeta de cotizaciones")
    try:
        from googleapiclient.http import MediaIoBaseDownload
    except ImportError as exc:
        raise RuntimeError("Faltan las dependencias de Google Drive") from exc
    output = io.BytesIO()
    request = service.files().get_media(fileId=file_id, supportsAllDrives=True)
    downloader = MediaIoBaseDownload(output, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    name = file_info["name"]
    return {
        "name": name,
        "content": output.getvalue(),
        "mime_type": file_info.get("mimeType") or mimetypes.guess_type(name)[0] or "application/octet-stream",
    }


def send_quote_email(quote_id: str, payload: QuoteEmailRequest, profile: AccessProfile) -> dict[str, object]:
    recipients = []
    seen: set[str] = set()
    for item in payload.recipients:
        email = str(item or "").strip().casefold()
        if not email:
            continue
        if "@" not in email or email.startswith("@") or email.endswith("@"):
            raise ValueError(f"Destinatario inválido: {item}")
        if email not in seen:
            recipients.append(email)
            seen.add(email)
    if not recipients:
        raise ValueError("Captura al menos un destinatario")

    row = _quote_by_id(quote_id)
    service = build_client_folder_drive_service()
    folder = _quote_document_folder_for_listing(service, row, "cotizaciones")
    folder_id = str((folder or {}).get("id") or "")
    allowed_files = {item["id"]: item for item in _list_drive_files_in_folder(service, folder_id)}
    attachments = [
        _download_drive_file_attachment(service, file_id, allowed_files)
        for file_id in payload.file_ids
    ]
    if not attachments:
        raise ValueError("Selecciona al menos un archivo de cotización")

    settings = smtp_settings_for(profile.username)
    if not settings:
        raise ValueError("Configura primero tu cuenta de correo en Configuración de Mail")
    send_email_smtp(
        payload.subject.strip(),
        payload.body.strip(),
        recipients,
        attachments,
        cc_recipients=[],
        settings=settings,
    )
    return {
        "sent": True,
        "recipients": recipients,
        "attachment_count": len(attachments),
    }


async def upload_quote_document(quote_id: str, document_kind: str, document: UploadFile) -> dict[str, object]:
    column_name = QUOTE_DOCUMENT_COLUMNS.get(document_kind)
    if not column_name:
        raise ValueError("Tipo de documento no soportado")

    try:
        from googleapiclient.http import MediaIoBaseUpload
    except ImportError as exc:
        raise RuntimeError("Faltan las dependencias de Google Drive") from exc

    content = await document.read(MAX_QUOTE_DOCUMENT_BYTES + 1)
    if not content:
        raise ValueError("El archivo está vacío")
    if len(content) > MAX_QUOTE_DOCUMENT_BYTES:
        raise ValueError("El archivo supera el límite de 25 MB")

    with _workbook_lock:
        workbook = _load_workbook()
        sheet, headers = _sheet_and_headers(workbook)
        row_number = _quote_row_by_id(sheet, headers, quote_id)
        row = _serialize_row(sheet, headers, row_number)
        rfc = normalize_rfc(row["rfc"])
        if not valid_client_rfc(rfc):
            raise ValueError("Captura un RFC válido antes de cargar documentos")

        service = build_client_folder_drive_service()
        client_folder = _resolve_quote_client_folder(service, row, rfc)
        client_folder_id = str(client_folder.get("id") or "")
        if not client_folder_id:
            raise RuntimeError("No se pudo resolver la carpeta del cliente")

        subfolder_name = _quote_document_folder_name(document_kind, row["producto"], row["id"])
        destination_folder = find_or_create_quote_document_subfolder(
            service,
            parent_id=client_folder_id,
            folder_name=subfolder_name,
        )
        folder_id = str(destination_folder.get("id") or "")
        folder_link = _folder_link(destination_folder)
        if not folder_id:
            raise RuntimeError("No se pudo resolver la subcarpeta destino")

        final_name = _safe_upload_filename(document.filename)
        existing = service.files().list(
            q=f"'{_drive_query_literal(folder_id)}' in parents and trashed = false",
            fields="files(id,name)",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
            pageSize=1000,
        ).execute().get("files", [])
        existing_document = next(
            (item for item in existing if str(item.get("name") or "").casefold() == final_name.casefold()),
            None,
        )
        media = MediaIoBaseUpload(
            io.BytesIO(content),
            mimetype=document.content_type or "application/octet-stream",
            resumable=False,
        )
        if existing_document:
            saved = service.files().update(
                fileId=existing_document["id"],
                media_body=media,
                fields="id,name,mimeType,webViewLink,modifiedTime,size",
                supportsAllDrives=True,
            ).execute()
        else:
            saved = service.files().create(
                body={"name": final_name, "parents": [folder_id]},
                media_body=media,
                fields="id,name,mimeType,webViewLink,modifiedTime,size",
                supportsAllDrives=True,
            ).execute()

        sheet.cell(row=row_number, column=headers[column_name], value=folder_link)
        output = io.BytesIO()
        workbook.save(output)
        _upload_workbook(output.getvalue())
        return {
            "uploaded": True,
            "replaced": bool(existing_document),
            "document": saved,
            "quote": _serialize_row(sheet, headers, row_number),
        }


def parse_agent_directory(workbook_bytes: bytes) -> list[dict[str, str]]:
    excel = pd.ExcelFile(io.BytesIO(workbook_bytes))
    sheet_name = "Datos" if "Datos" in excel.sheet_names else excel.sheet_names[0]
    table = pd.read_excel(
        io.BytesIO(workbook_bytes),
        sheet_name=sheet_name,
        dtype=str,
        keep_default_na=False,
    )
    required = {"RFC", "Promotoria", "CLAVE_DEFINITIVA", "CLAVE_ARRANQUE"}
    missing = sorted(required.difference(table.columns))
    if missing:
        raise ValueError("La base de agentes no contiene: " + ", ".join(missing))

    agents: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for _, row in table.iterrows():
        rfc = "".join(str(row.get("RFC") or "").upper().split())
        promotoria = str(row.get("Promotoria") or "").strip().upper()
        definitive_key = str(row.get("CLAVE_DEFINITIVA") or "").strip()
        start_key = str(row.get("CLAVE_ARRANQUE") or "").strip()
        key = definitive_key or start_key
        if not rfc or not promotoria or not key or (rfc, promotoria) in seen:
            continue
        parts = [
            str(row.get("Nombres") or "").strip(),
            str(row.get("Apellido_Paterno") or "").strip(),
            str(row.get("Apellido_Materno") or "").strip(),
        ]
        name = " ".join(part for part in parts if part)
        if not name:
            name = str(row.get("Nombre") or "").strip()
        agents.append(
            {
                "rfc": rfc,
                "name": name.title() or "Nombre no registrado",
                "promotoria": promotoria,
                "key": key,
                "key_source": "CLAVE_DEFINITIVA" if definitive_key else "CLAVE_ARRANQUE",
            }
        )
        seen.add((rfc, promotoria))
    return sorted(agents, key=lambda item: (item["promotoria"], item["name"], item["rfc"]))


def load_agent_directory() -> list[dict[str, str]]:
    global _agent_cache
    now = time.monotonic()
    cache_seconds = max(0, int(os.getenv("QUOTES_AGENTS_CACHE_SECONDS", "300")))
    with _agent_cache_lock:
        if _agent_cache and now < _agent_cache[0]:
            return _agent_cache[1]
        file_id = os.getenv(AGENTS_FILE_ID_ENV, "").strip() or DEFAULT_AGENTS_FILE_ID
        agents = parse_agent_directory(download_drive_file_bytes(file_id))
        _agent_cache = (now + cache_seconds, agents)
        return agents


def agents_for_profile(profile: AccessProfile) -> list[dict[str, str]]:
    allowed_promotorias = set(profile.promotorias)
    agents = [agent for agent in load_agent_directory() if agent["promotoria"] in allowed_promotorias]
    if profile.is_agent:
        profile_rfc = "".join(profile.rfc.upper().split())
        agents = [agent for agent in agents if agent["rfc"] == profile_rfc]
    return agents


def assigned_agent(
    profile: AccessProfile,
    requested_rfc: str | None,
    requested_promotoria: str | None = None,
) -> dict[str, str]:
    options = agents_for_profile(profile)
    if profile.is_agent:
        if not profile.rfc:
            raise ValueError("Tu usuario no tiene RFC de agente configurado en Accesos")
        if len(options) != 1:
            raise ValueError("No fue posible identificar una única clave MetLife para tu usuario")
        return options[0]
    selected = "".join(str(requested_rfc or "").upper().split())
    selected_promotoria = str(requested_promotoria or "").strip().upper()
    matches = [
        agent for agent in options
        if agent["rfc"] == selected and agent["promotoria"] == selected_promotoria
    ]
    if not selected:
        raise ValueError("Selecciona el agente responsable de la cotización")
    if len(matches) != 1:
        raise ValueError("El agente no pertenece a una promotoría autorizada para tu usuario")
    return matches[0]


def list_quotes() -> list[dict[str, str]]:
    workbook = _load_workbook()
    sheet, headers = _sheet_and_headers(workbook)
    rows = []
    for row_number in range(2, sheet.max_row + 1):
        row = _serialize_row(sheet, headers, row_number)
        if any(row.values()):
            rows.append(row)
    return rows


def create_quote(payload: QuoteCreate, profile: AccessProfile) -> dict[str, str]:
    agent = assigned_agent(profile, payload.agent_rfc, payload.agent_promotoria)
    db = SessionLocal()
    try:
        if payload.client_id:
            client = db.query(Client).filter(Client.id == payload.client_id).first()
            if not client:
                raise ValueError("El cliente seleccionado ya no existe")
            name = client.full_name.strip()
            rfc = "".join(str(client.rfc or "").upper().split())
        else:
            name = " ".join(str(payload.prospect_name or "").split())
            rfc = ""
            if len(name) < 2:
                raise ValueError("Captura el nombre del prospecto")
    finally:
        db.close()

    quote_id = f"COT-{datetime.now(timezone.utc):%Y%m%d}-{uuid.uuid4().hex[:6].upper()}"
    values = {
        "ID": quote_id,
        "Agente": agent["name"],
        "Promotoría": agent["promotoria"],
        "Aseguradora": "MetLife",
        "Clave de agente": agent["key"],
        "Cliente / Prospecto": name,
        "RFC": rfc,
        "Ramo": payload.ramo,
        "Producto": payload.producto,
        "Estatus": INITIAL_STATUS,
        "Cotizaciones": "",
        "Documentos Adicionales": "",
    }
    with _workbook_lock:
        workbook = _load_workbook()
        sheet, headers = _sheet_and_headers(workbook)
        row_number = sheet.max_row + 1
        for header, value in values.items():
            sheet.cell(row=row_number, column=headers[header], value=value)
        output = io.BytesIO()
        workbook.save(output)
        _upload_workbook(output.getvalue())
        return _serialize_row(sheet, headers, row_number)


def update_quote(quote_id: str, payload: QuoteUpdate, profile: AccessProfile) -> dict[str, str]:
    agent = assigned_agent(profile, payload.agent_rfc, payload.agent_promotoria)
    normalized_id = str(quote_id or "").strip()
    with _workbook_lock:
        workbook = _load_workbook()
        sheet, headers = _sheet_and_headers(workbook)
        row_number = _quote_row_by_id(sheet, headers, normalized_id)

        stored_id = str(sheet.cell(row=row_number, column=headers["ID"]).value or "").strip()
        if not stored_id:
            stored_id = f"COT-{datetime.now(timezone.utc):%Y%m%d}-{uuid.uuid4().hex[:6].upper()}"
        values = {
            "ID": stored_id,
            "Agente": agent["name"],
            "Promotoría": agent["promotoria"],
            "Aseguradora": "MetLife",
            "Clave de agente": agent["key"],
            "Cliente / Prospecto": " ".join(payload.cliente.split()),
            "RFC": "".join(str(payload.rfc or "").upper().split()),
            "Ramo": payload.ramo,
            "Producto": payload.producto,
        }
        for header, value in values.items():
            sheet.cell(row=row_number, column=headers[header], value=value)
        output = io.BytesIO()
        workbook.save(output)
        _upload_workbook(output.getvalue())
        return _serialize_row(sheet, headers, row_number)


def start_quote(quote_id: str, payload: QuoteStartRequest) -> dict[str, str]:
    with _workbook_lock:
        workbook = _load_workbook()
        sheet, headers = _sheet_and_headers(workbook)
        row_number = _quote_row_by_id(sheet, headers, quote_id)
        row = _serialize_row(sheet, headers, row_number)
        rfc = normalize_rfc(payload.rfc or row["rfc"])
        if not valid_client_rfc(rfc):
            raise ValueError("Captura un RFC válido antes de iniciar la cotización")
        client_name = row["cliente"].strip()
        service = build_client_folder_drive_service()
        folder = find_or_create_quote_client_folder(
            service,
            rfc=rfc,
            client_name=client_name,
        )
        sheet.cell(row=row_number, column=headers["RFC"], value=rfc)
        sheet.cell(row=row_number, column=headers["Cotizaciones"], value=_folder_link(folder))
        if headers.get("Documentos Adicionales"):
            sheet.cell(row=row_number, column=headers["Documentos Adicionales"], value=_folder_link(folder))
        sheet.cell(row=row_number, column=headers["Estatus"], value=READY_STATUS)
        output = io.BytesIO()
        workbook.save(output)
        _upload_workbook(output.getvalue())
        updated = _serialize_row(sheet, headers, row_number)
        updated["folder_id"] = str(folder.get("id") or "")
        updated["folder_name"] = str(folder.get("name") or "")
        return updated


def quotation_portal_credentials_configured() -> bool:
    return quotation_credentials_configured()


def quote_browser_task(quote_id: str) -> MetLifeQuotationTask:
    workbook = _load_workbook()
    sheet, headers = _sheet_and_headers(workbook)
    row_number = _quote_row_by_id(sheet, headers, quote_id)
    row = _serialize_row(sheet, headers, row_number)
    rfc = normalize_rfc(row["rfc"])
    if not valid_client_rfc(rfc):
        raise ValueError("Captura un RFC válido antes de abrir el agente cotizador")
    if row["estatus"] != READY_STATUS:
        raise ValueError("Primero prepara la cotización para crear/validar la carpeta del cliente")
    return MetLifeQuotationTask(
        id=row["id"],
        rfc=rfc,
        client_name=row["cliente"],
        branch=row["ramo"],
        product=row["producto"],
        agent_name=row["agente"],
        drive_folder_link=row["cotizaciones"] or row["documentos_adicionales"] or None,
    )


def reserve_quote_browser_session(quote_id: str) -> dict[str, object] | None:
    global _active_quotation_browser_quote_id
    with _quotation_browser_lock:
        if _active_quotation_browser_quote_id and _active_quotation_browser_quote_id != quote_id:
            return {
                "status": "busy",
                "task_id": quote_id,
                "rfc": "",
                "current_url": None,
                "steps": [],
                "error_message": (
                    f"La cotización {_active_quotation_browser_quote_id} ya tiene una sesión de MetLife activa. "
                    "Ciérrala antes de iniciar otra."
                ),
            }
        _active_quotation_browser_quote_id = quote_id
    return None


def open_quote_browser_session(quote_id: str) -> dict[str, object]:
    busy = reserve_quote_browser_session(quote_id)
    if busy:
        return busy
    task = quote_browser_task(quote_id)
    result = MetLifeQuotationPortalAdapter().run(task)
    result_payload = quotation_result_to_dict(result)
    if result_payload.get("status") in {"failed", "busy"}:
        with _quotation_browser_lock:
            if _active_quotation_browser_quote_id == quote_id:
                _active_quotation_browser_quote_id = None
    return result_payload


def answer_quote_browser_session(quote_id: str, payload: QuoteBrowserAnswerRequest) -> dict[str, object]:
    busy = reserve_quote_browser_session(quote_id)
    if busy:
        return busy
    task = quote_browser_task(quote_id)
    result = MetLifeQuotationPortalAdapter().continue_with_answer(
        task,
        MetLifeQuotationAnswer(
            question_id=payload.question_id,
            option_id=payload.option_id,
            value=payload.value,
        ),
    )
    result_payload = quotation_result_to_dict(result)
    if result_payload.get("status") in {"failed", "busy"}:
        with _quotation_browser_lock:
            if _active_quotation_browser_quote_id == quote_id:
                _active_quotation_browser_quote_id = None
    return result_payload


def release_quote_browser_session(quote_id: str) -> dict[str, object]:
    global _active_quotation_browser_quote_id
    with _quotation_browser_lock:
        released = _active_quotation_browser_quote_id == quote_id
        if released:
            _active_quotation_browser_quote_id = None
    return {"released": released, "active_quote_id": _active_quotation_browser_quote_id}


@router.get("/config")
def get_quotes_config(profile: AccessProfile = Depends(current_access_profile)):
    agents = agents_for_profile(profile)
    return {
        "products": PRODUCTS,
        "initial_status": INITIAL_STATUS,
        "insurer": "MetLife",
        "agents": agents,
        "agent_is_automatic": profile.is_agent,
        "quotation_portal_credentials_configured": quotation_portal_credentials_configured(),
    }


@router.get("")
def get_quotes():
    try:
        return {"quotes": list_quotes()}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"No se pudo leer Cotizaciones.xlsx: {exc}") from exc


@router.get("/clients")
def search_clients(q: str = Query(min_length=2, max_length=120)):
    term = q.strip()
    normalized_rfc = "".join(term.upper().split())
    db = SessionLocal()
    try:
        clients = (
            db.query(Client)
            .filter(
                or_(
                    Client.full_name.ilike(f"%{term}%"),
                    Client.rfc.ilike(f"%{normalized_rfc}%"),
                )
            )
            .order_by(Client.full_name)
            .limit(20)
            .all()
        )
        return {
            "clients": [
                {"id": client.id, "nombre": client.full_name, "rfc": client.rfc or ""}
                for client in clients
            ]
        }
    finally:
        db.close()


@router.get("/public/data-requests/{token}")
def get_quote_data_request_route(token: str):
    try:
        return get_quote_data_request(token)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"No fue posible leer la solicitud: {exc}") from exc


@router.post("/public/data-requests/{token}")
async def submit_quote_data_request_route(
    token: str,
    payload: str = Form(...),
    documents: list[UploadFile] | None = File(default=None),
):
    try:
        return await submit_quote_data_request(token, payload, documents)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"No fue posible guardar la solicitud: {exc}") from exc
    finally:
        for document in documents or []:
            await document.close()


@router.post("", status_code=201)
def add_quote(
    payload: QuoteCreate,
    profile: AccessProfile = Depends(current_access_profile),
):
    try:
        return {"quote": create_quote(payload, profile)}
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"No se pudo registrar la cotización: {exc}") from exc


@router.put("/{quote_id}")
def edit_quote(
    quote_id: str,
    payload: QuoteUpdate,
    profile: AccessProfile = Depends(current_access_profile),
):
    try:
        return {"quote": update_quote(quote_id, payload, profile)}
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"No se pudo actualizar la cotización: {exc}") from exc


@router.post("/{quote_id}/data-request-link")
def create_quote_data_request_link_route(
    quote_id: str,
    profile: AccessProfile = Depends(current_access_profile),
):
    try:
        return create_quote_data_request_link(quote_id, profile)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"No fue posible crear la liga de solicitud de datos: {exc}") from exc


@router.post("/{quote_id}/start")
def begin_quote(
    quote_id: str,
    payload: QuoteStartRequest,
    profile: AccessProfile = Depends(current_access_profile),
):
    try:
        return {"quote": start_quote(quote_id, payload)}
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"No se pudo iniciar la cotización: {exc}") from exc


@router.post("/{quote_id}/browser-session")
def open_browser_session(
    quote_id: str,
    profile: AccessProfile = Depends(current_access_profile),
):
    try:
        return {"session": open_quote_browser_session(quote_id)}
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        with _quotation_browser_lock:
            active_quote_id = _active_quotation_browser_quote_id
        return {
            "session": {
                "status": "failed",
                "task_id": quote_id,
                "rfc": "",
                "current_url": None,
                "steps": [],
                "error_message": f"No se pudo abrir el agente cotizador: {exc}",
                "active_quote_id": active_quote_id,
            }
        }


@router.post("/{quote_id}/browser-session/answer")
def answer_browser_session(
    quote_id: str,
    payload: QuoteBrowserAnswerRequest,
    profile: AccessProfile = Depends(current_access_profile),
):
    try:
        return {"session": answer_quote_browser_session(quote_id, payload)}
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        with _quotation_browser_lock:
            active_quote_id = _active_quotation_browser_quote_id
        return {
            "session": {
                "status": "failed",
                "task_id": quote_id,
                "rfc": "",
                "current_url": None,
                "steps": [],
                "error_message": f"No se pudo continuar el agente cotizador: {exc}",
                "active_quote_id": active_quote_id,
            }
        }


@router.post("/{quote_id}/browser-session/release")
def release_browser_session(
    quote_id: str,
    profile: AccessProfile = Depends(current_access_profile),
):
    return release_quote_browser_session(quote_id)


@router.post("/{quote_id}/documents/{document_kind}")
async def upload_quote_document_route(
    quote_id: str,
    document_kind: str,
    document: UploadFile = File(...),
    profile: AccessProfile = Depends(current_access_profile),
):
    try:
        return await upload_quote_document(quote_id, document_kind, document)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"No fue posible cargar el documento: {exc}") from exc
    finally:
        await document.close()


@router.get("/{quote_id}/quote-email-draft")
def get_quote_email_draft_route(
    quote_id: str,
    profile: AccessProfile = Depends(current_access_profile),
):
    try:
        return quote_email_draft(quote_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"No fue posible leer las cotizaciones: {exc}") from exc


@router.post("/{quote_id}/send-quote-email")
def send_quote_email_route(
    quote_id: str,
    payload: QuoteEmailRequest,
    profile: AccessProfile = Depends(current_access_profile),
):
    try:
        return send_quote_email(quote_id, payload, profile)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"No fue posible enviar la cotización: {exc}") from exc
