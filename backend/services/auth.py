from __future__ import annotations

import hmac
import io
import os
import posixpath
import re
import threading
import time
import zipfile
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree
from xml.sax.saxutils import escape

import pandas as pd
from dotenv import load_dotenv
from pydantic import BaseModel, Field

from drive.client import download_drive_file_bytes
from services.performance import mark_cache


load_dotenv(Path(__file__).resolve().parents[1] / ".env")


USERS_FILE_ID_ENV = "GOOGLE_DRIVE_USERS_FILE_ID"
USERS_CACHE_SECONDS_ENV = "AUTH_USERS_CACHE_SECONDS"
USERS_SNAPSHOT_PATH_ENV = "AUTH_USERS_SNAPSHOT_PATH"
DEFAULT_CACHE_SECONDS = 300
STALE_CACHE_RETRY_SECONDS = 30
REQUIRED_COLUMNS = {"Usuario", "Password"}
DEFAULT_USERS_SNAPSHOT_PATH = (
    Path(__file__).resolve().parents[2]
    / ".runtime"
    / "auth"
    / "users-directory.xlsx"
)

_cache_lock = threading.RLock()
_cached_credentials: dict[str, str] | None = None
_cached_profiles: dict[str, "AccessProfile"] | None = None
_cache_expires_at = 0.0

PROMOTORIAS = (
    "ABBONDANZA",
    "CELAVI",
    "EKILIBRA",
    "FENIX PRE-VISION",
    "TAIICO",
    "URQUIZA GARCIA",
)
MODULES = (
    "inicio",
    "cobranza",
    "renovaciones",
    "cumpleanos",
    "cumpleanos_agentes",
    "pendientes",
    "cartera",
    "clientes",
    "recluta",
    "dashboards",
    "configuracion_mail",
    "carga_bases",
    "accesos",
    "cotizaciones",
    "logs",
    "rrhh",
    "campanas",
    "finanzas",
)
MODULE_COLUMNS = {
    module: f"Permiso_{module.title()}"
    for module in MODULES
}
MODULE_COLUMNS["configuracion_mail"] = "Permiso_Configuracion_Mail"
MODULE_COLUMNS["cumpleanos_agentes"] = "Permiso_Cumpleanos_Agentes"
MODULE_COLUMNS["carga_bases"] = "Permiso_Carga_Bases"
MODULE_COLUMNS["accesos"] = "Permiso_Accesos"
MODULE_COLUMNS["cotizaciones"] = "Permiso_Cotizaciones"
MODULE_COLUMNS["logs"] = "Permiso_Logs"
MODULE_COLUMNS["rrhh"] = "Permiso_RRHH"
MODULE_COLUMNS["campanas"] = "Permiso_Campanas"
MODULE_COLUMNS["finanzas"] = "Permiso_Finanzas"

MODULE_LABELS = {
    "inicio": "Inicio",
    "cobranza": "Cobranza",
    "renovaciones": "Renovaciones",
    "cumpleanos": "Cumpleaños",
    "cumpleanos_agentes": "Cumpleaños de agentes",
    "pendientes": "Pendientes",
    "cartera": "Cartera de Prospectadores",
    "clientes": "Clientes",
    "recluta": "Recluta",
    "dashboards": "Dashboards",
    "configuracion_mail": "Configuración de Mail",
    "carga_bases": "Carga de bases",
    "accesos": "Accesos",
    "cotizaciones": "Cotizaciones",
    "logs": "Logs",
    "rrhh": "RRHH",
    "campanas": "Campañas",
    "finanzas": "Finanzas",
}

PERMISSION_LABELS = {
    "ninguno": "Ninguno",
    "lectura": "Lectura",
    "operacion": "Operación",
}


class AccessUserPayload(BaseModel):
    username: str = Field(min_length=3, max_length=320)
    password: str | None = Field(default=None, max_length=256)
    role: str = Field(default="agente")
    promotorias: list[str] = Field(default_factory=list)
    rfc: str = ""
    aseguradoras: list[str] = Field(default_factory=list)
    module_permissions: dict[str, str] = Field(default_factory=dict)


@dataclass(frozen=True)
class AccessProfile:
    username: str
    role: str
    promotorias: tuple[str, ...]
    rfc: str
    aseguradoras: tuple[str, ...]
    module_permissions: dict[str, str]

    @property
    def is_admin(self) -> bool:
        return self.role == "admin"

    @property
    def is_agent(self) -> bool:
        return self.role == "agente"

    @property
    def is_central_admin(self) -> bool:
        return self.is_admin and set(self.promotorias) == set(PROMOTORIAS)

    def permission_for(self, module: str) -> str:
        return self.module_permissions.get(module, "ninguno")

    def can_read(self, module: str) -> bool:
        return self.permission_for(module) in {"lectura", "operacion"}

    def can_operate(self, module: str) -> bool:
        return self.is_admin and self.permission_for(module) == "operacion"


def _download_users_workbook(file_id: str) -> bytes:
    # Authorization is evaluated for every protected request. Reuse the
    # requests-based Drive downloader so concurrent FastAPI requests do not
    # share googleapiclient/httplib2 sockets on macOS (which intermittently
    # fails with Errno 49 and turns an otherwise valid upload into HTTP 500).
    return download_drive_file_bytes(file_id)


def _build_writable_drive_service():
    try:
        from google.auth import default
        from googleapiclient.discovery import build
    except ImportError as exc:
        raise RuntimeError(
            "Google Drive dependencies are not installed. "
            "Run `pip install -r backend/requirements.txt`."
        ) from exc

    credentials, _ = default(scopes=["https://www.googleapis.com/auth/drive"])
    return build("drive", "v3", credentials=credentials, cache_discovery=False)


def _upload_users_workbook(file_id: str, workbook: bytes) -> None:
    try:
        from googleapiclient.http import MediaIoBaseUpload
    except ImportError as exc:
        raise RuntimeError(
            "Google Drive dependencies are not installed. "
            "Run `pip install -r backend/requirements.txt`."
        ) from exc

    media = MediaIoBaseUpload(
        io.BytesIO(workbook),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False,
    )
    _build_writable_drive_service().files().update(
        fileId=file_id,
        media_body=media,
        supportsAllDrives=True,
    ).execute()


def _split_access_values(value: object) -> tuple[str, ...]:
    text = str(value or "").strip()
    if not text:
        return ()
    return tuple(
        item.strip().upper()
        for item in re.split(r"[,;\n]+", text)
        if item.strip()
    )


def _normalize_permission(value: object) -> str:
    normalized = (
        str(value or "").strip().casefold()
        .replace("ó", "o")
        .replace("á", "a")
    )
    if normalized in {"operacion", "operativo", "escritura", "admin"}:
        return "operacion"
    if normalized in {"lectura", "consulta", "read"}:
        return "lectura"
    return "ninguno"


def _default_module_permissions(role: str, promotorias: tuple[str, ...]) -> dict[str, str]:
    if not promotorias:
        permissions = {module: "ninguno" for module in MODULES}
    elif role == "agente":
        permissions = {
            module: ("lectura" if module == "pendientes" else "ninguno")
            for module in MODULES
        }
    elif role == "admin" and set(promotorias) == set(PROMOTORIAS):
        permissions = {module: "operacion" for module in MODULES}
    elif role == "admin":
        permissions = {
            module: ("operacion" if module == "pendientes" else "ninguno")
            for module in MODULES
        }
    else:
        permissions = {module: "ninguno" for module in MODULES}

    # New modules must be explicitly enabled in the access workbook.
    permissions["cumpleanos"] = "ninguno"
    permissions["cumpleanos_agentes"] = "ninguno"
    permissions["carga_bases"] = "ninguno"
    permissions["accesos"] = "ninguno"
    permissions["cotizaciones"] = "ninguno"
    permissions["logs"] = "ninguno"
    permissions["rrhh"] = "ninguno"
    permissions["campanas"] = "ninguno"
    permissions["finanzas"] = "ninguno"
    return permissions


def _read_user_directory(
    workbook: bytes,
) -> tuple[dict[str, str], dict[str, AccessProfile]]:
    table = pd.read_excel(io.BytesIO(workbook), dtype=str, keep_default_na=False)
    missing_columns = REQUIRED_COLUMNS.difference(table.columns)
    if missing_columns:
        raise ValueError(
            "Users workbook is missing required columns: "
            + ", ".join(sorted(missing_columns))
        )

    credentials: dict[str, str] = {}
    profiles: dict[str, AccessProfile] = {}
    for _, row in table.iterrows():
        username = str(row["Usuario"]).strip().casefold()
        password = str(row["Password"])
        if username and password:
            credentials[username] = password
        if not username:
            continue
        role = str(row.get("Rol", "")).strip().casefold()
        promotorias = _split_access_values(row.get("Promotoria", ""))
        if "*" in promotorias:
            promotorias = PROMOTORIAS
        defaults = _default_module_permissions(role, promotorias)
        permissions = {
            module: (
                _normalize_permission(row.get(column, ""))
                if column in table.columns
                else defaults[module]
            )
            for module, column in MODULE_COLUMNS.items()
        }
        profiles[username] = AccessProfile(
            username=username,
            role=role,
            promotorias=tuple(
                promotoria for promotoria in PROMOTORIAS if promotoria in promotorias
            ),
            rfc=str(row.get("RFC", "")).strip().upper(),
            aseguradoras=_split_access_values(row.get("Aseguradoras", "")),
            module_permissions=permissions,
        )
    return credentials, profiles


def _read_credentials(workbook: bytes) -> dict[str, str]:
    return _read_user_directory(workbook)[0]


def _permission_label(value: object) -> str:
    return PERMISSION_LABELS[_normalize_permission(value)]


def _normalize_role(value: object) -> str:
    role = str(value or "").strip().casefold()
    if role == "admin":
        return "admin"
    if role == "agente":
        return "agente"
    raise ValueError("El rol debe ser Admin o Agente")


def _normalize_promotorias(values: list[str]) -> list[str]:
    normalized = {str(value).strip().upper() for value in values if str(value).strip()}
    if "*" in normalized:
        return list(PROMOTORIAS)
    invalid = sorted(normalized.difference(PROMOTORIAS))
    if invalid:
        raise ValueError("Promotoría no válida: " + ", ".join(invalid))
    return [promotoria for promotoria in PROMOTORIAS if promotoria in normalized]


def _normalize_module_permissions(
    values: dict[str, str],
    *,
    role: str,
    promotorias: tuple[str, ...],
) -> dict[str, str]:
    invalid = sorted(set(values).difference(MODULES))
    if invalid:
        raise ValueError("Módulo no válido: " + ", ".join(invalid))
    defaults = _default_module_permissions(role, promotorias)
    return {
        module: _normalize_permission(values.get(module, defaults[module]))
        for module in MODULES
    }


def _validate_username(username: str) -> str:
    normalized = str(username or "").strip().casefold()
    if "@" not in normalized or "." not in normalized.rsplit("@", 1)[-1]:
        raise ValueError("El usuario debe ser un correo válido")
    return normalized


def _workbook_users_sheet(workbook):
    if "Usuarios" in workbook.sheetnames:
        return workbook["Usuarios"]
    return workbook.active


def _headers_for_sheet(sheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for index, cell in enumerate(sheet[1], start=1):
        value = str(cell.value or "").strip()
        if value:
            headers[value] = index
    return headers


def _ensure_user_headers(sheet) -> dict[str, int]:
    headers = _headers_for_sheet(sheet)
    required = [
        "Usuario",
        "Password",
        "Rol",
        "Promotoria",
        "RFC",
        "Aseguradoras",
        *MODULE_COLUMNS.values(),
    ]
    for header in required:
        if header not in headers:
            column = sheet.max_column + 1
            sheet.cell(row=1, column=column, value=header)
            headers[header] = column
    return headers


def _find_user_row(sheet, headers: dict[str, int], username: str) -> int | None:
    username_column = headers["Usuario"]
    for row_number in range(2, sheet.max_row + 1):
        value = sheet.cell(row=row_number, column=username_column).value
        if str(value or "").strip().casefold() == username:
            return row_number
    return None


def _serialize_user_payload(payload: AccessUserPayload, *, existing_password: str = "") -> dict[str, str]:
    username = _validate_username(payload.username)
    role = _normalize_role(payload.role)
    promotorias = _normalize_promotorias(payload.promotorias)
    permissions = _normalize_module_permissions(
        payload.module_permissions,
        role=role,
        promotorias=tuple(promotorias),
    )
    password = payload.password if payload.password is not None else existing_password
    return {
        "Usuario": username,
        "Password": str(password or ""),
        "Rol": "Admin" if role == "admin" else "Agente",
        "Promotoria": ", ".join(promotorias),
        "RFC": str(payload.rfc or "").strip().upper(),
        "Aseguradoras": ", ".join(
            str(value).strip().upper()
            for value in payload.aseguradoras
            if str(value).strip()
        ),
        **{
            column: _permission_label(permissions[module])
            for module, column in MODULE_COLUMNS.items()
        },
    }


def access_modules_configuration() -> dict[str, object]:
    return {
        "modules": [
            {
                "key": module,
                "label": MODULE_LABELS[module],
                "column": MODULE_COLUMNS[module],
            }
            for module in MODULES
        ],
        "promotorias": list(PROMOTORIAS),
        "roles": ["admin", "agente"],
        "permissions": [
            {"key": key, "label": label}
            for key, label in PERMISSION_LABELS.items()
        ],
    }


def list_access_users() -> list[dict[str, object]]:
    file_id = os.getenv(USERS_FILE_ID_ENV, "").strip()
    if not file_id:
        raise RuntimeError(f"{USERS_FILE_ID_ENV} is not configured")
    workbook = _download_users_workbook(file_id)
    _, profiles = _read_user_directory(workbook)
    table = pd.read_excel(io.BytesIO(workbook), dtype=str, keep_default_na=False)
    password_by_user = {
        str(row.get("Usuario", "")).strip().casefold(): bool(str(row.get("Password", "")))
        for _, row in table.iterrows()
    }
    return [
        {
            "username": profile.username,
            "role": profile.role,
            "promotorias": list(profile.promotorias),
            "rfc": profile.rfc,
            "aseguradoras": list(profile.aseguradoras),
            "module_permissions": profile.module_permissions,
            "has_password": password_by_user.get(profile.username, False),
        }
        for profile in profiles.values()
    ]


def save_access_user(payload: AccessUserPayload, *, create: bool) -> dict[str, object]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to update the users workbook. "
            "Run `pip install -r backend/requirements.txt`."
        ) from exc

    file_id = os.getenv(USERS_FILE_ID_ENV, "").strip()
    if not file_id:
        raise RuntimeError(f"{USERS_FILE_ID_ENV} is not configured")

    username = _validate_username(payload.username)
    with _cache_lock:
        workbook_bytes = _download_users_workbook(file_id)
        workbook = load_workbook(io.BytesIO(workbook_bytes))
        sheet = _workbook_users_sheet(workbook)
        headers = _ensure_user_headers(sheet)
        row_number = _find_user_row(sheet, headers, username)
        if create and row_number is not None:
            raise ValueError("El usuario ya existe")
        if not create and row_number is None:
            raise KeyError("Usuario no encontrado")
        if create and not payload.password:
            raise ValueError("La contraseña es obligatoria para usuarios nuevos")
        if row_number is None:
            row_number = sheet.max_row + 1
            existing_password = ""
        else:
            existing_password = str(
                sheet.cell(row=row_number, column=headers["Password"]).value or ""
            )
        values = _serialize_user_payload(payload, existing_password=existing_password)
        for header, value in values.items():
            sheet.cell(row=row_number, column=headers[header], value=value)
        output = io.BytesIO()
        workbook.save(output)
        updated_workbook = output.getvalue()
        _upload_users_workbook(file_id, updated_workbook)
        _write_users_snapshot(updated_workbook)
        clear_credentials_cache()
    _, profiles = _read_user_directory(updated_workbook)
    profile = profiles[username]
    return {
        "username": profile.username,
        "role": profile.role,
        "promotorias": list(profile.promotorias),
        "rfc": profile.rfc,
        "aseguradoras": list(profile.aseguradoras),
        "module_permissions": profile.module_permissions,
        "has_password": bool(values["Password"]),
    }


def delete_access_user(username: str) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to update the users workbook. "
            "Run `pip install -r backend/requirements.txt`."
        ) from exc

    normalized = _validate_username(username)
    file_id = os.getenv(USERS_FILE_ID_ENV, "").strip()
    if not file_id:
        raise RuntimeError(f"{USERS_FILE_ID_ENV} is not configured")
    with _cache_lock:
        workbook_bytes = _download_users_workbook(file_id)
        workbook = load_workbook(io.BytesIO(workbook_bytes))
        sheet = _workbook_users_sheet(workbook)
        headers = _ensure_user_headers(sheet)
        row_number = _find_user_row(sheet, headers, normalized)
        if row_number is None:
            raise KeyError("Usuario no encontrado")
        sheet.delete_rows(row_number, 1)
        output = io.BytesIO()
        workbook.save(output)
        updated_workbook = output.getvalue()
        _upload_users_workbook(file_id, updated_workbook)
        _write_users_snapshot(updated_workbook)
        clear_credentials_cache()


def set_user_module_permission(username: str, module: str, permission: str) -> None:
    if module not in MODULE_COLUMNS:
        raise ValueError("Módulo no válido")
    file_id = os.getenv(USERS_FILE_ID_ENV, "").strip()
    if not file_id:
        raise RuntimeError(f"{USERS_FILE_ID_ENV} is not configured")
    normalized = _validate_username(username)
    with _cache_lock:
        workbook_bytes = _download_users_workbook(file_id)
        _, profiles = _read_user_directory(workbook_bytes)
        if normalized not in profiles:
            raise KeyError("Usuario no encontrado")
        values = {
            profile.username: _permission_label(profile.permission_for(module))
            for profile in profiles.values()
        }
        values[normalized] = _permission_label(permission)
        updated_workbook = _set_permission_column_in_xlsx(
            workbook_bytes,
            MODULE_COLUMNS[module],
            values,
        )
        _upload_users_workbook(file_id, updated_workbook)
        _write_users_snapshot(updated_workbook)
        clear_credentials_cache()


def _cache_seconds() -> int:
    value = int(os.getenv(USERS_CACHE_SECONDS_ENV, str(DEFAULT_CACHE_SECONDS)))
    return max(0, value)


def _users_snapshot_path() -> Path:
    configured = os.getenv(USERS_SNAPSHOT_PATH_ENV, "").strip()
    return Path(configured).expanduser() if configured else DEFAULT_USERS_SNAPSHOT_PATH


def _read_users_snapshot() -> bytes:
    return _users_snapshot_path().read_bytes()


def _write_users_snapshot(workbook: bytes) -> None:
    """Atomically persist the last valid directory with owner-only access."""
    destination = _users_snapshot_path()
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_suffix(destination.suffix + ".tmp")
    temporary.write_bytes(workbook)
    os.chmod(temporary, 0o600)
    os.replace(temporary, destination)
    os.chmod(destination, 0o600)


def _load_credentials() -> dict[str, str]:
    global _cached_credentials, _cached_profiles, _cache_expires_at

    file_id = os.getenv(USERS_FILE_ID_ENV, "").strip()
    if not file_id:
        raise RuntimeError(f"{USERS_FILE_ID_ENV} is not configured")

    now = time.monotonic()
    with _cache_lock:
        if _cached_credentials is not None and now < _cache_expires_at:
            mark_cache("auth", "hit")
            return _cached_credentials

        # A valid local snapshot keeps authorization available after a process
        # restart even when Google Drive is temporarily slow or unavailable.
        # It is short-lived in memory so the canonical Drive source is retried.
        if _cached_credentials is None:
            try:
                credentials, profiles = _read_user_directory(
                    _read_users_snapshot()
                )
            except (FileNotFoundError, OSError, ValueError):
                pass
            else:
                mark_cache("auth", "snapshot")
                _cached_credentials = credentials
                _cached_profiles = profiles
                _cache_expires_at = now + STALE_CACHE_RETRY_SECONDS
                return credentials

        try:
            mark_cache("auth", "miss")
            workbook = _download_users_workbook(file_id)
            credentials, profiles = _read_user_directory(workbook)
        except Exception:
            # A transient Drive outage must not invalidate a directory that was
            # already downloaded and parsed successfully. Keep the last known
            # access profile briefly, then retry the canonical Drive source.
            # With no prior valid snapshot we still fail closed.
            if _cached_credentials is None or _cached_profiles is None:
                raise
            mark_cache("auth", "stale")
            _cache_expires_at = now + STALE_CACHE_RETRY_SECONDS
            return _cached_credentials
        try:
            _write_users_snapshot(workbook)
        except OSError as exc:
            # A local persistence failure must not reject a valid directory
            # that was just read from the canonical Drive source.
            print(f"Authentication snapshot unavailable: {type(exc).__name__}: {exc}")
        _cached_credentials = credentials
        _cached_profiles = profiles
        _cache_expires_at = now + _cache_seconds()
        return credentials


def clear_credentials_cache() -> None:
    """Clear the in-memory workbook cache (primarily for tests)."""
    global _cached_credentials, _cached_profiles, _cache_expires_at
    with _cache_lock:
        _cached_credentials = None
        _cached_profiles = None
        _cache_expires_at = 0.0


def get_access_profile(username: str) -> AccessProfile:
    normalized = str(username).strip().casefold()
    _load_credentials()
    with _cache_lock:
        profile = (_cached_profiles or {}).get(normalized)
    if profile is None:
        raise KeyError("El usuario no tiene un perfil de acceso configurado")
    return profile


def list_access_profiles() -> tuple[AccessProfile, ...]:
    """Return the cached canonical access directory without downloading it twice."""
    _load_credentials()
    with _cache_lock:
        return tuple((_cached_profiles or {}).values())


def verify_credentials(username, password) -> bool:
    """Verify credentials against the configured read-only Drive workbook."""
    try:
        stored_password = _load_credentials().get(str(username).strip().casefold())
        if stored_password is None:
            return False
        return hmac.compare_digest(stored_password, str(password))
    except Exception as exc:
        # Fail closed without logging credential values.
        print(f"Authentication unavailable: {type(exc).__name__}: {exc}")
        return False


def registered_user(username: str) -> bool:
    """Return whether an email exists without exposing any stored password."""
    try:
        return str(username).strip().casefold() in _load_credentials()
    except Exception as exc:
        print(f"Authentication directory unavailable: {type(exc).__name__}: {exc}")
        return False


_SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


def _xlsx_cell_text(cell, shared_strings: list[str]) -> str:
    cell_type = cell.get("t")
    if cell_type == "inlineStr":
        return "".join(
            node.text or ""
            for node in cell.findall(f".//{{{_SPREADSHEET_NS}}}t")
        )
    value = cell.find(f"{{{_SPREADSHEET_NS}}}v")
    if value is None or value.text is None:
        return ""
    if cell_type == "s":
        return shared_strings[int(value.text)]
    return value.text


def _replace_password_in_xlsx(
    workbook_bytes: bytes,
    normalized_username: str,
    new_password: str,
) -> bytes:
    source = io.BytesIO(workbook_bytes)
    with zipfile.ZipFile(source, "r") as archive:
        names = archive.namelist()
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in names:
            shared_root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
            shared_strings = [
                "".join(
                    node.text or ""
                    for node in item.findall(f".//{{{_SPREADSHEET_NS}}}t")
                )
                for item in shared_root.findall(f"{{{_SPREADSHEET_NS}}}si")
            ]

        workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        relationship_root = ElementTree.fromstring(
            archive.read("xl/_rels/workbook.xml.rels")
        )
        targets = {
            relationship.get("Id"): relationship.get("Target")
            for relationship in relationship_root.findall(
                f"{{{_PACKAGE_REL_NS}}}Relationship"
            )
        }

        updated_path = None
        updated_xml = None
        for sheet in workbook_root.findall(
            f".//{{{_SPREADSHEET_NS}}}sheet"
        ):
            relationship_id = sheet.get(f"{{{_OFFICE_REL_NS}}}id")
            target = targets.get(relationship_id)
            if not target:
                continue
            cleaned_target = target.lstrip("/")
            sheet_path = (
                posixpath.normpath(cleaned_target)
                if cleaned_target.startswith("xl/")
                else posixpath.normpath(posixpath.join("xl", cleaned_target))
            )
            sheet_root = ElementTree.fromstring(archive.read(sheet_path))
            rows = sheet_root.findall(
                f".//{{{_SPREADSHEET_NS}}}sheetData/{{{_SPREADSHEET_NS}}}row"
            )
            if not rows:
                continue

            headers = {}
            for cell in rows[0].findall(f"{{{_SPREADSHEET_NS}}}c"):
                reference = cell.get("r", "")
                column = "".join(character for character in reference if character.isalpha())
                headers[_xlsx_cell_text(cell, shared_strings).strip()] = column
            if not REQUIRED_COLUMNS.issubset(headers):
                continue

            username_column = headers["Usuario"]
            password_column = headers["Password"]
            for row in rows[1:]:
                cells = {
                    "".join(
                        character
                        for character in cell.get("r", "")
                        if character.isalpha()
                    ): cell
                    for cell in row.findall(f"{{{_SPREADSHEET_NS}}}c")
                }
                username_cell = cells.get(username_column)
                if username_cell is None:
                    continue
                if (
                    _xlsx_cell_text(username_cell, shared_strings).strip().casefold()
                    != normalized_username
                ):
                    continue
                password_cell = cells.get(password_column)
                if password_cell is None:
                    raise ValueError(
                        "Registered user has no password cell in the workbook"
                    )

                cell_reference = password_cell.get("r", "")
                original_xml = archive.read(sheet_path)
                encoded_reference = re.escape(cell_reference.encode("utf-8"))
                cell_pattern = re.compile(
                    rb"<c\b(?=[^>]*\br=[\"']"
                    + encoded_reference
                    + rb"[\"'])[^>]*>.*?</c>",
                    re.DOTALL,
                )
                match = cell_pattern.search(original_xml)
                prefix = b""
                if match is None:
                    prefixed_cell_pattern = re.compile(
                        rb"<(?P<prefix>[A-Za-z_][\w.-]*:)c\b"
                        rb"(?=[^>]*\br=[\"']"
                        + encoded_reference
                        + rb"[\"'])[^>]*>.*?</(?P=prefix)c>",
                        re.DOTALL,
                    )
                    match = prefixed_cell_pattern.search(original_xml)
                    if match is not None:
                        prefix = match.group("prefix")
                if match is None:
                    raise ValueError(
                        f"Password cell {cell_reference} was not found in worksheet XML"
                    )

                original_cell = match.group(0)
                start_tag_end = original_cell.find(b">")
                start_tag = original_cell[: start_tag_end + 1]
                start_tag = re.sub(
                    rb"\s+t=[\"'][^\"']*[\"']",
                    b"",
                    start_tag,
                    count=1,
                )
                start_tag = start_tag[:-1] + b' t="inlineStr">'

                password_text = str(new_password)
                preserve_space = (
                    ' xml:space="preserve"'
                    if password_text != password_text.strip()
                    else ""
                )
                escaped_password = escape(password_text).encode("utf-8")
                replacement = (
                    start_tag
                    + b"<"
                    + prefix
                    + b"is><"
                    + prefix
                    + b"t"
                    + preserve_space.encode("utf-8")
                    + b">"
                    + escaped_password
                    + b"</"
                    + prefix
                    + b"t></"
                    + prefix
                    + b"is></"
                    + prefix
                    + b"c>"
                )
                updated_path = sheet_path
                updated_xml = (
                    original_xml[: match.start()]
                    + replacement
                    + original_xml[match.end() :]
                )
                break
            if updated_path:
                break

        if not updated_path or updated_xml is None:
            raise KeyError("Registered user not found")

        output = io.BytesIO()
        with zipfile.ZipFile(output, "w") as updated_archive:
            for item in archive.infolist():
                content = updated_xml if item.filename == updated_path else archive.read(item.filename)
                updated_archive.writestr(item, content)
        return output.getvalue()


def _column_number(letters: str) -> int:
    value = 0
    for letter in letters.upper():
        value = value * 26 + ord(letter) - 64
    return value


def _column_letter(number: int) -> str:
    value = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        value = chr(65 + remainder) + value
    return value


def _replace_or_append_inline_cell(
    worksheet_xml: bytes,
    *,
    row_number: str,
    cell_reference: str,
    value: str,
    style_id: str = "",
) -> bytes:
    encoded_reference = re.escape(cell_reference.encode("utf-8"))
    cell_pattern = re.compile(
        rb"<c\b(?=[^>]*\br=[\"']"
        + encoded_reference
        + rb"[\"'])[^>]*>.*?</c>",
        re.DOTALL,
    )
    escaped_value = escape(str(value)).encode("utf-8")
    style_attribute = (
        b' s="' + escape(style_id).encode("utf-8") + b'"'
        if style_id
        else b""
    )
    replacement = (
        b'<c r="'
        + cell_reference.encode("utf-8")
        + b'"'
        + style_attribute
        + b' t="inlineStr"><is><t>'
        + escaped_value
        + b"</t></is></c>"
    )
    match = cell_pattern.search(worksheet_xml)
    if match is not None:
        return (
            worksheet_xml[: match.start()]
            + replacement
            + worksheet_xml[match.end() :]
        )

    row_pattern = re.compile(
        rb"<row\b(?=[^>]*\br=[\"']"
        + re.escape(row_number.encode("utf-8"))
        + rb"[\"'])[^>]*>.*?</row>",
        re.DOTALL,
    )
    row_match = row_pattern.search(worksheet_xml)
    if row_match is None:
        raise ValueError(f"Worksheet row {row_number} was not found")
    row_xml = row_match.group(0)
    closing_position = row_xml.rfind(b"</row>")
    updated_row = (
        row_xml[:closing_position] + replacement + row_xml[closing_position:]
    )
    return (
        worksheet_xml[: row_match.start()]
        + updated_row
        + worksheet_xml[row_match.end() :]
    )


def _set_permission_column_in_xlsx(
    workbook_bytes: bytes,
    column_name: str,
    values_by_username: dict[str, str],
    *,
    default_value: str = "Ninguno",
) -> bytes:
    """Set one access column without reserializing the rest of the workbook."""
    normalized_values = {
        str(username).strip().casefold(): str(value)
        for username, value in values_by_username.items()
    }
    with zipfile.ZipFile(io.BytesIO(workbook_bytes), "r") as archive:
        names = archive.namelist()
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in names:
            shared_root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
            shared_strings = [
                "".join(
                    node.text or ""
                    for node in item.findall(f".//{{{_SPREADSHEET_NS}}}t")
                )
                for item in shared_root.findall(f"{{{_SPREADSHEET_NS}}}si")
            ]

        workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        relationship_root = ElementTree.fromstring(
            archive.read("xl/_rels/workbook.xml.rels")
        )
        targets = {
            relationship.get("Id"): relationship.get("Target")
            for relationship in relationship_root.findall(
                f"{{{_PACKAGE_REL_NS}}}Relationship"
            )
        }

        updated_path = ""
        updated_xml = b""
        for sheet in workbook_root.findall(f".//{{{_SPREADSHEET_NS}}}sheet"):
            relationship_id = sheet.get(f"{{{_OFFICE_REL_NS}}}id")
            target = targets.get(relationship_id)
            if not target:
                continue
            cleaned_target = target.lstrip("/")
            sheet_path = (
                posixpath.normpath(cleaned_target)
                if cleaned_target.startswith("xl/")
                else posixpath.normpath(posixpath.join("xl", cleaned_target))
            )
            worksheet_bytes = archive.read(sheet_path)
            sheet_root = ElementTree.fromstring(worksheet_bytes)
            rows = sheet_root.findall(
                f".//{{{_SPREADSHEET_NS}}}sheetData/{{{_SPREADSHEET_NS}}}row"
            )
            if not rows:
                continue

            headers: dict[str, str] = {}
            for cell in rows[0].findall(f"{{{_SPREADSHEET_NS}}}c"):
                reference = cell.get("r", "")
                column = "".join(character for character in reference if character.isalpha())
                headers[_xlsx_cell_text(cell, shared_strings).strip()] = column
            if not REQUIRED_COLUMNS.issubset(headers):
                continue

            username_column = headers["Usuario"]
            target_column = headers.get(column_name)
            is_new_column = not target_column
            if not target_column:
                target_column = _column_letter(
                    max(_column_number(column) for column in headers.values()) + 1
                )

            header_style = ""
            header_cells = rows[0].findall(f"{{{_SPREADSHEET_NS}}}c")
            if header_cells:
                header_style = header_cells[-1].get("s", "")
            worksheet_bytes = _replace_or_append_inline_cell(
                worksheet_bytes,
                row_number=rows[0].get("r", "1"),
                cell_reference=f"{target_column}{rows[0].get('r', '1')}",
                value=column_name,
                style_id=header_style,
            )

            for row in rows[1:]:
                row_number = row.get("r", "")
                if not row_number:
                    continue
                cells = {
                    "".join(
                        character
                        for character in cell.get("r", "")
                        if character.isalpha()
                    ): cell
                    for cell in row.findall(f"{{{_SPREADSHEET_NS}}}c")
                }
                username_cell = cells.get(username_column)
                if username_cell is None:
                    continue
                username = (
                    _xlsx_cell_text(username_cell, shared_strings).strip().casefold()
                )
                if not username:
                    continue
                row_cells = row.findall(f"{{{_SPREADSHEET_NS}}}c")
                style_id = row_cells[-1].get("s", "") if row_cells else ""
                worksheet_bytes = _replace_or_append_inline_cell(
                    worksheet_bytes,
                    row_number=row_number,
                    cell_reference=f"{target_column}{row_number}",
                    value=normalized_values.get(username, default_value),
                    style_id=style_id,
                )

            if is_new_column:
                target_number = _column_number(target_column)

                def extend_dimension(match: re.Match[bytes]) -> bytes:
                    start, end_letters, end_row = match.groups()
                    end_number = max(_column_number(end_letters.decode()), target_number)
                    return (
                        b'<dimension ref="'
                        + start
                        + b":"
                        + _column_letter(end_number).encode()
                        + end_row
                        + b'"'
                    )

                worksheet_bytes = re.sub(
                    rb'<dimension ref="([A-Z]+\d+):([A-Z]+)(\d+)"',
                    extend_dimension,
                    worksheet_bytes,
                    count=1,
                )
            updated_path = sheet_path
            updated_xml = worksheet_bytes
            break

        if not updated_path:
            raise ValueError("Users worksheet was not found")

        output = io.BytesIO()
        with zipfile.ZipFile(output, "w") as updated_archive:
            for item in archive.infolist():
                content = (
                    updated_xml
                    if item.filename == updated_path
                    else archive.read(item.filename)
                )
                updated_archive.writestr(item, content)
        return output.getvalue()


def update_password(username: str, new_password: str) -> None:
    """Update one password cell while preserving the existing workbook."""
    global _cached_credentials, _cached_profiles, _cache_expires_at

    normalized_username = str(username).strip().casefold()
    file_id = os.getenv(USERS_FILE_ID_ENV, "").strip()
    if not file_id:
        raise RuntimeError(f"{USERS_FILE_ID_ENV} is not configured")

    with _cache_lock:
        workbook_bytes = _download_users_workbook(file_id)
        updated_workbook = _replace_password_in_xlsx(
            workbook_bytes,
            normalized_username,
            new_password,
        )
        _upload_users_workbook(file_id, updated_workbook)
        _write_users_snapshot(updated_workbook)
        _cached_credentials = None
        _cached_profiles = None
        _cache_expires_at = 0.0
