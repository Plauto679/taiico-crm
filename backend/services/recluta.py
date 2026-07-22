from __future__ import annotations

import hashlib
import io
import os
import re
import unicodedata
from pathlib import Path

import pandas as pd
from dotenv import load_dotenv
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from google.auth import default
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
from openpyxl import load_workbook
from pydantic import BaseModel

from services.session_auth import current_username


load_dotenv(Path(__file__).resolve().parents[1] / ".env")

router = APIRouter(prefix="/recluta", tags=["recluta"])

RECLUTA_SOURCE_FILE_ID = os.getenv(
    "GOOGLE_DRIVE_RECLUTA_SOURCE_FILE_ID",
    "1a4YYy-vF4pESre60BJWXwdp_1DdSmObT",
)
RECLUTA_DOCUMENTS_FOLDER_ID = os.getenv(
    "GOOGLE_DRIVE_RECLUTA_DOCUMENTS_FOLDER_ID",
    "1you3U-LTCi8iVbT_GX8zC7cpT44atKpr",
)
RECLUTA_SHEET_NAME = "Reclutamiento"
DRIVE_SCOPE = "https://www.googleapis.com/auth/drive"
FOLDER_MIME_TYPE = "application/vnd.google-apps.folder"
MAX_DOCUMENT_BYTES = 25 * 1024 * 1024


class ReclutaCreateRequest(BaseModel):
    nombre: str
    telefono: str = ""
    correo: str = ""
    rfc: str = ""
    fase: str = ""
    estatus: str = ""


def normalize_header(value: str) -> str:
    text = " ".join(str(value or "").strip().split()).casefold()
    return "".join(
        character
        for character in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(character)
    )


def normalize_name(value: str) -> str:
    return " ".join(str(value or "").strip().split())


def prospect_id(row: dict[str, str], row_number: int) -> str:
    identity = "|".join([
        row.get("rfc", "").strip().upper(),
        normalize_name(row.get("nombre", "")).casefold(),
        row.get("correo", "").strip().casefold(),
        re.sub(r"\D+", "", row.get("telefono", "")),
    ])
    stable_value = identity if identity.strip("|") else f"row:{row_number}"
    return hashlib.sha256(stable_value.encode()).hexdigest()[:20]


def parse_recluta_workbook(workbook: bytes) -> tuple[list[str], list[dict]]:
    excel = pd.ExcelFile(io.BytesIO(workbook))
    sheet_name = RECLUTA_SHEET_NAME if RECLUTA_SHEET_NAME in excel.sheet_names else excel.sheet_names[0]
    table = pd.read_excel(
        io.BytesIO(workbook),
        sheet_name=sheet_name,
        dtype=str,
        keep_default_na=False,
    )
    columns = [str(column).strip() for column in table.columns]
    normalized_columns = {normalize_header(column): column for column in columns}
    required = {"nombre", "telefono", "correo", "rfc", "fase", "estatus"}
    missing = sorted(required.difference(normalized_columns))
    if missing:
        raise ValueError("La base de Recluta no contiene: " + ", ".join(missing))

    prospects = []
    for row_index, source_row in table.iterrows():
        raw = {
            column: str(source_row.get(column, "")).strip()
            for column in columns
        }
        if not any(raw.values()):
            continue
        canonical = {
            key: raw[normalized_columns[key]]
            for key in required
        }
        prospects.append({
            "id": prospect_id(canonical, row_index + 2),
            "source_row": row_index + 2,
            "nombre": normalize_name(canonical["nombre"]),
            "telefono": canonical["telefono"],
            "correo": canonical["correo"],
            "rfc": canonical["rfc"].upper(),
            "fase": canonical["fase"] or "Sin fase",
            "estatus": canonical["estatus"] or "Sin estatus",
            "raw": raw,
        })
    return columns, prospects


def build_drive_service():
    credentials, _ = default(scopes=[DRIVE_SCOPE])
    return build("drive", "v3", credentials=credentials, cache_discovery=False)


def download_source_workbook(service) -> bytes:
    output = io.BytesIO()
    request = service.files().get_media(
        fileId=RECLUTA_SOURCE_FILE_ID,
        supportsAllDrives=True,
    )
    downloader = MediaIoBaseDownload(output, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return output.getvalue()


def append_prospect_to_workbook(workbook: bytes, prospect: ReclutaCreateRequest) -> bytes:
    document = load_workbook(io.BytesIO(workbook))
    sheet = document[RECLUTA_SHEET_NAME] if RECLUTA_SHEET_NAME in document.sheetnames else document.active
    columns = [str(cell.value or "").strip() for cell in sheet[1]]
    normalized_columns = {normalize_header(column): index + 1 for index, column in enumerate(columns)}
    required = {"nombre", "telefono", "correo", "rfc", "fase", "estatus"}
    missing = sorted(required.difference(normalized_columns))
    if missing:
        raise ValueError("La base de Recluta no contiene: " + ", ".join(missing))

    values = {
        "nombre": normalize_name(prospect.nombre),
        "telefono": prospect.telefono.strip(),
        "correo": prospect.correo.strip().casefold(),
        "rfc": prospect.rfc.strip().upper(),
        "fase": prospect.fase.strip(),
        "estatus": prospect.estatus.strip(),
    }
    if not values["nombre"]:
        raise ValueError("El nombre del recluta es obligatorio")

    if values["rfc"]:
        rfc_column = normalized_columns["rfc"]
        for row_number in range(2, sheet.max_row + 1):
            existing_rfc = str(sheet.cell(row=row_number, column=rfc_column).value or "").strip().upper()
            if existing_rfc == values["rfc"]:
                raise ValueError(f"Ya existe un recluta con el RFC {values['rfc']}")

    next_row = 2
    while any(sheet.cell(row=next_row, column=index).value not in (None, "") for index in range(1, len(columns) + 1)):
        next_row += 1
    for key, value in values.items():
        sheet.cell(row=next_row, column=normalized_columns[key], value=value)

    output = io.BytesIO()
    document.save(output)
    return output.getvalue()


def upload_source_workbook(service, workbook: bytes) -> None:
    media = MediaIoBaseUpload(
        io.BytesIO(workbook),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False,
    )
    service.files().update(
        fileId=RECLUTA_SOURCE_FILE_ID,
        media_body=media,
        supportsAllDrives=True,
    ).execute()


def sanitize_folder_name(value: str) -> str:
    cleaned = re.sub(r"[\\/:*?\"<>|]+", "-", normalize_name(value))
    return cleaned[:180] or "Prospecto sin nombre"


def document_name_for(value: str, original_filename: str) -> str:
    if not normalize_name(value):
        raise ValueError("El nombre del documento es obligatorio")
    cleaned = sanitize_folder_name(value)
    original_suffix = Path(original_filename or "").suffix
    if original_suffix and not Path(cleaned).suffix:
        cleaned = f"{cleaned}{original_suffix.lower()}"
    return cleaned[:220]


def folder_name_for(prospect: dict, all_prospects: list[dict]) -> str:
    name = sanitize_folder_name(prospect["nombre"])
    same_name = [
        item for item in all_prospects
        if normalize_name(item["nombre"]).casefold() == normalize_name(prospect["nombre"]).casefold()
    ]
    if len(same_name) > 1:
        discriminator = prospect["rfc"] or prospect["id"][-6:]
        return sanitize_folder_name(f"{name} - {discriminator}")
    return name


def list_child_folders(service) -> list[dict]:
    response = service.files().list(
        q=(
            f"'{RECLUTA_DOCUMENTS_FOLDER_ID}' in parents and "
            f"mimeType = '{FOLDER_MIME_TYPE}' and trashed = false"
        ),
        fields="files(id,name,webViewLink,modifiedTime)",
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
        pageSize=1000,
    ).execute()
    return response.get("files", [])


def load_prospects(service=None) -> tuple[list[str], list[dict]]:
    service = service or build_drive_service()
    columns, prospects = parse_recluta_workbook(download_source_workbook(service))
    folders = {
        str(folder.get("name", "")).casefold(): folder
        for folder in list_child_folders(service)
    }
    for prospect in prospects:
        expected_name = folder_name_for(prospect, prospects)
        folder = folders.get(expected_name.casefold())
        prospect["folder_name"] = expected_name
        prospect["folder_id"] = folder.get("id") if folder else None
        prospect["folder_url"] = folder.get("webViewLink") if folder else None
    return columns, prospects


def get_prospect(prospect_id_value: str, service=None) -> tuple[dict, list[dict]]:
    _, prospects = load_prospects(service)
    prospect = next((item for item in prospects if item["id"] == prospect_id_value), None)
    if not prospect:
        raise HTTPException(status_code=404, detail="Prospecto no encontrado en la base de Recluta")
    return prospect, prospects


def create_folder_for_prospect(service, prospect: dict) -> dict:
    created = service.files().create(
        body={
            "name": prospect["folder_name"],
            "parents": [RECLUTA_DOCUMENTS_FOLDER_ID],
            "mimeType": FOLDER_MIME_TYPE,
        },
        fields="id,name,webViewLink",
        supportsAllDrives=True,
    ).execute()
    prospect["folder_id"] = created.get("id")
    prospect["folder_url"] = created.get("webViewLink")
    return prospect


@router.get("/prospects")
def get_recluta_prospects():
    try:
        columns, prospects = load_prospects()
        phases = list(dict.fromkeys(item["fase"] for item in prospects))
        return {
            "source_file_id": RECLUTA_SOURCE_FILE_ID,
            "source_url": f"https://docs.google.com/spreadsheets/d/{RECLUTA_SOURCE_FILE_ID}",
            "documents_folder_id": RECLUTA_DOCUMENTS_FOLDER_ID,
            "documents_folder_url": f"https://drive.google.com/drive/folders/{RECLUTA_DOCUMENTS_FOLDER_ID}",
            "columns": columns,
            "phases": phases,
            "prospects": prospects,
        }
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"No fue posible leer la base de Recluta: {exc}") from exc


@router.post("/prospects")
def create_recluta_prospect(
    request: ReclutaCreateRequest,
    _username: str = Depends(current_username),
):
    try:
        service = build_drive_service()
        updated_workbook = append_prospect_to_workbook(download_source_workbook(service), request)
        upload_source_workbook(service, updated_workbook)
        _, prospects = load_prospects(service)
        expected_id = prospect_id({
            "nombre": request.nombre,
            "telefono": request.telefono,
            "correo": request.correo,
            "rfc": request.rfc,
        }, 0)
        prospect = next((item for item in prospects if item["id"] == expected_id), None)
        if not prospect:
            raise RuntimeError("El registro se guardó, pero no pudo releerse desde Drive")

        folder_warning = None
        if not prospect["folder_id"]:
            try:
                create_folder_for_prospect(service, prospect)
            except Exception as exc:
                folder_warning = f"El registro se guardó, pero no fue posible crear su carpeta: {exc}"
        return {
            "created": True,
            "prospect": prospect,
            "folder_warning": folder_warning,
        }
    except ValueError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"No fue posible agregar el recluta: {exc}") from exc


@router.get("/prospects/{prospect_id_value}/documents")
def get_recluta_documents(prospect_id_value: str):
    try:
        service = build_drive_service()
        prospect, _ = get_prospect(prospect_id_value, service)
        if not prospect["folder_id"]:
            return {"prospect": prospect, "folder_missing": True, "documents": []}
        response = service.files().list(
            q=f"'{prospect['folder_id']}' in parents and trashed = false",
            fields="files(id,name,mimeType,webViewLink,modifiedTime,size)",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
            orderBy="folder,name",
            pageSize=1000,
        ).execute()
        return {
            "prospect": prospect,
            "folder_missing": False,
            "documents": response.get("files", []),
        }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"No fue posible consultar los documentos: {exc}") from exc


@router.post("/prospects/{prospect_id_value}/documents")
async def upload_recluta_document(
    prospect_id_value: str,
    document_name: str = Form(...),
    document: UploadFile = File(...),
    _username: str = Depends(current_username),
):
    try:
        service = build_drive_service()
        prospect, _ = get_prospect(prospect_id_value, service)
        if not prospect["folder_id"]:
            raise HTTPException(status_code=409, detail="Primero debe crearse la carpeta documental del recluta")

        final_name = document_name_for(document_name, document.filename or "")
        existing = service.files().list(
            q=f"'{prospect['folder_id']}' in parents and trashed = false",
            fields="files(id,name)",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
            pageSize=1000,
        ).execute().get("files", [])
        if any(str(item.get("name", "")).casefold() == final_name.casefold() for item in existing):
            raise HTTPException(status_code=409, detail=f"Ya existe un documento llamado {final_name}")

        content = await document.read(MAX_DOCUMENT_BYTES + 1)
        if not content:
            raise HTTPException(status_code=400, detail="El archivo está vacío")
        if len(content) > MAX_DOCUMENT_BYTES:
            raise HTTPException(status_code=413, detail="El archivo supera el límite de 25 MB")

        media = MediaIoBaseUpload(
            io.BytesIO(content),
            mimetype=document.content_type or "application/octet-stream",
            resumable=False,
        )
        created = service.files().create(
            body={"name": final_name, "parents": [prospect["folder_id"]]},
            media_body=media,
            fields="id,name,mimeType,webViewLink,modifiedTime,size",
            supportsAllDrives=True,
        ).execute()
        return {"uploaded": True, "document": created}
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"No fue posible cargar el documento: {exc}") from exc
    finally:
        await document.close()


@router.post("/prospects/{prospect_id_value}/folder")
def create_recluta_folder(
    prospect_id_value: str,
    _username: str = Depends(current_username),
):
    try:
        service = build_drive_service()
        prospect, _ = get_prospect(prospect_id_value, service)
        if prospect["folder_id"]:
            return {"created": False, "prospect": prospect}
        create_folder_for_prospect(service, prospect)
        return {"created": True, "prospect": prospect}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"No fue posible crear la carpeta: {exc}") from exc
