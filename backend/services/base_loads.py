from __future__ import annotations

import hashlib
import json
import logging
import os
import re
import shutil
import tempfile
import uuid
from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Iterable

from fastapi import APIRouter, File, HTTPException, UploadFile
from openpyxl import load_workbook
from starlette.concurrency import run_in_threadpool

from config import GOOGLE_DRIVE_SOURCE_FOLDERS, METLIFE_PATHS
from drive.client import download_drive_file_bytes


router = APIRouter(prefix="/base-loads", tags=["base-loads"])
logger = logging.getLogger(__name__)

BUSINESS_COLUMN_COUNT = 24  # A:X
POLICY_COLUMN_INDEX = 4  # E / NPOLIZA
END_DATE_COLUMN_INDEX = 7  # H / FFINVIG
START_DATE_COLUMN_INDEX = 6  # G / FINIVIG
AGENT_COLUMN_INDEX = 13  # N / AGENTE
EXPECTED_HEADERS = (
    "CONTRATANTE",
    "RFC",
    "RAMSUBRAMO",
    "PRODUCTO",
    "NPOLIZA",
    "POLORIG",
    "FINIVIG",
    "FFINVIG",
    "NESQFPAGO",
    "NOMBREL",
    "ESTATUS",
    "CONDCOB",
    "PROMOTORIA",
    "AGENTE",
    "NOMBRE",
    "PRIMA",
    "PRIMA",
    "RECARGO",
    "GTOSEXP",
    "IVA",
    "MONEDA",
    "PAGADOHASTA",
    "DEDUCIBLE",
    "COASEGURO",
)
MAX_UPLOAD_BYTES = 100 * 1024 * 1024
PREVIEW_TTL_HOURS = 24
DEFAULT_HISTORY_FOLDER_ID = "1VRzdS1Oqpf1XXnT1hHrzz024ocjhLQzM"
HISTORY_FOLDER_ID_ENV = "BASE_LOAD_HISTORY_FOLDER_ID"
DEFAULT_AGENTS_METLIFE_FILE_ID = "1IoeLDCQe4T3DofStiBSaI09xjX2-RSby"
AGENTS_METLIFE_FILE_ID_ENV = "GOOGLE_DRIVE_AGENTS_METLIFE_FILE_ID"
DEFAULT_METLIFE_GMM_FILE_ID = "1e1fL1qH4jBJLSNdSO-izTg2eDjYV6VNn"


def staging_root() -> Path:
    configured = os.getenv("BASE_LOAD_STAGING_DIR", "").strip()
    root = (
        Path(configured).expanduser()
        if configured
        else Path(__file__).resolve().parents[2] / ".runtime" / "base-loads"
    )
    root.mkdir(parents=True, exist_ok=True)
    return root


def stage_agents_workbook(destination: Path) -> Path:
    """Download the canonical MetLife agent catalog from Google Drive."""
    file_id = os.getenv(
        AGENTS_METLIFE_FILE_ID_ENV,
        DEFAULT_AGENTS_METLIFE_FILE_ID,
    ).strip()
    if not file_id:
        raise ValueError("No está configurado el archivo de Agentes MetLife en Drive")

    destination.write_bytes(download_drive_file_bytes(file_id))
    return destination


def history_folder_id() -> str:
    return os.getenv(HISTORY_FOLDER_ID_ENV, DEFAULT_HISTORY_FOLDER_ID).strip()


def canonical_drive_file_id() -> str:
    configured = (
        GOOGLE_DRIVE_SOURCE_FOLDERS.get("renovaciones.metlife_gmm", {}).get("file_id")
        or DEFAULT_METLIFE_GMM_FILE_ID
    )
    file_id = str(configured).strip()
    if not file_id:
        raise ValueError("No está configurado el archivo canónico MetLife GMM en Drive")
    return file_id


def upload_history_backup(source_path: Path, backup_name: str) -> dict[str, str]:
    """Upload the previous canonical workbook before replacing it locally."""
    folder_id = history_folder_id()
    if not folder_id:
        raise ValueError("No está configurada la carpeta histórica de Carga de bases")
    try:
        from googleapiclient.http import MediaFileUpload
        from services.auth import _build_writable_drive_service
    except ImportError as exc:
        raise RuntimeError(
            "No están instaladas las dependencias de escritura de Google Drive"
        ) from exc

    media = MediaFileUpload(
        str(source_path),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False,
    )
    created = (
        _build_writable_drive_service()
        .files()
        .create(
            body={"name": backup_name, "parents": [folder_id]},
            media_body=media,
            fields="id,name,webViewLink",
            supportsAllDrives=True,
        )
        .execute()
    )
    file_id = str(created["id"])
    return {
        "backup_file_id": file_id,
        "backup_name": str(created.get("name") or backup_name),
        "backup_url": str(
            created.get("webViewLink")
            or f"https://drive.google.com/file/d/{file_id}/view"
        ),
        "backup_folder_id": folder_id,
    }


def drive_file_metadata(file_id: str) -> dict:
    from services.auth import _build_writable_drive_service

    return (
        _build_writable_drive_service()
        .files()
        .get(
            fileId=file_id,
            fields="id,name,webViewLink,md5Checksum,size,modifiedTime,version",
            supportsAllDrives=True,
        )
        .execute()
    )


def backup_drive_workbook(file_id: str, backup_name: str) -> dict[str, str]:
    """Copy the exact Drive version that will be replaced into history."""
    folder_id = history_folder_id()
    if not folder_id:
        raise ValueError("No está configurada la carpeta histórica de Carga de bases")
    from services.auth import _build_writable_drive_service

    created = (
        _build_writable_drive_service()
        .files()
        .copy(
            fileId=file_id,
            body={"name": backup_name, "parents": [folder_id]},
            fields="id,name,webViewLink",
            supportsAllDrives=True,
        )
        .execute()
    )
    backup_file_id = str(created["id"])
    return {
        "backup_file_id": backup_file_id,
        "backup_name": str(created.get("name") or backup_name),
        "backup_url": str(
            created.get("webViewLink")
            or f"https://drive.google.com/file/d/{backup_file_id}/view"
        ),
        "backup_folder_id": folder_id,
    }


def update_drive_workbook(file_id: str, source_path: Path) -> dict[str, str]:
    """Replace Drive bytes in place, preserving the canonical ID and sharing URL."""
    try:
        from googleapiclient.http import MediaFileUpload
        from services.auth import _build_writable_drive_service
    except ImportError as exc:
        raise RuntimeError(
            "No están instaladas las dependencias de escritura de Google Drive"
        ) from exc

    media = MediaFileUpload(
        str(source_path),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False,
    )
    updated = (
        _build_writable_drive_service()
        .files()
        .update(
            fileId=file_id,
            media_body=media,
            fields="id,name,webViewLink,md5Checksum,size,modifiedTime,version",
            supportsAllDrives=True,
        )
        .execute()
    )
    return {
        "drive_file_id": str(updated["id"]),
        "drive_name": str(updated.get("name") or "Metlife GMM.xlsx"),
        "drive_url": str(
            updated.get("webViewLink")
            or f"https://drive.google.com/file/d/{file_id}/view"
        ),
        "drive_md5": str(updated.get("md5Checksum") or ""),
        "drive_size": str(updated.get("size") or ""),
        "drive_modified_time": str(updated.get("modifiedTime") or ""),
        "drive_version": str(updated.get("version") or ""),
    }


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def row_values(row: Iterable[object], width: int) -> tuple[object, ...]:
    values = tuple(row)
    return values[:width] + (None,) * max(0, width - len(values))


def normalized_row(row: Iterable[object], width: int) -> tuple[str, ...]:
    return tuple(normalize_cell(value) for value in row_values(row, width))


def load_allowed_agent_keys(path: Path) -> set[str]:
    if not path.exists():
        raise ValueError(f"No se encontró el catálogo de agentes: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["Datos"] if "Datos" in workbook.sheetnames else workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [normalize_cell(value) for value in next(rows)]
        try:
            key_index = headers.index("CLAVE_DEFINITIVA")
        except ValueError as exc:
            raise ValueError(
                "El catálogo de agentes no contiene CLAVE_DEFINITIVA"
            ) from exc
        return {
            normalize_cell(row[key_index])
            for row in rows
            if key_index < len(row) and normalize_cell(row[key_index])
        }
    finally:
        workbook.close()


def incoming_rows(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    # Some MetLife exports incorrectly declare their dimension as A1:A1.
    sheet.reset_dimensions()
    rows = sheet.iter_rows(values_only=True)
    try:
        header = row_values(next(rows), BUSINESS_COLUMN_COUNT)
    except StopIteration as exc:
        workbook.close()
        raise ValueError("El archivo cargado está vacío") from exc
    normalized_headers = tuple(normalize_cell(value) for value in header)
    if normalized_headers != EXPECTED_HEADERS:
        workbook.close()
        raise ValueError(
            "El archivo no tiene las columnas esperadas de MetLife GMM en A:X"
        )
    try:
        yield workbook, header, rows
    finally:
        workbook.close()


def business_key(row: Iterable[object]) -> tuple[str, ...]:
    return normalized_row(row, BUSINESS_COLUMN_COUNT)


def period_key(row: Iterable[object]) -> tuple[str, str, str]:
    values = row_values(row, BUSINESS_COLUMN_COUNT)
    return (
        normalize_cell(values[POLICY_COLUMN_INDEX]),
        normalize_cell(values[START_DATE_COLUMN_INDEX]),
        normalize_cell(values[END_DATE_COLUMN_INDEX]),
    )


def select_incoming_rows(path: Path, allowed_keys: set[str]) -> dict:
    total_rows = 0
    filtered_rows = 0
    exact_business_rows: dict[tuple[str, ...], tuple[object, ...]] = {}
    policy_occurrences: Counter[str] = Counter()
    period_occurrences: Counter[tuple[str, str, str]] = Counter()

    for _workbook, header, rows in incoming_rows(path):
        for raw_row in rows:
            total_rows += 1
            business = row_values(raw_row, BUSINESS_COLUMN_COUNT)
            agent_key = normalize_cell(business[AGENT_COLUMN_INDEX])
            if agent_key not in allowed_keys:
                continue
            filtered_rows += 1
            normalized_business = tuple(normalize_cell(value) for value in business)
            exact_business_rows.setdefault(normalized_business, business)
            policy_number = normalized_business[POLICY_COLUMN_INDEX]
            if not policy_number:
                continue
            policy_occurrences[policy_number] += 1
            period_occurrences[period_key(business)] += 1

    return {
        "header": header,
        "rows": list(exact_business_rows.values()),
        "total_rows": total_rows,
        "filtered_rows": filtered_rows,
        "unique_business_rows": len(exact_business_rows),
        "duplicate_business_rows": filtered_rows - len(exact_business_rows),
        "policies_with_multiple_rows": sum(
            1 for count in policy_occurrences.values() if count > 1
        ),
        "unique_policies": len(policy_occurrences),
        "unique_periods": len(period_occurrences),
    }


def load_current_policies(path: Path) -> dict:
    if not path.exists():
        raise ValueError(f"No se encontró la base canónica: {path}")
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if "GMM" not in workbook.sheetnames:
            raise ValueError("La base canónica no contiene la hoja GMM")
        sheet = workbook["GMM"]
        rows = sheet.iter_rows(values_only=True)
        header = tuple(next(rows))
        if len(header) < BUSINESS_COLUMN_COUNT:
            raise ValueError("La hoja GMM no contiene las columnas A:X")
        width = len(header)
        rows_by_business: dict[tuple[str, ...], tuple[object, ...]] = {}
        policies: set[str] = set()
        for raw_row in rows:
            values = row_values(raw_row, width)
            policy_number = normalize_cell(values[POLICY_COLUMN_INDEX])
            if policy_number:
                policies.add(policy_number)
                rows_by_business[business_key(values)] = values
        return {
            "header": header,
            "width": width,
            "rows": list(rows_by_business.values()),
            "policies": policies,
        }
    finally:
        workbook.close()


def preview_from_loaded_data(allowed_keys: set[str], incoming: dict, current: dict) -> dict:
    incoming_policies = {
        normalize_cell(row[POLICY_COLUMN_INDEX]) for row in incoming["rows"]
    }
    current_policies = current["policies"]
    incoming_periods = {period_key(row) for row in incoming["rows"]}
    preserved_comment_rows = sum(
        period_key(row) in incoming_periods
        and any(normalize_cell(value) for value in row[BUSINESS_COLUMN_COUNT:])
        for row in current["rows"]
    )
    exception_rows = [
        row for row in current["rows"] if period_key(row) not in incoming_periods
    ]
    return {
        "allowed_agent_keys": len(allowed_keys),
        "source_rows": incoming["total_rows"],
        "rows_after_agent_filter": incoming["filtered_rows"],
        "unique_a_x_rows": incoming["unique_business_rows"],
        "duplicate_a_x_rows": incoming["duplicate_business_rows"],
        "unique_incoming_policies": incoming["unique_policies"],
        "unique_policy_periods": incoming["unique_periods"],
        "policies_with_multiple_rows": incoming["policies_with_multiple_rows"],
        "existing_policies_updated": len(incoming_policies & current_policies),
        "new_policies_added": len(incoming_policies - current_policies),
        "current_policies_preserved_as_exceptions": len(current_policies - incoming_policies),
        "current_rows_preserved_as_exceptions": len(exception_rows),
        "rows_with_preserved_y_plus_data": preserved_comment_rows,
        "final_policy_count": len(incoming_policies | current_policies),
        "final_row_count": len(incoming["rows"]) + len(exception_rows),
    }


def build_preview(upload_path: Path, canonical_path: Path, agent_path: Path) -> dict:
    allowed_keys = load_allowed_agent_keys(agent_path)
    incoming = select_incoming_rows(upload_path, allowed_keys)
    current = load_current_policies(canonical_path)
    return preview_from_loaded_data(allowed_keys, incoming, current)


def current_policies_from_sheet(sheet) -> dict:
    rows = sheet.iter_rows(values_only=True)
    try:
        header = tuple(next(rows))
    except StopIteration as exc:
        raise ValueError("La hoja GMM está vacía") from exc
    if len(header) < BUSINESS_COLUMN_COUNT:
        raise ValueError("La hoja GMM no contiene las columnas A:X")
    width = len(header)
    rows_by_business: dict[tuple[str, ...], tuple[object, ...]] = {}
    policies: set[str] = set()
    for raw_row in rows:
        values = row_values(raw_row, width)
        policy_number = normalize_cell(values[POLICY_COLUMN_INDEX])
        if policy_number:
            policies.add(policy_number)
            rows_by_business[business_key(values)] = values
    return {
        "header": header,
        "width": width,
        "rows": list(rows_by_business.values()),
        "policies": policies,
    }


def prepare_candidate_workbook(
    upload_path: Path,
    canonical_path: Path,
    agent_path: Path,
    destination: Path,
) -> dict:
    """Build the exact workbook represented by the preview."""
    allowed_keys = load_allowed_agent_keys(agent_path)
    incoming = select_incoming_rows(upload_path, allowed_keys)
    workbook = load_workbook(canonical_path)
    try:
        if "GMM" not in workbook.sheetnames:
            raise ValueError("La base canónica no contiene la hoja GMM")
        sheet = workbook["GMM"]
        current = current_policies_from_sheet(sheet)
        preview = preview_from_loaded_data(allowed_keys, incoming, current)
        width = current["width"]
        incoming_rows = incoming["rows"]
        current_rows = current["rows"]

        exact_indexes = {
            business_key(row): index for index, row in enumerate(incoming_rows)
        }
        period_indexes: dict[tuple[str, str, str], list[int]] = defaultdict(list)
        for index, row in enumerate(incoming_rows):
            period_indexes[period_key(row)].append(index)

        preserved_by_index: dict[int, tuple[object, ...]] = {}
        exceptions: list[tuple[object, ...]] = []
        for old_row in current_rows:
            matched_index = exact_indexes.get(business_key(old_row))
            if matched_index is None:
                candidates = period_indexes.get(period_key(old_row), [])
                matched_index = candidates[-1] if candidates else None
            if matched_index is None:
                exceptions.append(old_row)
                continue
            if matched_index not in preserved_by_index:
                preserved_by_index[matched_index] = row_values(
                    old_row[BUSINESS_COLUMN_COUNT:],
                    width - BUSINESS_COLUMN_COUNT,
                )

        merged_rows = [
            row_values(business, BUSINESS_COLUMN_COUNT)
            + preserved_by_index.get(
                index,
                (None,) * (width - BUSINESS_COLUMN_COUNT),
            )
            for index, business in enumerate(incoming_rows)
        ] + exceptions

        if sheet.max_row > 1:
            sheet.delete_rows(2, sheet.max_row - 1)
        for merged_row in merged_rows:
            sheet.append(list(row_values(merged_row, width)))
        final_row = max(1, sheet.max_row)
        for table in sheet.tables.values():
            table.ref = f"A1:{sheet.cell(1, width).column_letter}{final_row}"

        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(destination)
        return preview
    finally:
        workbook.close()


def file_sha256(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        while chunk := source.read(chunk_size):
            digest.update(chunk)
    return digest.hexdigest()


def file_md5(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.md5(usedforsecurity=False)
    with path.open("rb") as source:
        while chunk := source.read(chunk_size):
            digest.update(chunk)
    return digest.hexdigest()


def apply_prepared_workbook(
    candidate_path: Path,
    canonical_path: Path,
    expected_candidate_sha256: str,
    expected_canonical_sha256: str,
) -> dict:
    """Back up and atomically install a previously prepared workbook."""
    if file_sha256(candidate_path) != expected_candidate_sha256:
        raise ValueError("El libro preparado cambió después de la vista previa")
    if file_sha256(canonical_path) != expected_canonical_sha256:
        raise RuntimeError(
            "La base canónica cambió después de la vista previa. Genera una nueva vista previa."
        )

    drive_file_id = canonical_drive_file_id()
    local_md5 = file_md5(canonical_path)
    drive_before = drive_file_metadata(drive_file_id)
    drive_before_md5 = str(drive_before.get("md5Checksum") or "")
    if drive_before_md5 and drive_before_md5 != local_md5:
        raise RuntimeError(
            "El archivo de Drive cambió o no coincide con la copia local. "
            "Sincronízalos antes de aplicar una nueva carga."
        )

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_name = f"{canonical_path.stem} antes de carga {timestamp}.xlsx"
    backup = backup_drive_workbook(drive_file_id, backup_name)
    drive = update_drive_workbook(drive_file_id, candidate_path)
    candidate_md5 = file_md5(candidate_path)
    if drive.get("drive_md5") and drive["drive_md5"] != candidate_md5:
        raise RuntimeError(
            "Drive recibió el archivo pero su checksum no coincide con el libro preparado"
        )
    canonical_path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        prefix="taiico-base-load-",
        suffix=".xlsx",
        dir=canonical_path.parent,
        delete=False,
    ) as temporary:
        temporary_path = Path(temporary.name)
    try:
        shutil.copyfile(candidate_path, temporary_path)
        temporary_path.replace(canonical_path)
    finally:
        temporary_path.unlink(missing_ok=True)
    return {**backup, **drive, "canonical_path": str(canonical_path)}


def replace_canonical_workbook(
    upload_path: Path,
    canonical_path: Path,
    agent_path: Path,
) -> dict:
    """Compatibility helper for prepare-and-apply callers."""
    canonical_sha256 = file_sha256(canonical_path)
    with tempfile.NamedTemporaryFile(
        prefix="taiico-base-load-prepared-",
        suffix=".xlsx",
        delete=False,
    ) as temporary:
        candidate_path = Path(temporary.name)
    try:
        preview = prepare_candidate_workbook(
            upload_path, canonical_path, agent_path, candidate_path
        )
        applied = apply_prepared_workbook(
            candidate_path,
            canonical_path,
            file_sha256(candidate_path),
            canonical_sha256,
        )
        return {**preview, **applied}
    finally:
        candidate_path.unlink(missing_ok=True)


def cleanup_expired_previews() -> None:
    cutoff = datetime.now(timezone.utc) - timedelta(hours=PREVIEW_TTL_HOURS)
    for candidate in staging_root().iterdir():
        try:
            modified = datetime.fromtimestamp(candidate.stat().st_mtime, timezone.utc)
        except FileNotFoundError:
            continue
        if candidate.is_dir() and modified < cutoff:
            shutil.rmtree(candidate, ignore_errors=True)


async def save_upload(upload: UploadFile, destination: Path) -> tuple[int, str]:
    size = 0
    digest = hashlib.sha256()
    with destination.open("wb") as output:
        while chunk := await upload.read(1024 * 1024):
            size += len(chunk)
            if size > MAX_UPLOAD_BYTES:
                raise HTTPException(
                    status_code=413,
                    detail="El archivo supera el límite de 100 MB",
                )
            digest.update(chunk)
            output.write(chunk)
    return size, digest.hexdigest()


def safe_filename(value: str | None) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._ -]+", "_", value or "carga.xlsx").strip()
    return cleaned or "carga.xlsx"


@router.post("/metlife-gmm/preview")
async def preview_metlife_gmm_base(file: UploadFile = File(...)):
    cleanup_expired_previews()
    filename = safe_filename(file.filename)
    if not filename.casefold().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Sólo se aceptan archivos .xlsx")
    token = uuid.uuid4().hex
    token_dir = staging_root() / token
    token_dir.mkdir(parents=True)
    upload_path = token_dir / "source.xlsx"
    agents_path = token_dir / "agents.xlsx"
    candidate_path = token_dir / "prepared.xlsx"
    try:
        size, sha256 = await save_upload(file, upload_path)
        canonical_path = Path(METLIFE_PATHS["RENOVACIONES_GMM"])
        canonical_sha256 = await run_in_threadpool(file_sha256, canonical_path)
        await run_in_threadpool(stage_agents_workbook, agents_path)
        preview = await run_in_threadpool(
            prepare_candidate_workbook,
            upload_path,
            canonical_path,
            agents_path,
            candidate_path,
        )
        manifest = {
            "token": token,
            "filename": filename,
            "size": size,
            "sha256": sha256,
            "canonical_sha256": canonical_sha256,
            "candidate_sha256": await run_in_threadpool(file_sha256, candidate_path),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "preview": preview,
        }
        (token_dir / "manifest.json").write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2) + "\n"
        )
        return manifest
    except HTTPException:
        shutil.rmtree(token_dir, ignore_errors=True)
        raise
    except Exception as exc:
        shutil.rmtree(token_dir, ignore_errors=True)
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    finally:
        await file.close()


@router.post("/metlife-gmm/apply/{token}")
async def apply_metlife_gmm_base(token: str):
    if not re.fullmatch(r"[a-f0-9]{32}", token):
        raise HTTPException(status_code=400, detail="Token de carga inválido")
    token_dir = staging_root() / token
    upload_path = token_dir / "source.xlsx"
    candidate_path = token_dir / "prepared.xlsx"
    manifest_path = token_dir / "manifest.json"
    if not upload_path.exists() or not candidate_path.exists() or not manifest_path.exists():
        raise HTTPException(status_code=404, detail="La vista previa expiró o no existe")
    manifest = json.loads(manifest_path.read_text())
    digest = await run_in_threadpool(file_sha256, upload_path)
    if digest != manifest.get("sha256"):
        raise HTTPException(status_code=409, detail="El archivo cambió después de la vista previa")
    try:
        result = await run_in_threadpool(
            apply_prepared_workbook,
            candidate_path,
            Path(METLIFE_PATHS["RENOVACIONES_GMM"]),
            manifest["candidate_sha256"],
            manifest["canonical_sha256"],
        )
        try:
            from services.renewal_ingestion import sync_local_canonical_renewals

            renewal_sync = await run_in_threadpool(
                sync_local_canonical_renewals,
                "renovaciones.metlife_gmm",
            )
        except Exception as sync_exc:
            logger.exception("Canonical workbook applied but renewal synchronization failed")
            renewal_sync = {"status": "failed", "error": str(sync_exc)}
        shutil.rmtree(token_dir, ignore_errors=True)
        return {
            "applied": True,
            "filename": manifest["filename"],
            **manifest["preview"],
            **result,
            "renewal_sync": renewal_sync,
        }
    except Exception as exc:
        logger.exception(
            "MetLife GMM base apply failed; token=%s staging=%s",
            token,
            token_dir,
        )
        raise HTTPException(status_code=500, detail=str(exc)) from exc
