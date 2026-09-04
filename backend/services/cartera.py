from datetime import date, datetime
from decimal import Decimal
import io
import os
from pathlib import Path
from threading import Lock
import unicodedata
from uuid import uuid4

from fastapi import APIRouter, Depends, HTTPException, Query
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import func

from config import CARTERA_SOURCE_FILE_IDS, METLIFE_PATHS, SURA_PATHS
from database import Client, Insurer, Policy, Product, SessionLocal, User
from drive.client import download_drive_file_bytes
from services.auth import AccessProfile
from services.authorization import require_module_access


router = APIRouter(prefix="/cartera", tags=["cartera"])
_WORKBOOK_LOCK = Lock()


class CarteraRecordPayload(BaseModel):
    policy_number: str = Field(min_length=1, max_length=100)
    current_policy_number: str | None = Field(default=None, max_length=100)
    contractor: str = Field(min_length=1, max_length=255)
    prospector: str = Field(min_length=1, max_length=255)
    percentage: float = Field(ge=0, le=100)
    payment_start_date: date | None = None
    insurer: str = Field(min_length=2, max_length=50)
    policy_type: str = Field(default="VIDA", min_length=2, max_length=20)


def _normalized_insurer(value: str) -> str:
    insurer = value.strip().casefold()
    if insurer not in {"metlife", "sura", "axa", "aarco"}:
        raise HTTPException(status_code=422, detail="Aseguradora no válida")
    return insurer


def _percentage_for_ui(value) -> float:
    number = float(value or 0)
    return number * 100 if abs(number) <= 1 else number


def _commission_anniversary(start_date: date) -> date:
    try:
        return start_date.replace(year=start_date.year + 1)
    except ValueError:
        return start_date.replace(year=start_date.year + 1, day=28)


def prospector_commission_is_expired(policy: Policy, on_date: date | None = None) -> bool:
    metadata = dict(policy.metadata_json or {})
    raw_start = metadata.get("payment_start_date")
    try:
        start_date = date.fromisoformat(str(raw_start)) if raw_start else None
    except ValueError:
        start_date = None
    return bool(start_date and (on_date or date.today()) >= _commission_anniversary(start_date))


def effective_prospector_percentage(policy: Policy, on_date: date | None = None) -> float:
    if prospector_commission_is_expired(policy, on_date):
        return 0.0
    return _percentage_for_ui(policy.commission_percentage)


def _serialize(policy: Policy) -> dict:
    metadata = dict(policy.metadata_json or {})
    client_metadata = dict(policy.client.metadata_json or {}) if policy.client else {}
    return {
        "id": policy.id,
        "policy_number": policy.policy_number,
        "current_policy_number": metadata.get("current_policy_number") or policy.policy_number,
        "contractor": policy.client.full_name if policy.client else "",
        "prospector": metadata.get("prospector") or client_metadata.get("prospectador", ""),
        "percentage": effective_prospector_percentage(policy),
        "payment_start_date": metadata.get("payment_start_date"),
        "insurer": policy.insurer_id,
        "policy_type": policy.product.branch if policy.product else "",
    }


def _product_for(db, insurer_id: str, policy_type: str) -> Product:
    branch = policy_type.strip().upper() or "VIDA"
    product = db.query(Product).filter(Product.insurer_id == insurer_id, Product.branch == branch).first()
    if product:
        return product
    product = Product(
        id=f"prod_{insurer_id}_{branch.casefold()}",
        insurer_id=insurer_id,
        name=f"{insurer_id.upper()} {branch}",
        branch=branch,
    )
    db.add(product)
    db.flush()
    return product


def _owner_for(db, profile: AccessProfile) -> User:
    user = db.query(User).filter(func.lower(User.email) == profile.username.casefold()).first()
    user = user or db.query(User).filter(User.id == "usr_admin").first()
    if not user:
        raise HTTPException(status_code=422, detail="No se encontró un usuario responsable")
    return user


def _canonical_sheet(payload: CarteraRecordPayload):
    insurer = _normalized_insurer(payload.insurer)
    if insurer == "metlife":
        return Path(METLIFE_PATHS["CARTERA"]), "GMM" if payload.policy_type.upper() == "GMM" else "Vida"
    if insurer == "sura":
        return Path(SURA_PATHS["CARTERA"]), "SURA"
    return None


def _build_writable_drive_service():
    try:
        from google.auth import default
        from googleapiclient.discovery import build
    except ImportError as exc:
        raise RuntimeError("Faltan las dependencias de Google Drive") from exc
    credentials, _ = default(scopes=["https://www.googleapis.com/auth/drive"])
    return build("drive", "v3", credentials=credentials, cache_discovery=False)


def _upload_drive_workbook(file_id: str, contents: bytes) -> None:
    try:
        from googleapiclient.http import MediaIoBaseUpload
    except ImportError as exc:
        raise RuntimeError("Faltan las dependencias de Google Drive") from exc
    media = MediaIoBaseUpload(
        io.BytesIO(contents),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False,
    )
    _build_writable_drive_service().files().update(
        fileId=file_id,
        media_body=media,
        supportsAllDrives=True,
    ).execute()


def _canonical_contents(insurer: str, path: Path) -> tuple[bytes, str]:
    file_id = CARTERA_SOURCE_FILE_IDS.get(insurer, "").strip()
    if file_id:
        return download_drive_file_bytes(file_id), file_id
    if not path.exists():
        raise HTTPException(status_code=503, detail=f"No se encontró el archivo canónico {path.name}")
    return path.read_bytes(), ""


def _save_canonical_contents(path: Path, file_id: str, contents: bytes) -> None:
    if file_id:
        _upload_drive_workbook(file_id, contents)
    # Keep the historical desktop copy current when that directory is mounted.
    if path.exists():
        temporary = path.with_name(f".{path.name}.{uuid4().hex}.tmp")
        try:
            temporary.write_bytes(contents)
            os.replace(temporary, path)
        finally:
            temporary.unlink(missing_ok=True)


def _write_canonical(payload: CarteraRecordPayload, original_policy_number: str | None = None):
    target = _canonical_sheet(payload)
    if not target:
        return None
    path, sheet_name = target
    insurer = _normalized_insurer(payload.insurer)
    with _WORKBOOK_LOCK:
        snapshot, file_id = _canonical_contents(insurer, path)
        workbook = load_workbook(io.BytesIO(snapshot))
        sheet = workbook[sheet_name]
        headers = {str(cell.value or "").strip().casefold(): cell.column for cell in sheet[1] if cell.value}
        if not any(key in headers for key in ("inicio de pago", "fecha inicio de pago")):
            column = sheet.max_column + 1
            sheet.cell(1, column).value = "Inicio de pago"
            headers["inicio de pago"] = column
        lookup = str(original_policy_number or "").strip()
        row_number = None
        if lookup:
            for row in range(2, sheet.max_row + 1):
                values = [str(sheet.cell(row, headers[key]).value or "").removesuffix(".0").strip() for key in ("poliza", "póliza", "poliza actual", "póliza actual") if key in headers]
                if lookup in values:
                    row_number = row
                    break
        row_number = row_number or sheet.max_row + 1

        def assign(keys: tuple[str, ...], value) -> None:
            for key in keys:
                if key in headers:
                    sheet.cell(row_number, headers[key]).value = value
                    return

        assign(("poliza", "póliza"), payload.policy_number.strip())
        assign(("poliza actual", "póliza actual"), (payload.current_policy_number or payload.policy_number).strip())
        assign(("contratante",), payload.contractor.strip())
        assign(("prospectador",), payload.prospector.strip())
        # Canonical workbooks store percentages as ratios (0.8 = 80%).
        assign(("porcentaje",), payload.percentage / 100 if payload.percentage > 1 else payload.percentage)
        assign(("inicio de pago", "fecha inicio de pago"), payload.payment_start_date)
        output = io.BytesIO()
        workbook.save(output)
        _save_canonical_contents(path, file_id, output.getvalue())
        return path, file_id, snapshot


def _restore_canonical(snapshot) -> None:
    if snapshot:
        path, file_id, contents = snapshot
        with _WORKBOOK_LOCK:
            _save_canonical_contents(path, file_id, contents)


def _header_key(value) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    return " ".join("".join(char for char in text if not unicodedata.combining(char)).casefold().split())


def _policy_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip().removesuffix(".0")


def parse_cartera_workbook(contents: bytes, insurer: str) -> list[dict]:
    """Parse a canonical workbook and resolve repeated policies deterministically.

    Administrators occasionally append a corrected row instead of replacing the
    prior one. The last occurrence therefore wins while preserving source order.
    """
    insurer_id = _normalized_insurer(insurer)
    workbook = load_workbook(io.BytesIO(contents), data_only=True)
    sheet_names = ("Vida", "GMM") if insurer_id == "metlife" else ("SURA",)
    rows_by_policy: dict[str, dict] = {}
    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        headers = {_header_key(cell.value): cell.column for cell in sheet[1] if cell.value}

        def value(row_number: int, *aliases: str):
            for alias in aliases:
                column = headers.get(_header_key(alias))
                if column:
                    return sheet.cell(row_number, column).value
            return None

        for row_number in range(2, sheet.max_row + 1):
            policy_number = _policy_text(value(row_number, "Póliza", "Poliza"))
            if not policy_number:
                continue
            percentage_value = value(row_number, "Porcentaje")
            try:
                percentage = Decimal(str(percentage_value or 0))
            except Exception:
                percentage = Decimal("0")
            payment_start = value(row_number, "Inicio de pago", "Fecha inicio de pago")
            if isinstance(payment_start, datetime):
                payment_start = payment_start.date()
            rows_by_policy[policy_number] = {
                "policy_number": policy_number,
                "current_policy_number": _policy_text(value(row_number, "Póliza actual", "Poliza actual")) or policy_number,
                "contractor": str(value(row_number, "Contratante") or "").strip(),
                "prospector": str(value(row_number, "Prospectador") or "").strip(),
                "percentage": percentage,
                "payment_start_date": payment_start.isoformat() if isinstance(payment_start, date) else None,
                "policy_type": sheet_name.upper() if sheet_name in {"Vida", "GMM"} else "GMM",
            }
    return list(rows_by_policy.values())


def sync_cartera_source(insurer: str, *, contents: bytes | None = None, db=None) -> dict:
    """Refresh the fast SQL index from its canonical Drive workbook."""
    insurer_id = _normalized_insurer(insurer)
    file_id = CARTERA_SOURCE_FILE_IDS.get(insurer_id, "").strip()
    if contents is None:
        if not file_id:
            raise RuntimeError(f"No hay un archivo de Drive configurado para {insurer_id}")
        contents = download_drive_file_bytes(file_id)
    rows = parse_cartera_workbook(contents, insurer_id)
    session = db or SessionLocal()
    owns_session = db is None
    created = updated = conflicts = 0
    try:
        insurer_record = session.query(Insurer).filter(Insurer.id == insurer_id).first()
        if not insurer_record:
            raise RuntimeError(f"La aseguradora {insurer_id} no está configurada")
        owner = session.query(User).filter(User.id == "usr_pamela").first()
        owner = owner or session.query(User).filter(User.id == "usr_admin").first()
        if not owner:
            raise RuntimeError("No se encontró un usuario responsable")
        existing = {
            policy.policy_number: policy
            for policy in session.query(Policy).filter(Policy.insurer_id == insurer_id).all()
        }
        policies_by_number = {
            policy.policy_number: policy
            for policy in session.query(Policy).filter(
                Policy.policy_number.in_([row["policy_number"] for row in rows])
            ).all()
        }
        for row in rows:
            policy = existing.get(row["policy_number"])
            contractor = row["contractor"]
            if policy:
                metadata = dict(policy.metadata_json or {})
                policy.commission_percentage = row["percentage"]
                policy.metadata_json = {
                    **metadata,
                    "prospector": row["prospector"],
                    "current_policy_number": row["current_policy_number"],
                    "payment_start_date": row["payment_start_date"],
                    "cartera_source": "google_drive",
                }
                if contractor:
                    policy.client.full_name = contractor
                client_metadata = dict(policy.client.metadata_json or {})
                policy.client.metadata_json = {**client_metadata, "prospectador": row["prospector"]}
                updated += 1
                continue
            if row["policy_number"] in policies_by_number:
                conflicts += 1
                continue
            client = Client(
                full_name=contractor or f"{insurer_id.upper()} CLIENT P-{row['policy_number']}",
                responsible_user_id=owner.id,
                status="active",
                metadata_json={"prospectador": row["prospector"]},
            )
            session.add(client)
            session.flush()
            product = _product_for(session, insurer_id, row["policy_type"])
            session.add(Policy(
                policy_number=row["policy_number"],
                client_id=client.id,
                insurer_id=insurer_id,
                product_id=product.id,
                effective_start_date=date.today(),
                effective_end_date=date.today(),
                status="in_force",
                premium_amount=Decimal("0"),
                payment_frequency="annual",
                responsible_user_id=owner.id,
                commission_percentage=row["percentage"],
                metadata_json={
                    "prospector": row["prospector"],
                    "current_policy_number": row["current_policy_number"],
                    "payment_start_date": row["payment_start_date"],
                    "cartera_source": "google_drive",
                },
            ))
            created += 1
        session.commit()
        return {
            "insurer": insurer_id,
            "source_file_id": file_id,
            "source_rows": len(rows),
            "created": created,
            "updated": updated,
            "conflicts": conflicts,
        }
    except Exception:
        session.rollback()
        raise
    finally:
        if owns_session:
            session.close()


@router.get("/data")
def get_cartera_data(
    insurer: str = Query(..., description="Insurer name"),
    type: str = Query("ALL", description="Policy type: ALL, VIDA, GMM"),
):
    db = SessionLocal()
    canonical_snapshot = None
    try:
        insurer_id = _normalized_insurer(insurer)
        query = db.query(Policy).join(Client).filter(Policy.insurer_id == insurer_id)
        if type.upper() != "ALL":
            query = query.join(Product).filter(Product.branch == type.upper())
        return [_serialize(policy) for policy in query.order_by(Policy.policy_number).all()]
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"No fue posible cargar la cartera: {exc}") from exc
    finally:
        db.close()


@router.post("/records", status_code=201)
def create_cartera_record(
    payload: CarteraRecordPayload,
    profile: AccessProfile = Depends(require_module_access("cartera", operation=True)),
):
    db = SessionLocal()
    canonical_snapshot = None
    try:
        policy_number = payload.policy_number.strip()
        insurer_id = _normalized_insurer(payload.insurer)
        existing_policy = db.query(Policy).filter(Policy.policy_number == policy_number).first()
        if existing_policy and existing_policy.insurer_id != insurer_id:
            raise HTTPException(
                status_code=409,
                detail="La póliza ya existe asociada a otra aseguradora",
            )
        owner = _owner_for(db, profile)
        if not db.query(Insurer).filter(Insurer.id == insurer_id).first():
            raise HTTPException(status_code=422, detail="La aseguradora no está configurada")
        contractor = payload.contractor.strip()
        client = db.query(Client).filter(
            func.upper(func.trim(Client.full_name)) == contractor.upper(),
            Client.status != "inactive",
        ).first()
        if not client:
            client = Client(
                full_name=contractor,
                responsible_user_id=owner.id,
                metadata_json={"prospectador": payload.prospector.strip()},
            )
            db.add(client)
            db.flush()
        requested_branch = payload.policy_type.strip().upper()
        product = (
            existing_policy.product
            if existing_policy
            and existing_policy.product
            and existing_policy.product.branch == requested_branch
            else _product_for(db, insurer_id, requested_branch)
        )
        if existing_policy:
            policy = existing_policy
            policy.client_id = client.id
            policy.product_id = product.id
            policy.commission_percentage = Decimal(str(payload.percentage))
            policy.responsible_user_id = owner.id
            policy.metadata_json = {
                **dict(policy.metadata_json or {}),
                "prospector": payload.prospector.strip(),
                "current_policy_number": (payload.current_policy_number or policy_number).strip(),
                "payment_start_date": payload.payment_start_date.isoformat() if payload.payment_start_date else None,
                "cartera_manual": True,
            }
        else:
            policy = Policy(
                policy_number=policy_number,
                client_id=client.id,
                insurer_id=insurer_id,
                product_id=product.id,
                effective_start_date=date.today(),
                effective_end_date=date.today(),
                premium_amount=Decimal("0"),
                payment_frequency="annual",
                responsible_user_id=owner.id,
                commission_percentage=Decimal(str(payload.percentage)),
                metadata_json={
                    "prospector": payload.prospector.strip(),
                    "current_policy_number": (payload.current_policy_number or policy_number).strip(),
                    "payment_start_date": payload.payment_start_date.isoformat() if payload.payment_start_date else None,
                    "cartera_manual": True,
                },
            )
            db.add(policy)
        canonical_snapshot = _write_canonical(
            payload,
            original_policy_number=policy_number if existing_policy else None,
        )
        db.commit()
        db.refresh(policy)
        return _serialize(policy)
    except HTTPException:
        db.rollback()
        _restore_canonical(canonical_snapshot)
        raise
    except Exception as exc:
        db.rollback()
        _restore_canonical(canonical_snapshot)
        raise HTTPException(status_code=500, detail=f"No fue posible registrar la relación: {exc}") from exc
    finally:
        db.close()


@router.put("/records/{record_id}")
def update_cartera_record(
    record_id: str,
    payload: CarteraRecordPayload,
    _profile: AccessProfile = Depends(require_module_access("cartera", operation=True)),
):
    db = SessionLocal()
    canonical_snapshot = None
    try:
        policy = db.query(Policy).filter(Policy.id == record_id).first()
        if not policy:
            raise HTTPException(status_code=404, detail="El registro ya no existe")
        policy_number = payload.policy_number.strip()
        duplicate = db.query(Policy).filter(Policy.policy_number == policy_number, Policy.id != record_id).first()
        if duplicate:
            raise HTTPException(status_code=409, detail="Ya existe una póliza con ese número")
        insurer_id = _normalized_insurer(payload.insurer)
        original_policy_number = policy.policy_number
        policy.policy_number = policy_number
        policy.insurer_id = insurer_id
        policy.product_id = _product_for(db, insurer_id, payload.policy_type).id
        policy.commission_percentage = Decimal(str(payload.percentage))
        policy.metadata_json = {
            **dict(policy.metadata_json or {}),
            "prospector": payload.prospector.strip(),
            "current_policy_number": (payload.current_policy_number or policy_number).strip(),
            "payment_start_date": payload.payment_start_date.isoformat() if payload.payment_start_date else None,
        }
        policy.client.full_name = payload.contractor.strip()
        canonical_snapshot = _write_canonical(payload, original_policy_number)
        db.commit()
        db.refresh(policy)
        return _serialize(policy)
    except HTTPException:
        db.rollback()
        _restore_canonical(canonical_snapshot)
        raise
    except Exception as exc:
        db.rollback()
        _restore_canonical(canonical_snapshot)
        raise HTTPException(status_code=500, detail=f"No fue posible actualizar la relación: {exc}") from exc
    finally:
        db.close()


@router.get("/search")
def search_cartera(query: str = Query(..., min_length=1)):
    db = SessionLocal()
    try:
        term = f"%{query}%"
        policies = db.query(Policy).join(Client).filter(
            (Policy.policy_number.like(term)) | (Client.full_name.like(term))
        ).limit(50).all()
        return [{
            "poliza": policy.policy_number,
            "contratante": policy.client.full_name if policy.client else "",
            "aseguradora": policy.insurer_id.upper(),
            "estatus": policy.status,
            "ramo": policy.product.branch if policy.product else "",
        } for policy in policies]
    finally:
        db.close()
