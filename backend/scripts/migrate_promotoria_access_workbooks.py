from __future__ import annotations

import argparse
import io
import re
import zipfile
from copy import copy
from pathlib import Path
from xml.sax.saxutils import escape

from openpyxl import load_workbook

from services.xlsx_integrity import repair_workbook_integrity


PROMOTORIAS = {
    "ABBONDANZA",
    "CELAVI",
    "EKILIBRA",
    "FENIX PRE-VISION",
    "TAIICO",
    "URQUIZA GARCIA",
}
MODULE_COLUMNS = (
    "Permiso_Inicio",
    "Permiso_Cobranza",
    "Permiso_Renovaciones",
    "Permiso_Cumpleanos",
    "Permiso_Cumpleanos_Agentes",
    "Permiso_Pendientes",
    "Permiso_Cartera",
    "Permiso_Clientes",
    "Permiso_Recluta",
    "Permiso_Dashboards",
    "Permiso_Configuracion_Mail",
)


def split_values(value: object) -> set[str]:
    return {
        item.strip().upper()
        for item in re.split(r"[,;\n]+", str(value or ""))
        if item.strip()
    }


def copy_column_style(sheet, source_column: int, target_column: int) -> None:
    sheet.column_dimensions[
        sheet.cell(1, target_column).column_letter
    ].width = sheet.column_dimensions[
        sheet.cell(1, source_column).column_letter
    ].width
    for row in range(1, sheet.max_row + 1):
        source = sheet.cell(row, source_column)
        target = sheet.cell(row, target_column)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def migrate_users(path: Path) -> None:
    workbook = load_workbook(path)
    sheet = workbook[workbook.sheetnames[0]]
    headers = {
        str(sheet.cell(1, column).value or "").strip(): column
        for column in range(1, sheet.max_column + 1)
    }
    for required in ("Usuario", "Password", "Rol", "Promotoria", "RFC", "Aseguradoras"):
        if required not in headers:
            raise ValueError(f"Users workbook is missing {required}")

    for header in MODULE_COLUMNS:
        if header not in headers:
            column = sheet.max_column + 1
            copy_column_style(sheet, max(1, column - 1), column)
            sheet.cell(1, column).value = header
            headers[header] = column

    for row in range(2, sheet.max_row + 1):
        role = str(sheet.cell(row, headers["Rol"]).value or "").strip().casefold()
        promotorias = split_values(sheet.cell(row, headers["Promotoria"]).value)
        central_admin = role == "admin" and promotorias == PROMOTORIAS
        for header in MODULE_COLUMNS:
            username = str(
                sheet.cell(row, headers["Usuario"]).value or ""
            ).strip().casefold()
            if header in {"Permiso_Cumpleanos", "Permiso_Cumpleanos_Agentes"}:
                permission = (
                    "Lectura"
                    if username == "alberto.alfaro@taiico.com"
                    else "Ninguno"
                )
            elif not promotorias:
                permission = "Ninguno"
            elif role == "agente":
                permission = "Lectura" if header == "Permiso_Pendientes" else "Ninguno"
            elif central_admin:
                permission = "Operación"
            elif role == "admin":
                permission = "Operación" if header == "Permiso_Pendientes" else "Ninguno"
            else:
                permission = "Ninguno"
            sheet.cell(row, headers[header]).value = permission
    workbook.save(path)


def migrate_pending(path: Path, sheet_name: str, minimum_core_columns: int) -> None:
    workbook = load_workbook(path)
    sheet = workbook[sheet_name]
    headers = [
        str(sheet.cell(1, column).value or "").strip()
        for column in range(1, sheet.max_column + 1)
    ]
    if all(header in headers for header in ("Promotoria", "RFC Agente")):
        workbook.save(path)
        return

    history_column = next(
        (
            column
            for column in range(minimum_core_columns + 1, sheet.max_column + 1)
            if re.match(
                r"^(?:fecha hoy|\d{4}-\d{1,2}-\d{1,2}|\d{1,2}[/-])",
                str(sheet.cell(1, column).value or "").strip().casefold(),
            )
        ),
        minimum_core_columns + 1,
    )
    sheet.insert_cols(history_column, amount=2)
    copy_column_style(sheet, max(1, history_column - 1), history_column)
    copy_column_style(sheet, max(1, history_column - 1), history_column + 1)
    sheet.cell(1, history_column).value = "Promotoria"
    sheet.cell(1, history_column + 1).value = "RFC Agente"

    for table in sheet.tables.values():
        start, end = table.ref.split(":")
        end_row = re.sub(r"^[A-Z]+", "", end)
        table.ref = f"{start}:{sheet.cell(int(end_row), sheet.max_column).coordinate}"
        if table.autoFilter:
            table.autoFilter.ref = table.ref
    workbook.save(path)


def column_number(letters: str) -> int:
    value = 0
    for letter in letters:
        value = value * 26 + ord(letter) - 64
    return value


def column_letter(number: int) -> str:
    value = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        value = chr(65 + remainder) + value
    return value


def insert_pending_columns_low_level(
    path: Path,
    insert_column: int,
    headers: tuple[str, ...],
    *,
    sheet_path: str = "xl/worksheets/sheet1.xml",
) -> None:
    """Insert columns without loading worksheets with styled million-row ranges."""
    if not headers:
        return
    amount = len(headers)
    with zipfile.ZipFile(path, "r") as archive:
        names = archive.namelist()
        sheet_xml = archive.read(sheet_path).decode("utf-8")
        if all(f"<t>{escape(header)}</t>" in sheet_xml for header in headers):
            return

        def shift_reference(match: re.Match[str]) -> str:
            letters, row = match.groups()
            number = column_number(letters)
            return f'r="{column_letter(number + amount if number >= insert_column else number)}{row}"'

        sheet_xml = re.sub(r'r="([A-Z]+)(\d+)"', shift_reference, sheet_xml)
        row_one = re.search(r'<row\b[^>]*\br="1"[^>]*>.*?</row>', sheet_xml, re.DOTALL)
        if not row_one:
            raise ValueError("No se encontró la fila de encabezados")
        insertion_ref = f"{column_letter(insert_column + amount)}1"
        position = row_one.group(0).find(f'<c r="{insertion_ref}"')
        if position < 0:
            position = row_one.group(0).rfind("</row>")
        new_cells = "".join(
            f'<c r="{column_letter(insert_column + offset)}1" t="inlineStr">'
            f"<is><t>{escape(header)}</t></is></c>"
            for offset, header in enumerate(headers)
        )
        updated_row = row_one.group(0)[:position] + new_cells + row_one.group(0)[position:]
        sheet_xml = sheet_xml[:row_one.start()] + updated_row + sheet_xml[row_one.end():]

        def extend_dimension(match: re.Match[str]) -> str:
            start, end_letters, end_row = match.groups()
            return (
                f'<dimension ref="{start}:'
                f'{column_letter(column_number(end_letters) + amount)}{end_row}"'
            )

        sheet_xml = re.sub(
            r'<dimension ref="([A-Z]+\d+):([A-Z]+)(\d+)"',
            extend_dimension,
            sheet_xml,
            count=1,
        )
        replacements = {sheet_path: sheet_xml.encode("utf-8")}
        sheet_filename = sheet_path.rsplit("/", 1)[-1]
        relations_path = f"xl/worksheets/_rels/{sheet_filename}.rels"
        relation_xml = (
            archive.read(relations_path).decode("utf-8")
            if relations_path in names
            else ""
        )
        table_paths = {
            "xl/" + target.removeprefix("../")
            for target in re.findall(
                r'<Relationship\b[^>]*\bType="[^"]*/table"[^>]*\bTarget="([^"]+)"',
                relation_xml,
            )
        }
        for table_path in table_paths:
            table_xml = archive.read(table_path).decode("utf-8")
            if all(f'name="{escape(header)}"' in table_xml for header in headers):
                continue
            table_xml = re.sub(
                r'(<table\b[^>]*\bref="[A-Z]+\d+:)([A-Z]+)(\d+")',
                lambda match: (
                    f"{match.group(1)}"
                    f"{column_letter(column_number(match.group(2)) + amount)}"
                    f"{match.group(3)}"
                ),
                table_xml,
                count=1,
            )
            table_xml = re.sub(
                r'(<autoFilter\b[^>]*\bref="[A-Z]+\d+:)([A-Z]+)(\d+")',
                lambda match: (
                    f"{match.group(1)}"
                    f"{column_letter(column_number(match.group(2)) + amount)}"
                    f"{match.group(3)}"
                ),
                table_xml,
                count=1,
            )
            columns = list(re.finditer(r'<tableColumn\b[^>]*/>', table_xml))
            if len(columns) >= insert_column - 1:
                max_id = max(
                    (int(value) for value in re.findall(r'\bid="(\d+)"', table_xml)),
                    default=len(columns),
                )
                insertion = columns[insert_column - 2].end() if insert_column > 1 else columns[0].start()
                new_columns = "".join(
                    f'<tableColumn id="{max_id + offset + 1}" name="{escape(header)}"/>'
                    for offset, header in enumerate(headers)
                )
                table_xml = table_xml[:insertion] + new_columns + table_xml[insertion:]
                table_xml = re.sub(
                    r'(<tableColumns\b[^>]*\bcount=")(\d+)(")',
                    lambda match: f"{match.group(1)}{int(match.group(2)) + amount}{match.group(3)}",
                    table_xml,
                    count=1,
                )
            replacements[table_path] = table_xml.encode("utf-8")

        output = io.BytesIO()
        with zipfile.ZipFile(output, "w") as destination:
            for item in archive.infolist():
                destination.writestr(item, replacements.get(item.filename, archive.read(item.filename)))
    path.write_bytes(repair_workbook_integrity(output.getvalue()))


def migrate_pending_low_level(
    path: Path,
    insert_column: int,
    *,
    sheet_path: str = "xl/worksheets/sheet1.xml",
) -> None:
    insert_pending_columns_low_level(
        path,
        insert_column,
        ("Promotoria", "RFC Agente"),
        sheet_path=sheet_path,
    )


def rename_pending_column_low_level(
    path: Path,
    old_header: str,
    new_header: str,
    *,
    sheet_path: str = "xl/worksheets/sheet1.xml",
) -> None:
    """Rename a header and its table column without expanding styled worksheets."""
    with zipfile.ZipFile(path, "r") as archive:
        names = archive.namelist()
        sheet_xml = archive.read(sheet_path).decode("utf-8")
        old_cell = f"<t>{escape(old_header)}</t>"
        new_cell = f"<t>{escape(new_header)}</t>"
        if new_cell in sheet_xml and old_cell not in sheet_xml:
            return
        replacements: dict[str, bytes] = {}
        if old_cell in sheet_xml:
            sheet_xml = sheet_xml.replace(old_cell, new_cell, 1)
        else:
            shared_path = "xl/sharedStrings.xml"
            if shared_path not in names:
                raise ValueError(f"No se encontró la columna {old_header}")
            shared_xml = archive.read(shared_path).decode("utf-8")
            shared_items = list(re.finditer(r"<si\b[^>]*>.*?</si>", shared_xml, re.DOTALL))
            old_index = next(
                (
                    index
                    for index, match in enumerate(shared_items)
                    if "".join(re.findall(r"<t\b[^>]*>(.*?)</t>", match.group(0), re.DOTALL))
                    == escape(old_header)
                ),
                None,
            )
            if old_index is None:
                raise ValueError(f"No se encontró la columna {old_header}")
            header_row = re.search(r'<row\b[^>]*\br="1"[^>]*>.*?</row>', sheet_xml, re.DOTALL)
            if not header_row:
                raise ValueError("No se encontró la fila de encabezados")
            cell_pattern = re.compile(
                rf'(<c\b[^>]*\bt="s"[^>]*>\s*<v>){old_index}(</v>\s*</c>)'
            )
            updated_header, substitutions = cell_pattern.subn(
                rf"\g<1>{len(shared_items)}\g<2>",
                header_row.group(0),
                count=1,
            )
            if not substitutions:
                raise ValueError(f"No se encontró la celda de encabezado {old_header}")
            sheet_xml = (
                sheet_xml[:header_row.start()]
                + updated_header
                + sheet_xml[header_row.end():]
            )
            shared_xml = shared_xml.replace(
                "</sst>",
                f"<si><t>{escape(new_header)}</t></si></sst>",
                1,
            )
            shared_xml = re.sub(
                r'(<sst\b[^>]*\buniqueCount=")(\d+)(")',
                lambda match: f'{match.group(1)}{int(match.group(2)) + 1}{match.group(3)}',
                shared_xml,
                count=1,
            )
            replacements[shared_path] = shared_xml.encode("utf-8")
        replacements[sheet_path] = sheet_xml.encode("utf-8")

        sheet_filename = sheet_path.rsplit("/", 1)[-1]
        relations_path = f"xl/worksheets/_rels/{sheet_filename}.rels"
        relation_xml = (
            archive.read(relations_path).decode("utf-8")
            if relations_path in names
            else ""
        )
        table_paths = {
            "xl/" + target.removeprefix("../")
            for target in re.findall(
                r'<Relationship\b[^>]*\bType="[^"]*/table"[^>]*\bTarget="([^"]+)"',
                relation_xml,
            )
        }
        for table_path in table_paths:
            table_xml = archive.read(table_path).decode("utf-8")
            table_xml = table_xml.replace(
                f'name="{escape(old_header)}"',
                f'name="{escape(new_header)}"',
                1,
            )
            replacements[table_path] = table_xml.encode("utf-8")

        output = io.BytesIO()
        with zipfile.ZipFile(output, "w") as destination:
            for item in archive.infolist():
                destination.writestr(item, replacements.get(item.filename, archive.read(item.filename)))
    path.write_bytes(repair_workbook_integrity(output.getvalue()))


def validate(path: Path, sheet_name: str, required_headers: set[str]) -> None:
    workbook = load_workbook(path, read_only=True, data_only=False)
    sheet = workbook[sheet_name]
    headers = {
        str(sheet.cell(1, column).value or "").strip()
        for column in range(1, sheet.max_column + 1)
    }
    missing = required_headers - headers
    if missing:
        raise ValueError(f"{path.name} is missing {sorted(missing)}")
    if sheet.max_row < 2:
        raise ValueError(f"{path.name} unexpectedly has no data rows")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--users", type=Path, required=True)
    parser.add_argument("--emision", type=Path, required=True)
    parser.add_argument("--siniestros", type=Path, required=True)
    args = parser.parse_args()

    migrate_users(args.users)
    migrate_pending_low_level(
        args.emision,
        18,
        sheet_path="xl/worksheets/sheet2.xml",
    )
    migrate_pending_low_level(args.siniestros, 15)
    validate(args.users, load_workbook(args.users, read_only=True).sheetnames[0], set(MODULE_COLUMNS))
    validate(args.emision, "Base1", {"Promotoria", "RFC Agente"})
    validate(args.siniestros, "Base", {"Promotoria", "RFC Agente"})


if __name__ == "__main__":
    main()
