import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table

from services.base_loads import (
    DEFAULT_AGENTS_METLIFE_FILE_ID,
    apply_prepared_workbook,
    build_preview,
    file_md5,
    file_sha256,
    replace_canonical_workbook,
    stage_agents_workbook,
)


HEADERS = [
    "CONTRATANTE", "RFC", "RAMSUBRAMO", "PRODUCTO", "NPOLIZA", "POLORIG",
    "FINIVIG", "FFINVIG", "NESQFPAGO", "NOMBREL", "ESTATUS", "CONDCOB",
    "PROMOTORIA", "AGENTE", "NOMBRE", "PRIMA", "PRIMA", "RECARGO",
    "GTOSEXP", "IVA", "MONEDA", "PAGADOHASTA", "DEDUCIBLE", "COASEGURO",
]


def business_row(policy, agent, end_date, client="CLIENTE", status=4):
    return [
        client, "RFC", 6001, "GMM", policy, policy, 20260101, end_date, 1,
        "ANUAL", status, "AGENTE", 119, agent, "AGENTE NOMBRE", 100, 100, 0,
        0, 16, "PESOS", 0, 10000, 10,
    ]


class BaseLoadsTest(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        root = Path(self.temp.name)
        self.agents = root / "agents.xlsx"
        self.incoming = root / "incoming.xlsx"
        self.canonical = root / "canonical.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"
        sheet.append(["CLAVE_DEFINITIVA", "Nombre"])
        sheet.append([73640, "Agente válido"])
        workbook.save(self.agents)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Reporte"
        sheet.append(HEADERS)
        sheet.append(business_row(100, 73640, 20270101, "ANTERIOR"))
        sheet.append(business_row(100, 73640, 20280101, "RECIENTE"))
        sheet.append(business_row(100, 73640, 20280101, "RECIENTE"))
        sheet.append(business_row(200, 99999, 20280101, "NO AUTORIZADO"))
        sheet.append(business_row(300, 73640, 20280201, "NUEVO"))
        workbook.save(self.incoming)

        workbook = Workbook()
        workbook.active.title = "Taiico OS Notes"
        sheet = workbook.create_sheet("GMM")
        full_headers = HEADERS + ["ESTATUS_DE_RENOVACION", "EXPEDIENTE", "Email"]
        sheet.append(full_headers)
        sheet.append(business_row(100, 73640, 20280101, "VIEJO") + ["En proceso", "drive://100", "cliente@example.com"])
        sheet.append(business_row(400, 73640, 20260101, "AUSENTE") + ["Comentario", "drive://400", None])
        sheet.add_table(Table(displayName="GMMTable", ref="A1:AA3"))
        workbook.save(self.canonical)

    def tearDown(self):
        self.temp.cleanup()

    def test_preview_reports_filter_dedup_and_preservation(self):
        preview = build_preview(self.incoming, self.canonical, self.agents)
        self.assertEqual(preview["source_rows"], 5)
        self.assertEqual(preview["rows_after_agent_filter"], 4)
        self.assertEqual(preview["duplicate_a_x_rows"], 1)
        self.assertEqual(preview["unique_incoming_policies"], 2)
        self.assertEqual(preview["unique_policy_periods"], 3)
        self.assertEqual(preview["existing_policies_updated"], 1)
        self.assertEqual(preview["new_policies_added"], 1)
        self.assertEqual(preview["current_policies_preserved_as_exceptions"], 1)
        self.assertEqual(preview["final_policy_count"], 3)
        self.assertEqual(preview["final_row_count"], 4)

    @patch("services.base_loads.download_drive_file_bytes")
    def test_agents_catalog_is_staged_from_canonical_drive_file(self, download):
        download.return_value = self.agents.read_bytes()
        destination = Path(self.temp.name) / "staged-agents.xlsx"

        result = stage_agents_workbook(destination)

        download.assert_called_once_with(DEFAULT_AGENTS_METLIFE_FILE_ID)
        self.assertEqual(result, destination)
        self.assertEqual(destination.read_bytes(), self.agents.read_bytes())

    @patch("services.base_loads.update_drive_workbook")
    @patch("services.base_loads.backup_drive_workbook")
    @patch("services.base_loads.drive_file_metadata")
    def test_apply_keeps_all_periods_and_preserves_matching_y_plus(
        self, drive_metadata, backup_drive, update_drive
    ):
        drive_metadata.return_value = {"md5Checksum": file_md5(self.canonical)}
        backup_drive.return_value = {
            "backup_file_id": "backup-id",
            "backup_name": "canonical antes de carga.xlsx",
            "backup_url": "https://drive.google.com/file/d/backup-id/view",
            "backup_folder_id": "history-folder-id",
        }
        update_drive.side_effect = lambda _file_id, path: {
            "drive_file_id": "canonical-drive-id",
            "drive_md5": file_md5(path),
        }
        result = replace_canonical_workbook(self.incoming, self.canonical, self.agents)
        backup_drive.assert_called_once()
        update_drive.assert_called_once()
        self.assertEqual(result["backup_file_id"], "backup-id")
        self.assertEqual(result["drive_file_id"], "canonical-drive-id")
        workbook = load_workbook(self.canonical, data_only=True)
        try:
            sheet = workbook["GMM"]
            rows = list(sheet.iter_rows(min_row=2, values_only=True))
            policy_100 = [row for row in rows if str(row[4]) == "100"]
            self.assertEqual(len(policy_100), 2)
            self.assertEqual({row[7] for row in policy_100}, {20270101, 20280101})
            recent = next(row for row in policy_100 if row[7] == 20280101)
            self.assertEqual(recent[0], "RECIENTE")
            old = next(row for row in policy_100 if row[7] == 20270101)
            self.assertEqual(old[24:], (None, None, None))
            policy_300 = next(row for row in rows if str(row[4]) == "300")
            self.assertEqual(policy_300[24:], (None, None, None))
            policy_400 = next(row for row in rows if str(row[4]) == "400")
            self.assertEqual(policy_400[0], "AUSENTE")
            self.assertEqual(sheet.tables["GMMTable"].ref, "A1:AA5")
        finally:
            workbook.close()

    @patch("services.base_loads.update_drive_workbook")
    @patch("services.base_loads.backup_drive_workbook")
    @patch("services.base_loads.drive_file_metadata")
    def test_apply_aborts_before_replacement_when_backup_fails(
        self, drive_metadata, backup_drive, update_drive
    ):
        original = self.canonical.read_bytes()
        drive_metadata.return_value = {"md5Checksum": file_md5(self.canonical)}
        backup_drive.side_effect = RuntimeError("Drive no disponible")

        with self.assertRaisesRegex(RuntimeError, "Drive no disponible"):
            replace_canonical_workbook(self.incoming, self.canonical, self.agents)

        self.assertEqual(self.canonical.read_bytes(), original)
        update_drive.assert_not_called()

    @patch("services.base_loads.backup_drive_workbook")
    @patch("services.base_loads.drive_file_metadata")
    def test_apply_refuses_to_overwrite_a_divergent_drive_file(
        self, drive_metadata, backup_drive
    ):
        candidate = Path(self.temp.name) / "prepared.xlsx"
        candidate.write_bytes(self.incoming.read_bytes())
        original = self.canonical.read_bytes()
        drive_metadata.return_value = {"md5Checksum": "different-drive-checksum"}

        with self.assertRaisesRegex(RuntimeError, "no coincide con la copia local"):
            apply_prepared_workbook(
                candidate,
                self.canonical,
                file_sha256(candidate),
                file_sha256(self.canonical),
            )

        backup_drive.assert_not_called()
        self.assertEqual(self.canonical.read_bytes(), original)

    @patch("services.base_loads.load_current_policies")
    @patch("services.base_loads.select_incoming_rows")
    @patch("services.base_loads.load_allowed_agent_keys")
    @patch("services.base_loads.load_workbook")
    @patch("services.base_loads.update_drive_workbook")
    @patch("services.base_loads.backup_drive_workbook")
    @patch("services.base_loads.drive_file_metadata")
    def test_apply_uses_prepared_workbook_without_reloading_sources(
        self,
        drive_metadata,
        backup_drive,
        update_drive,
        load_xlsx,
        load_agents,
        load_incoming,
        load_current,
    ):
        candidate = Path(self.temp.name) / "prepared.xlsx"
        candidate.write_bytes(self.incoming.read_bytes())
        drive_metadata.return_value = {"md5Checksum": file_md5(self.canonical)}
        backup_drive.return_value = {
            "backup_file_id": "backup-id",
            "backup_name": "canonical antes de carga.xlsx",
            "backup_url": "https://drive.google.com/file/d/backup-id/view",
            "backup_folder_id": "history-folder-id",
        }
        update_drive.side_effect = lambda _file_id, path: {
            "drive_file_id": "canonical-drive-id",
            "drive_md5": file_md5(path),
        }

        result = apply_prepared_workbook(
            candidate,
            self.canonical,
            file_sha256(candidate),
            file_sha256(self.canonical),
        )

        self.assertEqual(result["backup_file_id"], "backup-id")
        self.assertEqual(self.canonical.read_bytes(), candidate.read_bytes())
        load_xlsx.assert_not_called()
        load_agents.assert_not_called()
        load_incoming.assert_not_called()
        load_current.assert_not_called()
