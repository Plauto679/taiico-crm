import asyncio
import csv
import io
import sys
import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
from starlette.datastructures import UploadFile

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from database import Base, FinanceMovement
from services import finanzas


def canonical_csv(*rows):
    rows = rows or ({},)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=finanzas.CANONICAL_COLUMNS)
    writer.writeheader()
    for row in rows:
        base = {column: "" for column in finanzas.CANONICAL_COLUMNS}
        base.update({
            "id_movimiento": "mov-1", "empresa": "TLA", "banco": "AMEX",
            "tipo_cuenta": "Tarjeta", "naturaleza_cuenta": "Crédito", "moneda": "MXN",
            "fecha_operacion": "2026-08-01", "descripcion_original": "SERVICIO NUBE",
            "cargo": "100.00", "abono": "0", "importe_neto": "-100.00",
            "periodo_estado": "2026-08", "archivo_fuente": "estado.csv",
        })
        base.update(row)
        writer.writerow(base)
    return output.getvalue().encode("utf-8-sig")


def amex_monthly_csv():
    return (
        "Fecha,Fecha de Compra,Descripción,Titular de la Tarjeta,Cuenta,Importe\r\n"
        "27 Aug 2026,27 Aug 2026,PLAN DE PAGOS DIFERIDOS,ALBERTO ALFARO MENDOZA,-4006,13840.28\r\n"
        "14 Aug 2026,14 Aug 2026,GRACIAS POR SU PAGO EN BBVA,ALBERTO ALFARO MENDOZA,-4006,-140748.76\r\n"
        "28-jul-26,29-jul-26,CAFE SIRENA,PAMELA ASMARA ALFARO,-5011,746\r\n"
    ).encode("cp1252")


class FinanceModuleTests(unittest.TestCase):
    def setUp(self):
        engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
        Base.metadata.create_all(engine)
        self.Session = sessionmaker(bind=engine)
        self.temp = tempfile.TemporaryDirectory()
        self.source = Path(self.temp.name) / "Amex_historico.csv"
        self.session_patch = patch.object(finanzas, "SessionLocal", self.Session)
        self.paths_patch = patch.dict(finanzas.FINANCE_SOURCE_PATHS, {"tla_amex": self.source}, clear=True)
        self.file_ids_patch = patch.dict(finanzas.FINANCE_SOURCE_FILE_IDS, {"tla_amex": ""}, clear=True)
        self.runtime_patch = patch.object(finanzas, "RUNTIME_DIR", Path(self.temp.name) / "staging")
        self.session_patch.start(); self.paths_patch.start(); self.file_ids_patch.start(); self.runtime_patch.start()

    def tearDown(self):
        self.runtime_patch.stop(); self.file_ids_patch.stop(); self.paths_patch.stop(); self.session_patch.stop(); self.temp.cleanup()

    def test_parser_preserves_canonical_sign_convention(self):
        row = finanzas.parse_canonical_csv(canonical_csv())[0]
        self.assertEqual(row["operation_date"], date(2026, 8, 1))
        self.assertEqual(float(row["net_amount"]), -100.0)
        self.assertTrue(row["account_nature"], "Crédito")

    def test_parser_normalizes_american_express_bank_name(self):
        row = finanzas.parse_canonical_csv(canonical_csv({"banco": "AMERICAN EXPRESS"}))[0]
        self.assertEqual(row["bank"], "AMEX")

    def test_parser_rejects_duplicate_stable_ids(self):
        with self.assertRaisesRegex(ValueError, "duplicado"):
            finanzas.parse_canonical_csv(canonical_csv({}, {}))

    def test_monthly_amex_parser_enriches_original_export(self):
        rows = finanzas.parse_ingestion_csv("tla_amex", amex_monthly_csv(), filename="activity.csv")
        self.assertEqual(len(rows), 3)
        self.assertEqual(rows[0]["operation_date"], date(2026, 8, 27))
        self.assertEqual(rows[0]["account"], "-04006")
        self.assertEqual(rows[0]["debit"], finanzas.Decimal("13840.28"))
        self.assertEqual(rows[0]["net_amount"], finanzas.Decimal("-13840.28"))
        self.assertEqual(rows[1]["credit"], finanzas.Decimal("140748.76"))
        self.assertEqual(rows[1]["net_amount"], finanzas.Decimal("140748.76"))
        self.assertEqual(rows[2]["operation_date"], date(2026, 7, 28))
        self.assertEqual(rows[2]["settlement_date"], date(2026, 7, 29))
        self.assertEqual(rows[2]["account"], "-05011")
        self.assertEqual(rows[0]["source_filename"], "TLA/Estados Mensuales Amex/activity.csv")

    def test_monthly_amex_conversion_round_trips_as_canonical_csv(self):
        converted = finanzas.parse_ingestion_csv("tla_amex", amex_monthly_csv(), filename="activity.csv")
        reparsed = finanzas.parse_canonical_csv(finanzas.serialize_canonical_csv(converted))
        self.assertEqual(
            [(row["external_id"], row["operation_date"], row["net_amount"]) for row in reparsed],
            [(row["external_id"], row["operation_date"], row["net_amount"]) for row in converted],
        )

    def test_monthly_amex_parser_requires_original_six_columns(self):
        with self.assertRaisesRegex(ValueError, "Fecha de Compra"):
            finanzas.parse_ingestion_csv("tla_amex", b"Fecha,Importe\n27 Aug 2026,100\n")

    def test_preview_stages_monthly_amex_as_canonical_csv(self):
        upload = UploadFile(filename="activity.csv", file=io.BytesIO(amex_monthly_csv()))
        profile = finanzas.AccessProfile(
            username="admin@taiico.com", role="admin", promotorias=("TAIICO",),
            rfc="", aseguradoras=(), module_permissions={"finanzas": "operacion"},
        )
        preview = asyncio.run(finanzas.preview_ingestion("tla_amex", upload, profile))
        staged = (finanzas.RUNTIME_DIR / f"{preview['ingestion_id']}.csv").read_bytes()
        self.assertEqual(preview["rows"], 3)
        self.assertEqual(len(finanzas.parse_canonical_csv(staged)), 3)
        self.assertIn(b"id_movimiento,empresa,banco", staged)

    def test_publish_updates_drive_when_canonical_folder_is_not_mounted(self):
        upload = UploadFile(filename="activity.csv", file=io.BytesIO(amex_monthly_csv()))
        profile = finanzas.AccessProfile(
            username="admin@taiico.com", role="admin", promotorias=("TAIICO",),
            rfc="", aseguradoras=(), module_permissions={"finanzas": "operacion"},
        )
        preview = asyncio.run(finanzas.preview_ingestion("tla_amex", upload, profile))
        current = {"content": canonical_csv({"id_movimiento": "historical-1"})}

        def update_drive(file_id, content):
            self.assertEqual(file_id, "drive-amex")
            current["content"] = content

        finanzas.FINANCE_SOURCE_FILE_IDS["tla_amex"] = "drive-amex"
        with patch.object(finanzas, "download_drive_file_bytes", side_effect=lambda _file_id: current["content"]), patch.object(finanzas, "_update_drive_csv", side_effect=update_drive):
            result = finanzas.publish_ingestion(preview["ingestion_id"], profile)

        published = finanzas.parse_canonical_csv(current["content"])
        self.assertTrue(result["success"])
        self.assertEqual(result["published_rows"], 3)
        self.assertEqual(len(published), 4)
        self.assertEqual(published[0]["external_id"], "historical-1")

    def test_revert_restores_drive_backup(self):
        upload = UploadFile(filename="activity.csv", file=io.BytesIO(amex_monthly_csv()))
        profile = finanzas.AccessProfile(
            username="admin@taiico.com", role="admin", promotorias=("TAIICO",),
            rfc="", aseguradoras=(), module_permissions={"finanzas": "operacion"},
        )
        preview = asyncio.run(finanzas.preview_ingestion("tla_amex", upload, profile))
        original = canonical_csv({"id_movimiento": "historical-1"})
        current = {"content": original}
        finanzas.FINANCE_SOURCE_FILE_IDS["tla_amex"] = "drive-amex"

        def update_drive(_file_id, content):
            current["content"] = content

        with patch.object(finanzas, "download_drive_file_bytes", side_effect=lambda _file_id: current["content"]), patch.object(finanzas, "_update_drive_csv", side_effect=update_drive):
            finanzas.publish_ingestion(preview["ingestion_id"], profile)
            result = finanzas.revert_ingestion(preview["ingestion_id"], profile)

        self.assertTrue(result["success"])
        self.assertEqual(
            [row["external_id"] for row in finanzas.parse_canonical_csv(current["content"])],
            ["historical-1"],
        )

    def test_source_sync_is_idempotent_and_preserves_enrichment(self):
        self.source.write_bytes(canonical_csv())
        first = finanzas.sync_source("tla_amex")
        self.assertEqual(first["row_count"], 1)
        db = self.Session()
        movement = db.query(FinanceMovement).one()
        movement.category_override = "Software"; db.commit(); db.close()

        second = finanzas.sync_source("tla_amex", force=True)
        db = self.Session()
        rows = db.query(FinanceMovement).all()
        self.assertEqual(second["row_count"], 1)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].category_override, "Software")
        db.close()

    def test_source_sync_removes_rows_deleted_from_canonical_history(self):
        self.source.write_bytes(canonical_csv({}, {"id_movimiento": "mov-2"}))
        finanzas.sync_source("tla_amex")
        self.source.write_bytes(canonical_csv())
        finanzas.sync_source("tla_amex", force=True)
        db = self.Session()
        self.assertEqual([row.external_id for row in db.query(FinanceMovement).all()], ["mov-1"])
        db.close()

    def test_missing_source_is_reported_without_fake_data(self):
        result = finanzas.sync_source("tla_amex")
        self.assertFalse(result["available"])
        db = self.Session()
        self.assertEqual(db.query(FinanceMovement).count(), 0)
        db.close()

    def test_overview_and_movements_honor_bank_and_date_scope(self):
        self.source.write_bytes(canonical_csv(
            {"id_movimiento": "amex-aug", "banco": "AMEX", "fecha_operacion": "2026-08-01", "importe_neto": "100.00", "cargo": "0", "abono": "100.00"},
            {"id_movimiento": "bbva-aug", "banco": "BBVA", "fecha_operacion": "2026-08-15", "importe_neto": "200.00", "cargo": "0", "abono": "200.00"},
            {"id_movimiento": "amex-sep", "banco": "AMEX", "fecha_operacion": "2026-09-01", "importe_neto": "300.00", "cargo": "0", "abono": "300.00"},
        ))
        finanzas.sync_source("tla_amex")

        result = finanzas.overview("TLA", "AMEX", date(2026, 8, 1), date(2026, 8, 31))
        self.assertEqual(result["monthly"], [{"month": "2026-08", "entries": 100.0, "exits": 0.0, "net": 100.0}])
        self.assertEqual(result["kpis"]["entries_month"], 100.0)

        movements = finanzas.list_movements(
            company="TLA", search="", category="", bank="AMEX",
            start_date=date(2026, 8, 1), end_date=date(2026, 8, 31),
            page=1, page_size=5000, sort="operation_date", direction="desc",
        )
        self.assertEqual(movements["total"], 1)
        self.assertEqual(movements["items"][0]["id_movimiento"], "amex-aug")

    def test_amex_filter_includes_legacy_american_express_rows(self):
        db = self.Session()
        db.add(FinanceMovement(
            source_key="tla_amex", external_id="legacy-amex", company="TLA",
            bank="AMERICAN EXPRESS", currency="MXN", operation_date=date(2026, 6, 15),
            original_description="MOVIMIENTO HISTÓRICO", debit=0, credit=125,
            net_amount=125,
        ))
        db.commit()
        db.close()

        movements = finanzas.list_movements(
            company="TLA", search="", category="", bank="AMEX",
            start_date=date(2026, 1, 1), end_date=date(2026, 8, 27),
            page=1, page_size=5000, sort="operation_date", direction="desc",
        )
        self.assertEqual(movements["total"], 1)
        self.assertEqual(movements["items"][0]["banco"], "AMERICAN EXPRESS")

    def test_export_movements_returns_filtered_formatted_excel(self):
        self.source.write_bytes(canonical_csv(
            {"id_movimiento": "amex-export", "banco": "AMEX", "fecha_operacion": "2026-08-15", "importe_neto": "-100.00", "cargo": "100.00", "abono": "0"},
            {"id_movimiento": "bbva-hidden", "banco": "BBVA", "fecha_operacion": "2026-08-15"},
        ))
        finanzas.sync_source("tla_amex")

        response = finanzas.export_movements(
            company="TLA", search="amex-export", bank="AMEX",
            start_date=date(2026, 8, 1), end_date=date(2026, 8, 31),
        )

        self.assertEqual(response.media_type, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertIn(".xlsx", response.headers["content-disposition"])
        workbook = load_workbook(io.BytesIO(response.body), data_only=True)
        sheet = workbook["Movimientos"]
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet["A1"].value, "ID")
        self.assertEqual(sheet["A2"].value, "amex-export")
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet["D2"].value, datetime(2026, 8, 15))
        self.assertEqual(sheet["J2"].value, -100)
        self.assertEqual(sheet["J2"].number_format, '$#,##0.00;[Red]-$#,##0.00')
        self.assertEqual(sheet.auto_filter.ref, "A1:P2")

    def test_xml_invoice_extracts_cfdi_fields_without_claiming_validation(self):
        xml = b'''<cfdi:Comprobante xmlns:cfdi="http://www.sat.gob.mx/cfd/4" Fecha="2026-08-01T10:00:00" Total="116.00" Moneda="MXN"><cfdi:Emisor Rfc="AAA010101AAA"/><cfdi:Receptor Rfc="BBB010101BBB"/><cfdi:Complemento><tfd:TimbreFiscalDigital xmlns:tfd="http://www.sat.gob.mx/TimbreFiscalDigital" UUID="ABC-123"/></cfdi:Complemento></cfdi:Comprobante>'''
        path = Path(self.temp.name) / "factura.xml"; path.write_bytes(xml)
        result = finanzas._xml_invoice(path)
        self.assertEqual(result["uuid"], "ABC-123")
        self.assertEqual(result["issuer_rfc"], "AAA010101AAA")
        self.assertEqual(float(result["total"]), 116.0)


if __name__ == "__main__":
    unittest.main()
