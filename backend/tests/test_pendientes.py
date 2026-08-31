import io
import sys
import unittest
import zipfile
from datetime import date, datetime
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from services.pendientes import (
    EMISION_SERVICIOS_STATUS_OPTIONS,
    EMISION_SERVICIOS_PROCEDURE_OPTIONS,
    EmisionServiciosCreateRequest,
    GMM_REQUEST_OPTIONS,
    SiniestrosCreateRequest,
    SINIESTROS_STATUS_OPTIONS,
    SOURCES,
    VIDA_REQUEST_OPTIONS,
    PendingSource,
    _derived_day_values,
    _automatic_creation_values,
    _decorate_rows_with_folders,
    _folder_name_for_row,
    _filter_source_for_profile,
    _assigned_agent_rfc,
    _assigned_promotoria,
    _assigned_responsible,
    _notify_assigned_responsible,
    _insert_sheet_row,
    _normalize_spreadsheetml_namespace,
    add_pending_follow_up,
    append_pending_record,
    build_pending_report,
    build_pending_reminder_report,
    delete_pending_record,
    pending_report_html,
    pending_report_text,
    normalize_report_recipients,
    normalize_pending_amount,
    parse_pending_workbook,
    parse_agents_workbook,
    update_pending_record,
)
from services.auth import AccessProfile, PROMOTORIAS
from services.pending_document_requirements import (
    GMM_DOCUMENT_REQUIREMENTS,
    VIDA_DOCUMENT_REQUIREMENTS,
    requirements_for,
    split_request_types,
)


def workbook_bytes(sheet_name, headers, rows):
    output = io.BytesIO()
    with pd.ExcelWriter(output) as writer:
        pd.DataFrame(rows, columns=headers).to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
        )
    return output.getvalue()


def workbook_with_table(sheet_name, headers, rows):
    document = Workbook()
    sheet = document.active
    sheet.title = sheet_name
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    table = Table(displayName="Pendientes", ref=f"A1:{chr(64 + len(headers))}{len(rows) + 1}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)
    output = io.BytesIO()
    document.save(output)
    return output.getvalue()


class PendingWorkbookTests(unittest.TestCase):
    def test_non_contiguous_monto_column_is_core_and_not_history(self):
        source = PendingSource(
            "emision-servicios",
            "Emisión y Servicios",
            "TEST_ID",
            "file",
            "Base",
            2,
            ("Monto",),
        )
        original = workbook_bytes(
            "Base",
            ["Contratante", "Recordatorio Futuro", "Fecha Hoy", "31-ago-26", "Monto", "01-sep-26"],
            [["Cliente", "", "", "Primer avance", "1234.50", "Último avance"]],
        )

        parsed = parse_pending_workbook(original, source)

        self.assertEqual(parsed["core_headers"], ["Contratante", "Recordatorio Futuro", "Monto"])
        self.assertEqual(parsed["rows"][0]["summary"]["Monto"], "1234.50")
        self.assertEqual(
            parsed["rows"][0]["history"],
            [
                {"date": "31-ago-26", "update": "Primer avance"},
                {"date": "01-sep-26", "update": "Último avance"},
            ],
        )

        updated = update_pending_record(
            original,
            source,
            2,
            {"Monto": "2500.00"},
        )
        self.assertEqual(
            parse_pending_workbook(updated, source)["rows"][0]["summary"]["Monto"],
            "2500.00",
        )

    def test_pending_amount_normalizes_common_mexican_currency_inputs(self):
        self.assertEqual(normalize_pending_amount("$1,234.5"), "1234.50")
        self.assertEqual(normalize_pending_amount("(250)"), "-250.00")
        self.assertEqual(normalize_pending_amount(""), "")
        with self.assertRaisesRegex(ValueError, "cantidad válida"):
            normalize_pending_amount("mil pesos")

    def test_prefixed_spreadsheet_namespace_can_be_normalized_before_append(self):
        xml = (
            '<?xml version="1.0" encoding="utf-8"?>'
            '<s:worksheet xmlns:s="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<s:sheetData><s:row r="1"><s:c r="A1" t="inlineStr">'
            '<s:is><s:t>Contratante</s:t></s:is></s:c></s:row></s:sheetData>'
            '</s:worksheet>'
        )

        normalized = _normalize_spreadsheetml_namespace(xml)
        updated = _insert_sheet_row(normalized, 2, {1: "Alberto Alfaro"})

        self.assertIn('<sheetData>', updated)
        self.assertIn('<row r="2">', updated)
        self.assertIn('Alberto Alfaro', updated)
        self.assertNotIn('<s:sheetData>', updated)

    def test_pending_rows_are_scoped_by_promotoria_or_agent_rfc(self):
        source = {
            "source": "emision-servicios",
            "title": "Emisión y Servicios",
            "rows": [
                {"source_row": 2, "summary": {"Promotoria": "TAIICO", "RFC Agente": "RFC1"}},
                {"source_row": 3, "summary": {"Promotoria": "EKILIBRA", "RFC Agente": "RFC2"}},
                {"source_row": 4, "summary": {"Promotoria": "", "RFC Agente": ""}},
            ],
        }
        admin = AccessProfile(
            "admin@example.com",
            "admin",
            ("EKILIBRA",),
            "",
            (),
            {"pendientes": "operacion"},
        )
        agent = AccessProfile(
            "agent@example.com",
            "agente",
            ("TAIICO",),
            "RFC1",
            (),
            {"pendientes": "lectura"},
        )

        self.assertEqual(
            [row["source_row"] for row in _filter_source_for_profile(source, admin)["rows"]],
            [3],
        )
        self.assertEqual(
            [row["source_row"] for row in _filter_source_for_profile(source, agent)["rows"]],
            [2],
        )

    def test_single_promotoria_is_forced_when_admin_creates_pending(self):
        profile = AccessProfile(
            "admin@example.com",
            "admin",
            ("ABBONDANZA",),
            "",
            (),
            {"pendientes": "operacion"},
        )

        self.assertEqual(_assigned_promotoria("TAIICO", profile), "ABBONDANZA")

    def test_agents_workbook_builds_rfc_name_promotoria_labels(self):
        agents = parse_agents_workbook(
            workbook_bytes(
                "Datos",
                [
                    "RFC",
                    "Promotoria",
                    "Nombres",
                    "Apellido_Paterno",
                    "Apellido_Materno",
                ],
                [["aama950203i52", "Taiico", "ALBERTO", "ALFARO", "MENDOZA"]],
            ),
        )

        self.assertEqual(agents, [{
            "rfc": "AAMA950203I52",
            "name": "Alberto Alfaro Mendoza",
            "promotoria": "TAIICO",
            "label": "AAMA950203I52 - Alberto Alfaro Mendoza",
        }])

    def test_agent_assignment_must_match_selected_promotoria(self):
        profile = AccessProfile(
            "admin@example.com",
            "admin",
            ("TAIICO", "EKILIBRA"),
            "",
            (),
            {"pendientes": "operacion"},
        )
        agents = [
            {
                "rfc": "AAMA950203I52",
                "name": "Alberto Alfaro Mendoza",
                "promotoria": "TAIICO",
                "label": "AAMA950203I52 - Alberto Alfaro Mendoza",
            },
        ]

        with patch("services.pendientes.load_agent_directory", return_value=agents):
            self.assertEqual(
                _assigned_agent_rfc("aama950203i52", "TAIICO", profile),
                "AAMA950203I52",
            )
            with self.assertRaisesRegex(Exception, "promotoría asignada"):
                _assigned_agent_rfc("AAMA950203I52", "EKILIBRA", profile)

    def test_central_admin_sees_unassigned_records_and_inconsistencies(self):
        source = {
            "source": "siniestros",
            "title": "Siniestros",
            "rows": [{
                "source_row": 2,
                "summary": {
                    "ASEGURADO": "Cliente",
                    "Promotoria": "",
                    "RFC Agente": "",
                },
            }],
        }
        profile = AccessProfile(
            "central@example.com",
            "admin",
            PROMOTORIAS,
            "",
            (),
            {"pendientes": "operacion"},
        )

        filtered = _filter_source_for_profile(source, profile)

        self.assertEqual(len(filtered["rows"]), 1)
        self.assertEqual(len(filtered["inconsistencies"]), 1)
        self.assertIn(
            "Falta asignar promotoría",
            filtered["inconsistencies"][0]["problems"],
        )

    def test_summary_uses_core_columns_and_latest_update(self):
        source = PendingSource("test", "Test", "TEST_ID", "file", "Base", 2)
        result = parse_pending_workbook(
            workbook_bytes(
                "Base",
                ["Folio", "Cliente", "15-jul", "16-jul"],
                [["123", "Cliente Uno", "Primer avance", "Último avance"]],
            ),
            source,
        )
        row = result["rows"][0]
        self.assertEqual(row["summary"], {"Folio": "123", "Cliente": "Cliente Uno"})
        self.assertEqual(row["latest_update"], {"date": "16-jul", "update": "Último avance"})
        self.assertEqual(len(row["history"]), 2)

    def test_empty_history_cells_are_not_shown_in_detail(self):
        source = PendingSource("test", "Test", "TEST_ID", "file", "Base", 1)
        result = parse_pending_workbook(
            workbook_bytes("Base", ["Folio", "15-jul", "16-jul"], [["123", "", "Avance"]]),
            source,
        )
        self.assertEqual(
            result["rows"][0]["history"],
            [{"date": "16-jul", "update": "Avance"}],
        )

    def test_latest_update_is_last_non_empty_comment_for_each_row(self):
        source = PendingSource("test", "Test", "TEST_ID", "file", "Base", 1)
        result = parse_pending_workbook(
            workbook_bytes(
                "Base",
                ["Folio", "15-jul-26", "20-jul-26", "22-jul-26"],
                [
                    ["1", "En espera de aseguradora", "", ""],
                    ["2", "", "En espera de firma del cliente", ""],
                ],
            ),
            source,
        )
        self.assertEqual(result["rows"][0]["latest_update"], {"date": "15-jul-26", "update": "En espera de aseguradora"})
        self.assertEqual(result["rows"][1]["latest_update"], {"date": "20-jul-26", "update": "En espera de firma del cliente"})

    def test_follow_up_adds_one_date_column_and_updates_target_row(self):
        source = PendingSource("emision-servicios", "Test", "TEST_ID", "file", "Base", 1)
        original = workbook_bytes(
            "Base",
            ["Folio", "20-jul-26"],
            [["1", "Anterior"], ["2", ""]],
        )
        updated, header = add_pending_follow_up(
            original,
            source,
            3,
            "En espera de firma",
            date(2026, 7, 22),
        )
        parsed = parse_pending_workbook(updated, source)
        self.assertEqual(header, "22-jul-26")
        self.assertEqual(parsed["rows"][1]["latest_update"], {"date": header, "update": "En espera de firma"})

        with zipfile.ZipFile(io.BytesIO(original)) as before, zipfile.ZipFile(io.BytesIO(updated)) as after:
            changed = [name for name in before.namelist() if before.read(name) != after.read(name)]
        self.assertEqual(changed, ["xl/worksheets/sheet1.xml"])

    def test_second_follow_up_same_day_reuses_column(self):
        source = PendingSource("siniestros", "Test", "TEST_ID", "file", "Base", 1)
        original = workbook_bytes("Base", ["Folio", "21/07/2026"], [["1", ""]])
        first, header = add_pending_follow_up(original, source, 2, "Primer comentario", date(2026, 7, 22))
        second, second_header = add_pending_follow_up(first, source, 2, "Segundo comentario", date(2026, 7, 22))
        table = pd.read_excel(io.BytesIO(second), sheet_name="Base", dtype=str, keep_default_na=False)
        self.assertEqual(header, "22/07/2026")
        self.assertEqual(second_header, header)
        self.assertEqual(list(table.columns).count(header), 1)
        self.assertEqual(table.iloc[0][header], "Primer comentario | Segundo comentario")

    def test_follow_up_extends_native_excel_table(self):
        source = PendingSource("emision-servicios", "Test", "TEST_ID", "file", "Base", 1)
        original = workbook_with_table("Base", ["Folio", "21-jul-26"], [["1", "Anterior"]])
        updated, header = add_pending_follow_up(original, source, 2, "Nuevo", date(2026, 7, 22))
        with zipfile.ZipFile(io.BytesIO(updated)) as archive:
            table_xml = archive.read("xl/tables/table1.xml").decode()
        self.assertIn('ref="A1:C2"', table_xml)
        self.assertIn('count="3"', table_xml)
        self.assertIn(f'name="{header}"', table_xml)

    def test_append_pending_record_only_sets_requested_columns(self):
        source = PendingSource("test", "Test", "TEST_ID", "file", "Base", 3)
        original = workbook_bytes(
            "Base",
            ["Asegurado", "Póliza", "Casificacion", "16-jul"],
            [["Anterior", "123", "Vida", "Seguimiento"]],
        )
        updated, source_row = append_pending_record(original, source, {
            "Asegurado": "Nueva Persona",
            "Póliza": "456",
            "Casificacion": "GMM",
        })
        parsed = parse_pending_workbook(updated, source)
        self.assertEqual(source_row, 3)
        self.assertEqual(parsed["rows"][-1]["summary"], {
            "Asegurado": "Nueva Persona",
            "Póliza": "456",
            "Casificacion": "GMM",
        })
        self.assertEqual(parsed["rows"][-1]["latest_update"]["update"], "")

    def test_append_reuses_row_that_contains_only_excel_formatting(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Base"
        sheet.append(["Asegurado", "Póliza", "Casificacion", "16-jul"])
        sheet.append(["Anterior", "123", "Vida", "Seguimiento"])
        sheet.cell(row=3, column=1).number_format = "@"
        output = io.BytesIO()
        workbook.save(output)
        source = PendingSource("test", "Test", "TEST_ID", "file", "Base", 3)

        updated, source_row = append_pending_record(output.getvalue(), source, {
            "Asegurado": "Nueva Persona",
            "Póliza": "456",
            "Casificacion": "GMM",
        })

        self.assertEqual(source_row, 3)
        self.assertEqual(
            parse_pending_workbook(updated, source)["rows"][-1]["summary"]["Asegurado"],
            "Nueva Persona",
        )

    def test_append_preserves_physical_row_after_excel_sentinel(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Base"
        sheet.append(["Asegurado", "Póliza", "Casificacion", "16-jul"])
        sheet.append(["Anterior", "123", "Vida", "Seguimiento"])
        sheet.cell(row=3, column=1).number_format = "@"
        sheet.cell(row=1048575, column=1).number_format = "@"
        output = io.BytesIO()
        workbook.save(output)

        source = PendingSource("test", "Test", "TEST_ID", "file", "Base", 3)
        with zipfile.ZipFile(io.BytesIO(output.getvalue()), "r") as archive:
            sheet_path = "xl/worksheets/sheet1.xml"
            sheet_xml = archive.read(sheet_path).decode("utf-8")
            row_three_start = sheet_xml.index('<row r="3"')
            row_three_end = sheet_xml.index("</row>", row_three_start) + len("</row>")
            row_three = sheet_xml[row_three_start:row_three_end]
            sheet_xml = sheet_xml[:row_three_start] + sheet_xml[row_three_end:]
            sentinel_start = sheet_xml.index('<row r="1048575"')
            sheet_xml = sheet_xml[:sentinel_start] + row_three + sheet_xml[sentinel_start:]
            reordered = io.BytesIO()
            with zipfile.ZipFile(reordered, "w") as destination:
                for item in archive.infolist():
                    content = (
                        sheet_xml.encode("utf-8")
                        if item.filename == sheet_path
                        else archive.read(item.filename)
                    )
                    destination.writestr(item, content)

        updated, source_row = append_pending_record(reordered.getvalue(), source, {
            "Asegurado": "Nueva Persona",
            "Póliza": "456",
            "Casificacion": "GMM",
        })
        parsed = parse_pending_workbook(updated, source)

        self.assertEqual(source_row, 3)
        self.assertEqual(parsed["rows"][-1]["source_row"], 3)
        self.assertEqual(parsed["rows"][-1]["summary"]["Asegurado"], "Nueva Persona")

        followed_up, _ = add_pending_follow_up(
            updated,
            source,
            3,
            "Seguimiento en fila física",
            date(2026, 8, 19),
        )
        followed_up_row = next(
            row for row in parse_pending_workbook(followed_up, source)["rows"]
            if row["source_row"] == 3
        )
        self.assertEqual(
            followed_up_row["latest_update"]["update"],
            "Seguimiento en fila física",
        )

    def test_append_skips_row_with_data_outside_core_columns(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Base"
        sheet.append(["Asegurado", "Póliza", "Casificacion", "16-jul"])
        sheet.append(["Anterior", "123", "Vida", "Seguimiento"])
        sheet.cell(row=3, column=4, value="Comentario huérfano")
        output = io.BytesIO()
        workbook.save(output)
        source = PendingSource("test", "Test", "TEST_ID", "file", "Base", 3)

        updated, source_row = append_pending_record(output.getvalue(), source, {
            "Asegurado": "Nueva Persona",
            "Póliza": "456",
            "Casificacion": "GMM",
        })

        self.assertEqual(source_row, 4)
        self.assertEqual(
            parse_pending_workbook(updated, source)["rows"][-1]["summary"]["Asegurado"],
            "Nueva Persona",
        )

    def test_create_requests_allow_missing_rfc_and_emision_policy(self):
        emision = EmisionServiciosCreateRequest(
            asegurado="Cliente",
            casificacion="GMM",
            tipo_tramite="Servicios",
            solicitud_de="Rehabilitación GMM",
            promotoria="TAIICO",
            rfc_agente="AAMA950203I52",
        )
        siniestro = SiniestrosCreateRequest(
            asegurado="Cliente",
            poliza="123456",
            tipo_tramite="Reembolso",
            tramite="Complemento",
            promotoria="TAIICO",
            rfc_agente="AAMA950203I52",
        )
        self.assertEqual(emision.rfc, "")
        self.assertEqual(emision.poliza, "")
        self.assertEqual(siniestro.rfc, "")
        self.assertEqual(siniestro.poliza, "123456")

    def test_create_requests_require_promotoria_and_agent(self):
        with self.assertRaises(Exception):
            EmisionServiciosCreateRequest(
                asegurado="Cliente",
                casificacion="GMM",
                tipo_tramite="Servicios",
                solicitud_de="Rehabilitación GMM",
            )
        with self.assertRaises(Exception):
            SiniestrosCreateRequest(
                asegurado="Cliente",
                poliza="123456",
                tipo_tramite="Reembolso",
                tramite="Complemento",
            )

    def test_responsible_must_be_an_admin_user(self):
        profiles = (
            AccessProfile("admin@taiico.com", "admin", ("TAIICO",), "", ("*",), {}),
            AccessProfile("agent@taiico.com", "agente", ("TAIICO",), "RFC", ("*",), {}),
        )
        with patch("services.pendientes.list_access_profiles", return_value=profiles):
            self.assertEqual(_assigned_responsible("ADMIN@TAIICO.COM"), "admin@taiico.com")
            with self.assertRaises(Exception):
                _assigned_responsible("agent@taiico.com")

    def test_new_pending_notifies_assigned_responsible(self):
        row = {
            "summary": {
                "Responsable": "admin@taiico.com",
                "ASEGURADO": "Cliente Prueba",
                "Póliza": "123456",
                "Trámite": "Complemento",
            }
        }
        with (
            patch("services.pendientes.smtp_settings_for", return_value={"sender": "test"}),
            patch("services.pendientes.send_email_smtp") as send_email,
        ):
            warning = _notify_assigned_responsible(
                SOURCES["siniestros"],
                row,
                created_by="creator@taiico.com",
            )

        self.assertIsNone(warning)
        self.assertEqual(send_email.call_args.kwargs["recipients"], ["admin@taiico.com"])
        self.assertIn("Cliente Prueba", send_email.call_args.kwargs["subject"])
        self.assertIn("Póliza: 123456", send_email.call_args.kwargs["body"])

    def test_update_pending_record_updates_and_clears_core_values(self):
        source = PendingSource("test", "Test", "TEST_ID", "file", "Base", 3)
        original = workbook_bytes(
            "Base",
            ["Asegurado", "RFC", "Póliza", "22-jul"],
            [["Cliente", "", "123", "Registrado"]],
        )
        updated = update_pending_record(
            original,
            source,
            2,
            {"RFC": "AAMA950203I52", "Póliza": ""},
        )
        row = parse_pending_workbook(updated, source)["rows"][0]
        self.assertEqual(row["summary"]["RFC"], "AAMA950203I52")
        self.assertEqual(row["summary"]["Póliza"], "")
        self.assertEqual(row["latest_update"]["update"], "Registrado")

    def test_field_edits_are_added_to_update_history(self):
        source = PendingSource("siniestros", "Siniestros", "TEST_ID", "file", "Base", 3)
        original = workbook_bytes(
            "Base",
            ["Asegurado", "Estatus", "Comentarios", "22-jul-26"],
            [["Cliente", "Pendiente", "Nota anterior", "Seguimiento previo"]],
        )

        updated = update_pending_record(
            original,
            source,
            2,
            {"Estatus": "Pagado", "Comentarios": "Caso terminado"},
            history_values={"Estatus": "Pagado", "Comentarios": "Caso terminado"},
            history_date=date(2026, 8, 17),
        )
        row = parse_pending_workbook(updated, source)["rows"][0]

        self.assertEqual(row["summary"]["Estatus"], "Pagado")
        self.assertEqual(row["summary"]["Comentarios"], "Caso terminado")
        self.assertEqual(row["latest_update"]["date"], "17/08/2026")
        self.assertEqual(
            row["latest_update"]["update"],
            "Estatus: Pagado; Comentarios: Caso terminado",
        )

    def test_siniestros_status_update_accepts_only_current_catalog(self):
        source = PendingSource("siniestros", "Siniestros", "TEST_ID", "file", "Base", 2)
        original = workbook_bytes(
            "Base",
            ["ASEGURADO", "Estatus", "22-jul"],
            [["Cliente", "Estatus anterior", ""]],
        )
        for status in SINIESTROS_STATUS_OPTIONS:
            updated = update_pending_record(original, source, 2, {"Estatus": status})
            row = parse_pending_workbook(updated, source)["rows"][0]
            self.assertEqual(row["summary"]["Estatus"], status)
        with self.assertRaisesRegex(ValueError, "Estatus no válido"):
            update_pending_record(original, source, 2, {"Estatus": "Concluido"})

    def test_emision_status_update_accepts_only_current_catalog(self):
        source = PendingSource(
            "emision-servicios", "Emisión y Servicios", "TEST_ID", "file", "Base", 2
        )
        original = workbook_bytes(
            "Base",
            ["Asegurado", "Estatus actual", "22-jul"],
            [["Cliente", "EN ESPERA DE METLIFE", ""]],
        )

        for status in EMISION_SERVICIOS_STATUS_OPTIONS:
            updated = update_pending_record(
                original,
                source,
                2,
                {"Estatus actual": status},
            )
            row = parse_pending_workbook(updated, source)["rows"][0]
            self.assertEqual(row["summary"]["Estatus actual"], status)

        with self.assertRaisesRegex(ValueError, "Estatus actual no válido"):
            update_pending_record(
                original,
                source,
                2,
                {"Estatus actual": "EN ESPERA DE METLIFE"},
            )

    def test_emision_procedure_update_accepts_only_emision_and_servicios(self):
        source = PendingSource(
            "emision-servicios", "Emisión y Servicios", "TEST_ID", "file", "Base", 2
        )
        original = workbook_bytes(
            "Base",
            ["Asegurado", "Tipo de Trámite", "22-jul"],
            [["Cliente", "Trámite anterior", ""]],
        )

        for procedure in EMISION_SERVICIOS_PROCEDURE_OPTIONS:
            updated = update_pending_record(
                original,
                source,
                2,
                {"Tipo de Trámite": procedure},
            )
            row = parse_pending_workbook(updated, source)["rows"][0]
            self.assertEqual(row["summary"]["Tipo de Trámite"], procedure)

        with self.assertRaisesRegex(ValueError, "Tipo de trámite no válido"):
            update_pending_record(
                original,
                source,
                2,
                {"Tipo de Trámite": "Trámite anterior"},
            )

    def test_creation_dates_are_assigned_automatically_by_source(self):
        created_at = datetime(2026, 8, 10, 14, 35)
        self.assertEqual(
            _automatic_creation_values(SOURCES["emision-servicios"], created_at),
            {"Fecha Inicio": "2026-08-10 14:35"},
        )
        self.assertEqual(
            _automatic_creation_values(SOURCES["siniestros"], created_at),
            {"Fecha de registro de siniestro": "2026-08-10 14:35"},
        )

    def test_automatic_creation_dates_cannot_be_edited(self):
        cases = (
            (SOURCES["emision-servicios"], "Fecha Inicio"),
            (SOURCES["siniestros"], "Fecha de registro de siniestro"),
        )
        for source, date_header in cases:
            original = workbook_bytes(
                source.sheet_name,
                ["Asegurado", date_header, "22-jul"],
                [["Cliente", "2026-08-10", ""]],
            )
            with self.assertRaisesRegex(ValueError, "se asigna automáticamente"):
                update_pending_record(
                    original,
                    PendingSource(
                        source.key,
                        source.title,
                        source.file_id_env,
                        source.default_file_id,
                        source.sheet_name,
                        2,
                    ),
                    2,
                    {date_header: "2026-08-09"},
                )

    def test_delete_pending_record_removes_only_target_row_and_shrinks_table(self):
        source = PendingSource("test", "Test", "TEST_ID", "file", "Base", 3)
        original = workbook_with_table(
            "Base",
            ["Asegurado", "RFC", "Póliza", "22-jul"],
            [
                ["Primero", "RFC1", "101", "Seguimiento 1"],
                ["Eliminar", "RFC2", "102", "Seguimiento 2"],
                ["Tercero", "RFC3", "103", "Seguimiento 3"],
            ],
        )

        updated = delete_pending_record(original, source, 3)
        parsed = parse_pending_workbook(updated, source)

        self.assertEqual(
            [row["summary"]["Asegurado"] for row in parsed["rows"]],
            ["Primero", "Tercero"],
        )
        self.assertEqual(
            [row["source_row"] for row in parsed["rows"]],
            [2, 3],
        )
        with zipfile.ZipFile(io.BytesIO(updated)) as archive:
            self.assertIsNone(archive.testzip())
            table_xml = archive.read("xl/tables/table1.xml").decode()
        self.assertIn('ref="A1:D3"', table_xml)

    def test_assignment_update_preserves_untouched_comments(self):
        source = PendingSource("test", "Test", "TEST_ID", "file", "Base", 4)
        original = workbook_bytes(
            "Base",
            ["Asegurado", "Comentarios", "Promotoria", "RFC Agente", "22-jul"],
            [[
                "Cliente",
                "Pendiente de respuesta por parte de MetLife",
                "",
                "",
                "Seguimiento anterior",
            ]],
        )

        updated = update_pending_record(
            original,
            source,
            2,
            {"Promotoria": "TAIICO", "RFC Agente": "AAMA950203I52"},
        )
        row = parse_pending_workbook(updated, source)["rows"][0]

        self.assertEqual(
            row["summary"]["Comentarios"],
            "Pendiente de respuesta por parte de MetLife",
        )
        self.assertEqual(row["summary"]["Promotoria"], "TAIICO")
        self.assertEqual(row["summary"]["RFC Agente"], "AAMA950203I52")

    def test_folder_name_uses_pending_request_under_client_folder(self):
        self.assertEqual(
            _folder_name_for_row({
                "summary": {
                    "RFC": "aama950203i52",
                    "Solicitud de": "Rehabilitación póliza",
                    "Fecha Inicio": "2026-08-27 14:35",
                },
            }),
            "2026-08-27 14-35 Pendiente - Rehabilitación póliza",
        )

    def test_folder_names_distinguish_equal_claims_created_at_different_minutes(self):
        base_summary = {
            "RFC": "TLA180122DQ2",
            "Trámite": "Complemento",
        }
        first = _folder_name_for_row({
            "summary": {**base_summary, "Fecha de registro de siniestro": "2026-08-27 14:35"},
        })
        second = _folder_name_for_row({
            "summary": {**base_summary, "Fecha de registro de siniestro": "2026-08-27 14:36"},
        })
        self.assertEqual(first, "2026-08-27 14-35 Pendiente - Complemento")
        self.assertEqual(second, "2026-08-27 14-36 Pendiente - Complemento")
        self.assertNotEqual(first, second)

    def test_historical_folder_name_uses_date_without_inventing_a_time(self):
        self.assertEqual(
            _folder_name_for_row({
                "summary": {
                    "RFC": "TLA180122DQ2",
                    "Trámite": "Complemento",
                    "Fecha de registro de siniestro": "2026-08-27",
                },
            }),
            "2026-08-27 Pendiente - Complemento",
        )

    def test_one_legacy_folder_is_never_shared_by_two_pending_rows(self):
        class Request:
            def __init__(self, payload):
                self.payload = payload

            def execute(self):
                return self.payload

        class Files:
            def __init__(self):
                self.calls = 0

            def list(self, **_kwargs):
                self.calls += 1
                if self.calls == 1:
                    return Request({"files": [{"id": "client", "name": "TLA180122DQ2 - T&M2"}]})
                return Request({"files": [{
                    "id": "legacy",
                    "name": "Pendiente - Complemento",
                    "webViewLink": "https://drive.test/legacy",
                    "createdTime": "2026-08-24T10:00:00Z",
                }]})

        class Service:
            def __init__(self):
                self.api = Files()

            def files(self):
                return self.api

        result = {
            "rows": [
                {
                    "source_row": 21,
                    "summary": {
                        "RFC": "TLA180122DQ2",
                        "Trámite": "Complemento",
                        "Fecha de registro de siniestro": "2026-08-24",
                    },
                },
                {
                    "source_row": 32,
                    "summary": {
                        "RFC": "TLA180122DQ2",
                        "Trámite": "Complemento",
                        "Fecha de registro de siniestro": "2026-08-27",
                    },
                },
            ],
        }

        decorated = _decorate_rows_with_folders(result, Service())

        self.assertEqual(decorated["rows"][0]["folder_id"], "legacy")
        self.assertIsNone(decorated["rows"][1]["folder_id"])
        self.assertEqual(
            decorated["rows"][1]["folder_name"],
            "2026-08-27 Pendiente - Complemento",
        )

    def test_request_options_are_classification_specific(self):
        self.assertIn("Reembolso GMM", GMM_REQUEST_OPTIONS)
        self.assertNotIn("Reembolso GMM", VIDA_REQUEST_OPTIONS)
        self.assertIn("Rescate total / parcial VIDA", VIDA_REQUEST_OPTIONS)

    def test_rfc_is_preserved_as_core_column(self):
        source = PendingSource("test", "Test", "TEST_ID", "file", "Base", 3)
        result = parse_pending_workbook(
            workbook_bytes(
                "Base",
                ["Asegurado", "RFC", "Póliza", "22-jul"],
                [["Cliente", "AAMA950203I52", "123", "Registrado"]],
            ),
            source,
        )
        self.assertEqual(result["rows"][0]["summary"]["RFC"], "AAMA950203I52")

    def test_emision_new_date_columns_remain_core_and_days_are_current(self):
        source = PendingSource("emision-servicios", "Test", "TEST_ID", "file", "Base", 15)
        headers = [
            "Folio", "Asegurado", "RFC", "Póliza", "Producto", "Agente",
            "Casificacion", "Tipo de Trámite", "Solicitud de", "Estatus actual",
            "Fecha Inicio", "Días Transcurridos", "Fecha ingreso en la aseguradora",
            "Dias en la aseguradora", "Comentarios", "Responsable", "Fecha Hoy",
            "22-jul-26",
        ]
        row = [
            "1", "Cliente", "RFC1", "123", "GMM", "Pam", "GMM", "Servicios",
            "Rehabilitación GMM", "Ingresado", "2026-07-01", "0", "20/07/2026",
            "0", "Nota", "Pam", "", "Seguimiento",
        ]
        parsed = parse_pending_workbook(
            workbook_bytes("Base", headers, [row]),
            source,
            today=date(2026, 7, 23),
        )
        summary = parsed["rows"][0]["summary"]
        self.assertEqual(summary["Días Transcurridos"], "22")
        self.assertEqual(summary["Dias en la aseguradora"], "3")
        self.assertEqual(summary["Comentarios"], "Nota")
        self.assertNotIn("Fecha Hoy", summary)
        self.assertEqual(parsed["rows"][0]["latest_update"]["update"], "Seguimiento")

    def test_siniestros_new_date_columns_are_derived(self):
        source = PendingSource("siniestros", "Test", "TEST_ID", "file", "Base", 12)
        headers = [
            "Folio Titán", "ASEGURADO", "RFC", "Tipo de Trámite", "Trámite",
            "Estatus", "Comentarios", "Notas Especiales",
            "Fecha de registro de siniestro", "Dias desde registro del siniestro",
            "Fecha de envío a la aseguradora", "DIAS CUMPLIDOS EN LA ASEGURADORA",
            "Responsable", "Semaforo", "22/07/2026",
        ]
        row = [
            "1", "Cliente", "RFC1", "Reembolso", "Complemento", "Ingresado",
            "", "", "2026-07-01", "0", "2026-07-20", "0", "Pam", "Verde", "",
        ]
        parsed = parse_pending_workbook(
            workbook_bytes("Base", headers, [row]),
            source,
            today=date(2026, 7, 23),
        )
        summary = parsed["rows"][0]["summary"]
        self.assertEqual(summary["Dias desde registro del siniestro"], "22")
        self.assertEqual(summary["DIAS CUMPLIDOS EN LA ASEGURADORA"], "3")

    def test_derived_days_preserve_existing_counter_without_date_and_never_negative(self):
        self.assertEqual(
            _derived_day_values(
                {"Fecha Inicio": "", "Días Transcurridos": "99"},
                date(2026, 7, 23),
            ),
            {},
        )
        self.assertEqual(
            _derived_day_values(
                {"Fecha Inicio": "2026-07-25", "Días Transcurridos": ""},
                date(2026, 7, 23),
            ),
            {"Días Transcurridos": "0"},
        )

    def test_siniestros_preserve_existing_days_when_insurer_date_is_blank(self):
        source = PendingSource("siniestros", "Test", "TEST_ID", "file", "Base", 12)
        headers = [
            "Folio Titán", "ASEGURADO", "RFC", "Tipo de Trámite", "Trámite",
            "Estatus", "Comentarios", "Notas Especiales",
            "Fecha de registro de siniestro", "Dias desde registro del siniestro",
            "Fecha de envío a la aseguradora", "DIAS CUMPLIDOS EN LA ASEGURADORA",
            "Responsable", "Semaforo", "22/07/2026",
        ]
        row = [
            "1", "Cliente", "RFC1", "Reembolso", "Complemento", "Ingresado",
            "", "", "", "", "", "55", "Pam", "Rojo", "",
        ]

        parsed = parse_pending_workbook(
            workbook_bytes("Base", headers, [row]),
            source,
            today=date(2026, 7, 23),
        )

        summary = parsed["rows"][0]["summary"]
        self.assertEqual(summary["DIAS CUMPLIDOS EN LA ASEGURADORA"], "55")

    def test_report_classifies_each_counter_and_excludes_already_submitted_records(self):
        emision = {
            "rows": [
                {
                    "source_row": 2,
                    "summary": {
                        "Asegurado": "Previo Verde",
                        "RFC": "RFC1",
                        "Días Transcurridos": "5",
                        "Dias en la aseguradora": "",
                    },
                    "latest_update": {"date": "22-jul-26", "update": "Preparando"},
                },
                {
                    "source_row": 3,
                    "summary": {
                        "Asegurado": "Previo Amarillo",
                        "RFC": "RFC2",
                        "Días Transcurridos": "6",
                        "Dias en la aseguradora": "",
                    },
                    "latest_update": {},
                },
                {
                    "source_row": 4,
                    "summary": {
                        "Asegurado": "En aseguradora Rojo",
                        "RFC": "RFC3",
                        "Días Transcurridos": "20",
                        "Dias en la aseguradora": "11",
                    },
                    "latest_update": {},
                },
                {
                    "source_row": 5,
                    "summary": {
                        "Asegurado": "Sin fecha",
                        "Días Transcurridos": "",
                        "Dias en la aseguradora": "",
                    },
                    "latest_update": {},
                },
                {
                    "source_row": 60,
                    "summary": {
                        "Asegurado": "Emisión concluida",
                        "Estatus actual": "CONCLUIDO",
                        "Días Transcurridos": "40",
                        "Dias en la aseguradora": "",
                    },
                    "latest_update": {},
                },
            ],
        }
        siniestros = {
            "rows": [
                {
                    "source_row": 2,
                    "summary": {
                        "ASEGURADO": "Siniestro previo",
                        "Dias desde registro del siniestro": "10",
                        "DIAS CUMPLIDOS EN LA ASEGURADORA": "",
                    },
                    "latest_update": {},
                },
                {
                    "source_row": 3,
                    "summary": {
                        "ASEGURADO": "Siniestro aseguradora",
                        "Dias desde registro del siniestro": "12",
                        "DIAS CUMPLIDOS EN LA ASEGURADORA": "3",
                    },
                    "latest_update": {},
                },
                {
                    "source_row": 40,
                    "summary": {
                        "ASEGURADO": "Siniestro concluido",
                        "Estatus": "Concluido",
                        "Dias desde registro del siniestro": "40",
                        "DIAS CUMPLIDOS EN LA ASEGURADORA": "",
                    },
                    "latest_update": {},
                },
                {
                    "source_row": 41,
                    "summary": {
                        "ASEGURADO": "Siniestro pagado",
                        "Estatus": "Pagado",
                        "Dias desde registro del siniestro": "40",
                        "DIAS CUMPLIDOS EN LA ASEGURADORA": "",
                    },
                    "latest_update": {},
                },
                {
                    "source_row": 42,
                    "summary": {
                        "ASEGURADO": "Siniestro rechazado",
                        "Estatus": "Rechazado",
                        "Dias desde registro del siniestro": "40",
                        "DIAS CUMPLIDOS EN LA ASEGURADORA": "",
                    },
                    "latest_update": {},
                },
            ],
        }

        report = build_pending_report(emision, siniestros, date(2026, 7, 23))
        emision_before, emision_inside = report["sections"][0]["metrics"]
        siniestro_before, siniestro_inside = report["sections"][1]["metrics"]

        self.assertEqual(emision_before["counts"], {"verde": 1, "amarillo": 1, "rojo": 0})
        self.assertEqual(emision_inside["counts"], {"verde": 0, "amarillo": 0, "rojo": 1})
        self.assertEqual(siniestro_before["counts"], {"verde": 0, "amarillo": 1, "rojo": 0})
        self.assertEqual(siniestro_inside["counts"], {"verde": 1, "amarillo": 0, "rojo": 0})
        report_text = pending_report_text(report)
        self.assertNotIn("Emisión concluida", report_text)
        self.assertNotIn("Siniestro concluido", report_text)
        self.assertNotIn("Siniestro pagado", report_text)
        self.assertNotIn("Siniestro rechazado", report_text)
        self.assertFalse(any(item["source_row"] in {40, 41, 42, 60} for item in report["inconsistencies"]))

    def test_report_html_contains_summary_and_escapes_record_values(self):
        report = build_pending_report(
            {
                "rows": [{
                    "source_row": 2,
                    "summary": {
                        "Asegurado": "Cliente <Uno>",
                        "RFC": "RFC1",
                        "Solicitud de": "Rehabilitación GMM",
                        "Días Transcurridos": "3",
                        "Dias en la aseguradora": "",
                    },
                    "latest_update": {},
                }],
            },
            {"rows": []},
            date(2026, 7, 23),
        )
        html = pending_report_html(report)
        self.assertIn("Emisión y Servicios", html)
        self.assertIn("Verde (0-5)", html)
        self.assertIn("Cliente &lt;Uno&gt;", html)
        self.assertNotIn("Cliente <Uno>", html)

    def test_future_reminder_includes_next_15_days_regardless_of_status(self):
        emision = {
            "title": "Emisión y Servicios",
            "rows": [
                {"source_row": 2, "summary": {"Contratante": "Hoy", "Estatus actual": "Concluido", "Recordatorio Futuro": "2026-07-23"}},
                {"source_row": 3, "summary": {"Contratante": "Límite", "Recordatorio Futuro": "07/08/2026"}},
                {"source_row": 4, "summary": {"Contratante": "Fuera", "Recordatorio Futuro": "2026-08-08"}},
            ],
        }
        siniestros = {
            "title": "Siniestros",
            "rows": [
                {"source_row": 2, "summary": {"Contratante": "Pagado", "Estatus": "Pagado", "Recordatorio Futuro": "2026-07-30"}},
                {"source_row": 3, "summary": {"Contratante": "Anterior", "Recordatorio Futuro": "2026-07-22"}},
            ],
        }

        report = build_pending_reminder_report(emision, siniestros, date(2026, 7, 23))

        self.assertEqual(report["window_end"], "2026-08-07")
        self.assertEqual([item["contratante"] for item in report["items"]], ["Hoy", "Pagado", "Límite"])
        self.assertEqual(report["count"], 3)

    def test_report_recipients_accept_multiple_separators_and_remove_duplicates(self):
        self.assertEqual(
            normalize_report_recipients([
                "Pamela.Alfaro@taiico.com; veronica.alfaro@taiico.com",
                "pamela.alfaro@taiico.com\nalberto.alfaro@taiico.com",
            ]),
            [
                "pamela.alfaro@taiico.com",
                "veronica.alfaro@taiico.com",
                "alberto.alfaro@taiico.com",
            ],
        )

    def test_report_recipients_reject_invalid_or_empty_values(self):
        with self.assertRaises(ValueError):
            normalize_report_recipients(["no-es-correo"])
        with self.assertRaises(ValueError):
            normalize_report_recipients(["", "  "])

    def test_document_requirements_follow_classification_and_request(self):
        gmm = requirements_for("GMM", "Reembolso GMM")
        vida = requirements_for("Vida", "Rescate total / parcial VIDA")
        self.assertIn("Informe medico", gmm)
        self.assertIn("Estado de cuenta", vida)
        self.assertEqual(requirements_for("GMM", "Solicitud desconocida"), [])
        self.assertEqual(set(GMM_DOCUMENT_REQUIREMENTS), GMM_REQUEST_OPTIONS)
        self.assertEqual(set(VIDA_DOCUMENT_REQUIREMENTS), VIDA_REQUEST_OPTIONS)

    def test_siniestros_use_fixed_required_document_catalog(self):
        from services.pendientes import _requirements_for_row

        self.assertEqual(
            _requirements_for_row({"summary": {}}, "siniestros"),
            [
                "Identificación",
                "Comprobante de domicilio",
                "Informe Médico",
                "Facturas",
                "Finiquito",
            ],
        )

    def test_multiple_requests_combine_requirements_without_duplicates(self):
        selected = "Cambio de contratante GMM, Rehabilitación GMM & Cambio clave de agente"
        self.assertEqual(
            split_request_types(selected),
            [
                "Cambio de contratante GMM",
                "Rehabilitación GMM",
                "Cambio clave de agente",
            ],
        )
        documents = requirements_for("GMM", selected)
        self.assertEqual(documents.count("Solicitud de Cambios"), 1)
        self.assertIn("Cédula Fiscal", documents)
        self.assertIn("Formato de Rehabilitación", documents)
        self.assertIn("Carta Cliente", documents)


if __name__ == "__main__":
    unittest.main()
