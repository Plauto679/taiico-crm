import io
import sys
import unittest
from datetime import date
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from database import Base, Client, Insurer, Policy, Product, User
from services import cartera


def sura_workbook(*rows) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SURA"
    sheet.append(["Póliza", "Póliza actual", "Contratante", "Prospectador", "Porcentaje", "Inicio de pago"])
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


class CarteraModuleTests(unittest.TestCase):
    def setUp(self):
        engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
        Base.metadata.create_all(engine)
        self.Session = sessionmaker(bind=engine)
        db = self.Session()
        db.add_all([
            User(id="usr_pamela", name="Pamela", email="pamela@taiico.com", role="management"),
            Insurer(id="sura", name="SURA"),
            Product(id="prod_sura_gmm", insurer_id="sura", name="SURA GMM", branch="GMM"),
        ])
        db.commit()
        db.close()

    def test_parser_uses_last_repeated_policy_and_normalizes_columns(self):
        rows = cartera.parse_cartera_workbook(sura_workbook(
            (37, "", "", "PRIMERO", 0.5, None),
            (37, "37-A", "Cliente final", "CORRECTO", 0.8, date(2026, 9, 1)),
        ), "SURA")

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["policy_number"], "37")
        self.assertEqual(rows[0]["current_policy_number"], "37-A")
        self.assertEqual(rows[0]["prospector"], "CORRECTO")
        self.assertEqual(rows[0]["percentage"], Decimal("0.8"))
        self.assertEqual(rows[0]["payment_start_date"], "2026-09-01")

    def test_sync_creates_missing_records_and_updates_existing_without_erasing_name(self):
        db = self.Session()
        client = Client(full_name="Nombre existente", responsible_user_id="usr_pamela", metadata_json={})
        db.add(client)
        db.flush()
        db.add(Policy(
            policy_number="37", client_id=client.id, insurer_id="sura", product_id="prod_sura_gmm",
            effective_start_date=date(2026, 1, 1), effective_end_date=date(2027, 1, 1),
            premium_amount=0, payment_frequency="annual", responsible_user_id="usr_pamela",
            commission_percentage=0,
        ))
        db.commit()

        result = cartera.sync_cartera_source("sura", contents=sura_workbook(
            (37, 37, "", "ANA", 0.8, None),
            (38, 38, "", "LUIS", 0.5, None),
        ), db=db)

        self.assertEqual(result["updated"], 1)
        self.assertEqual(result["created"], 1)
        policies = {item.policy_number: item for item in db.query(Policy).filter_by(insurer_id="sura").all()}
        self.assertEqual(policies["37"].client.full_name, "Nombre existente")
        self.assertEqual(policies["37"].metadata_json["prospector"], "ANA")
        self.assertEqual(policies["38"].client.full_name, "SURA CLIENT P-38")
        db.close()

    def test_canonical_edit_updates_drive_and_keeps_percentage_as_ratio(self):
        original = sura_workbook((37, 37, "Cliente", "ANA", 0.5, None))
        uploaded = {}
        payload = cartera.CarteraRecordPayload(
            policy_number="37", current_policy_number="37", contractor="Cliente",
            prospector="ANA", percentage=80, payment_start_date=None,
            insurer="sura", policy_type="GMM",
        )
        with patch.object(cartera, "download_drive_file_bytes", return_value=original), \
             patch.object(cartera, "_upload_drive_workbook", side_effect=lambda file_id, contents: uploaded.update(file_id=file_id, contents=contents)), \
             patch.dict(cartera.CARTERA_SOURCE_FILE_IDS, {"sura": "drive-sura"}, clear=True), \
             patch.object(cartera, "SURA_PATHS", {"CARTERA": Path("/missing/local/Cartera SURA.xlsx")}):
            cartera._write_canonical(payload, "37")

        workbook = load_workbook(io.BytesIO(uploaded["contents"]), data_only=True)
        self.assertEqual(uploaded["file_id"], "drive-sura")
        self.assertEqual(workbook["SURA"]["E2"].value, 0.8)


if __name__ == "__main__":
    unittest.main()
