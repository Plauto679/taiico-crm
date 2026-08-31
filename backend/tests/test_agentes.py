from __future__ import annotations

import datetime as dt
import io
import sys
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from services.agentes import AgentFields, build_agent_directory, mutate_agent_workbook


HEADERS = [
    "CLAVE_ARRANQUE",
    "CLAVE_DEFINITIVA",
    "Nombres",
    "Nombre",
    "Apellido_Paterno",
    "Apellido_Materno",
    "Promotoria",
    "RFC",
    "Inicio_Vigencia_Cedula",
    "Fin_Vigencia_Cedula",
    "Clasificación Comercial",
    "Estatus_Met",
    "Campo_No_Editable",
]


def workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Datos"
    sheet.append(HEADERS)
    sheet.append([
        "00123",
        "00456",
        "ANA",
        "ANA LOPEZ DIAZ",
        "LOPEZ",
        "DIAZ",
        "TAIICO",
        "LODA900101AB1",
        dt.date(2026, 1, 15),
        dt.date(2029, 1, 15),
        "CONSOLIDADO",
        "ACTIVO",
        "CONSERVAR",
    ])
    sheet["I2"].number_format = "dd/mm/yyyy"
    sheet["J2"].number_format = "dd/mm/yyyy"
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


class AgentsDirectoryTests(unittest.TestCase):
    def test_directory_preserves_keys_and_serializes_dates(self):
        result = build_agent_directory(workbook_bytes(), can_operate=True)
        self.assertTrue(result["can_operate"])
        self.assertEqual(result["agents"][0]["clave_arranque"], "00123")
        self.assertEqual(result["agents"][0]["clave_definitiva"], "00456")
        self.assertEqual(result["agents"][0]["inicio_vigencia_cedula"], "2026-01-15")
        self.assertEqual(result["agents"][0]["fin_vigencia_cedula"], "2029-01-15")

    def test_update_preserves_unmanaged_columns(self):
        original = workbook_bytes()
        directory = build_agent_directory(original)
        agent = directory["agents"][0]
        payload = AgentFields(
            nombres="ANA MARIA",
            apellido_paterno="LOPEZ",
            apellido_materno="DIAZ",
            clave_arranque="00123",
            clave_definitiva="00456",
            promotoria="TAIICO",
            rfc="loda900101ab1",
            inicio_vigencia_cedula=dt.date(2026, 2, 20),
            fin_vigencia_cedula=dt.date(2029, 2, 20),
            clasificacion_comercial="ELITE",
            estatus_met="ACTIVO",
        )
        updated = mutate_agent_workbook(
            original,
            payload,
            row_number=agent["row_number"],
            fingerprint=agent["fingerprint"],
        )
        workbook = load_workbook(io.BytesIO(updated))
        sheet = workbook["Datos"]
        self.assertEqual(sheet["C2"].value, "ANA MARIA")
        self.assertEqual(sheet["D2"].value, "ANA MARIA LOPEZ DIAZ")
        self.assertEqual(sheet["H2"].value, "LODA900101AB1")
        self.assertEqual(sheet["M2"].value, "CONSERVAR")
        self.assertEqual(sheet["I2"].number_format, "dd/mm/yyyy")

    def test_append_adds_agent_without_changing_previous_row(self):
        payload = AgentFields(
            nombres="LUIS",
            apellido_paterno="PEREZ",
            apellido_materno="",
            clave_arranque="00999",
            clave_definitiva="",
            promotoria="CELAVI",
            rfc="PELJ800101AA1",
            inicio_vigencia_cedula=dt.date(2026, 3, 1),
            fin_vigencia_cedula=None,
            clasificacion_comercial="NUEVO",
            estatus_met="ACTIVO",
        )
        updated = mutate_agent_workbook(workbook_bytes(), payload)
        result = build_agent_directory(updated)
        self.assertEqual(len(result["agents"]), 2)
        added = next(agent for agent in result["agents"] if agent["rfc"] == "PELJ800101AA1")
        self.assertEqual(added["nombre"], "LUIS PEREZ")
        self.assertEqual(added["inicio_vigencia_cedula"], "2026-03-01")

    def test_duplicate_rfc_is_rejected(self):
        payload = AgentFields(
            nombres="OTRA",
            apellido_paterno="PERSONA",
            promotoria="TAIICO",
            rfc="LODA900101AB1",
        )
        with self.assertRaisesRegex(ValueError, "Ya existe otro agente con RFC"):
            mutate_agent_workbook(workbook_bytes(), payload)

    def test_stale_row_fingerprint_is_rejected(self):
        payload = AgentFields(nombres="ANA", promotoria="TAIICO")
        with self.assertRaisesRegex(RuntimeError, "cambió"):
            mutate_agent_workbook(
                workbook_bytes(),
                payload,
                row_number=2,
                fingerprint="0" * 64,
            )


if __name__ == "__main__":
    unittest.main()
